"""FastAPI 服务 v3 — 完整后端"""
import os, csv, hashlib, io, json, inspect, logging, math, time, uuid, threading, traceback
from datetime import datetime
from typing import Optional, Any
from fastapi import FastAPI, HTTPException, BackgroundTasks, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, PlainTextResponse, Response
from pydantic import BaseModel, Field
import openpyxl

from ..config import get_config, reload_config
from ..core.models import OpStatus, Operation, ShopFloor
from ..data.generator import InstanceGenerator
from ..core.rules import BUILTIN_RULES, get_all_rule_names
from ..core.sim_runtime import SimulationRuntime
from ..core.simulator import Simulator, SimResult
from ..ai.evolution import EvolutionEngine, EvolutionConfig, LLMInterface
from ..optimization.pareto import ParetoOptimizer, OBJECTIVES, NSGA2Optimizer
from ..optimization.objectives import OBJECTIVE_SPECS, build_schedule_analytics, objective_summary_payload
from ..optimization.hybrid_nsga3_alns import HybridConfig, HybridNSGA3ALNSOptimizer
from ..data.db import (
    init_db,
    get_instance_version,
    InstanceStore,
    GraphStore,
    DowntimeStore,
    WorkflowProgressStore,
    shifts_to_payload,
)
from ..data.graph_artifact_store import GraphArtifactStore
from ..data.template_builder import build_instance_template_bytes
from ..knowledge.canonical import compute_graph_fingerprint
from ..knowledge.context_service import (
    GraphContextMode,
    GraphContextService,
    resolve_graph_context_mode,
)
from ..optimization.exact import ExactSolver, EXACT_OBJECTIVES, exact_objective_catalog_payload
from ..scheduling.online import OnlineSchedulerV3
from ..core.time_utils import datetime_to_offset_hours

logging.basicConfig(level=logging.INFO)
app = FastAPI(title="LLM4DRD智能调度平台", version="3.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

FRONT = os.path.join(os.path.dirname(__file__), "..", "frontend")
if os.path.isdir(FRONT):
    app.mount("/static", StaticFiles(directory=FRONT), name="static")

shop: Optional[ShopFloor] = None
last_result = None
last_sim_payload = None
last_engine: Optional[EvolutionEngine] = None
inst_store = InstanceStore()
graph_store = GraphStore()
graph_artifact_store = GraphArtifactStore()
graph_context_service = GraphContextService(graph_artifact_store)
downtime_store = DowntimeStore()
workflow_store = WorkflowProgressStore()
_nsga2_tasks: dict = {}
_exact_tasks: dict = {}
_hybrid_tasks: dict = {}
_graph_tasks: dict = {}
_latest_hybrid_task_id: Optional[str] = None
OPTIMIZE_HEARTBEAT_INTERVAL_S = 1.0
online_scheduler_v3: Optional[OnlineSchedulerV3] = None
# 仿真在线程池里跑（def 端点），锁串行化对全局 shop/last_result/runtime 缓存的读写。
# runtime 构建要深拷贝整个实例（大实例十几秒），按内容键缓存：_active_shop() 每次
# 都从库重建 ShopFloor 新对象，对象身份不能当键。键 = 实例指纹（含班次/停机/交期，
# 见 knowledge.canonical）+ 指纹未覆盖的工序在制字段哈希。
_sim_lock = threading.Lock()
_sim_runtime_cache: Optional[tuple[str, SimulationRuntime]] = None
# (实例版本号, ShopFloor)，见 _active_shop()。
_active_shop_cache: Optional[tuple[int, ShopFloor]] = None


def _sim_runtime_cache_key(current_shop: ShopFloor) -> str:
    instance_hash = compute_graph_fingerprint(current_shop).instance_hash
    wip_fields = [
        (
            op_id,
            op.status.value,
            op.assigned_machine_id or "",
            tuple(op.assigned_tooling_ids),
            tuple(op.assigned_personnel_ids),
            op.start_time,
            op.end_time,
            op.remaining_processing_time,
        )
        for op_id, op in sorted(current_shop.operations.items())
        if op.status != OpStatus.PENDING or op.remaining_processing_time is not None
    ]
    wip_hash = hashlib.sha256(repr(wip_fields).encode()).hexdigest()
    return f"{instance_hash}|{wip_hash}"


def _cached_sim_runtime(current_shop: ShopFloor) -> SimulationRuntime:
    global _sim_runtime_cache
    try:
        key = _sim_runtime_cache_key(current_shop)
    except Exception:
        # 指纹计算失败（如数据含非有限值）不应阻断仿真，退化为不缓存。
        logging.exception("simulate: runtime cache key failed, building uncached runtime")
        return SimulationRuntime(current_shop)
    if _sim_runtime_cache is not None and _sim_runtime_cache[0] == key:
        return _sim_runtime_cache[1]
    runtime = SimulationRuntime(current_shop)
    _sim_runtime_cache = (key, runtime)
    return runtime


def _invalidate_graph_context(reason: str) -> None:
    graph_context_service.invalidate(reason)


def _build_graph_artifacts(
    current_shop,
    *,
    force_rebuild=False,
    progress_callback=None,
    deadline=None,
):
    return graph_context_service.get_or_build(
        current_shop,
        force_rebuild=force_rebuild,
        progress_callback=progress_callback,
        deadline=deadline,
    )


def _graph_context_diagnostics_payload(diagnostics) -> dict:
    fingerprint = diagnostics.fingerprint
    return {
        "cache_level": diagnostics.cache_level,
        "cache_hit": diagnostics.cache_hit,
        "instance_hash": fingerprint.instance_hash[:12],
        "topology_hash": fingerprint.topology_hash[:12],
        "feature_hash": fingerprint.feature_hash[:12],
        "schema_version": fingerprint.schema_version,
        "builder_version": fingerprint.builder_version,
        "load_time_ms": round(diagnostics.load_time_ms, 3),
        "build_time_ms": round(diagnostics.build_time_ms, 3),
        "validation_time_ms": round(diagnostics.validation_time_ms, 3),
        "operation_count": diagnostics.operation_count,
        "relation_count": diagnostics.relation_count,
        "invalid_reason": diagnostics.invalid_reason,
    }


def _save_workflow_progress(step: str, payload: dict) -> None:
    """存流程进度快照；失败只记日志。

    进度持久化是为了重启后能恢复，不是计算结果的一部分：写库出问题（磁盘满、
    payload 含无法序列化的值）不该把一次刚跑完的仿真或优化判成失败。
    """
    try:
        workflow_store.save(step, payload)
    except Exception:  # noqa: BLE001
        logging.exception("保存流程进度失败: step=%s", step)


def _restore_workflow_progress() -> None:
    """把上次进程存下的仿真/优化结果读回内存，让重启后的导出、评审接口继续可用。

    实例版本对不上的快照会被 WorkflowProgressStore 直接过滤掉，这里拿到的一定
    是与当前库内实例匹配的结果。校验和评审快照没有内存态，由前端直接取。
    """
    global last_sim_payload, _latest_hybrid_task_id
    try:
        snapshot = workflow_store.load_all()
    except Exception:  # noqa: BLE001
        logging.exception("读取流程进度失败，按无进度启动")
        return

    simulation = snapshot.get("simulation")
    if simulation:
        last_sim_payload = simulation

    optimization = snapshot.get("optimization") or {}
    task_id = optimization.get("task_id")
    task = optimization.get("task")
    if task_id and task:
        _hybrid_tasks[task_id] = task
        _latest_hybrid_task_id = task_id

    logging.info(
        "workflow progress restored: steps=%s", sorted(snapshot.keys()) or "none",
    )


@app.on_event("startup")
async def startup():
    init_db()
    _restore_workflow_progress()

@app.get("/")
async def index():
    # Keep one supported HTML/JavaScript contract for the application shell.
    f = os.path.join(FRONT, "index_v2.html")
    return FileResponse(f) if os.path.exists(f) else {"msg": "API running"}

@app.get("/v2")
async def index_v2():
    f = os.path.join(FRONT, "index_v2.html")
    return FileResponse(f) if os.path.exists(f) else {"msg": "V2 frontend not found"}

# === Models ===
class GenReq(BaseModel):
    num_orders: int = 10; tasks_per_order_min: int = 2; tasks_per_order_max: int = 5
    ops_per_task_min: int = 2; ops_per_task_max: int = 5
    machines_per_type: int = 3; due_date_factor: float = 1.5
    arrival_spread: float = 0.0; seed: Optional[int] = 42
    day_shift_hours: float = 10; night_shift_hours: float = 8
    schedule_days: int = 0; maintenance_prob: float = 0.05
    toolings_per_type: Optional[int] = None
    personnel_per_skill: Optional[int] = None
    plan_start_at: Optional[datetime] = None

class SimReq(BaseModel):
    rule_name: str = "ATC"

class ParetoReq(BaseModel):
    objectives: list[str] = ["total_tardiness", "makespan"]
    rule_names: Optional[list[str]] = None

class TrainReq(BaseModel):
    population_size: int = 6; max_generations: int = 5
    num_train_instances: int = 3; seed_rules: Optional[list[str]] = None

class LLMCfg(BaseModel):
    base_url: Optional[str] = None; api_key: Optional[str] = None
    model: Optional[str] = None

class DowntimeReq(BaseModel):
    machine_id: str
    downtime_type: str = "planned"
    start_time: float | datetime
    end_time: float | datetime

class NSGA2Req(BaseModel):
    objectives: list[str] = ["total_tardiness", "makespan"]
    pop_size: int = 20
    generations: int = 10
    seed: int = 42

class ExactReq(BaseModel):
    objectives: list[str] = ["makespan", "total_tardiness"]
    time_limit_s: int = 60

class HybridOptimizeReq(BaseModel):
    objective_keys: list[str] = Field(default_factory=lambda: ["total_tardiness", "makespan"])
    target_solution_count: int = 12
    time_limit_s: int = 90
    population_size: int = 24
    generations: int = 12
    alns_iterations_per_candidate: int = 6
    candidate_filter_multiplier: int = 3
    coarse_pool_multiplier: int = 4
    elite_refine_ratio: float = 0.4
    elite_refine_min: int = 4
    coarse_time_ratio: float = 0.68
    promotion_pool_multiplier: int = 3
    random_promotion_ratio: float = 0.12
    refine_rounds: int = 1
    alns_aggression: float = 1.0
    stagnation_generations: int = 3
    parallel_workers: int = 0
    seed: int = 42
    baseline_rule_name: str = "ATC"

class OnlineStartReq(BaseModel):
    rule_name: str = "ATC"

class OnlineAdvanceReq(BaseModel):
    delta: float = 1.0

class OnlineBreakdownReq(BaseModel):
    machine_id: str
    repair_at: float | datetime

class OnlineRepairReq(BaseModel):
    machine_id: str

class OnlineRescheduleReq(BaseModel):
    rule_name: Optional[str] = None

class ScenarioCompareReq(BaseModel):
    rule_names: Optional[list[str]] = None
    num_replications: int = 1
    breakdown_prob: float = 0.0
    pt_variance: float = 0.0

class NLScheduleReq(BaseModel):
    prompt: str
    current_rule: Optional[str] = "ATC"

class NLAnalysisReq(BaseModel):
    question: str


class ParetoCompareReq(BaseModel):
    task_id: Optional[str] = None
    solution_ids: list[str] = Field(default_factory=list)
    heuristic_rule_names: list[str] = Field(default_factory=list)
    requirement: Optional[str] = None
    conversation: list[dict[str, str]] = Field(default_factory=list)


class ParetoRecommendReq(BaseModel):
    task_id: Optional[str] = None
    solution_ids: list[str] = Field(default_factory=list)
    heuristic_rule_names: list[str] = Field(default_factory=list)
    requirement: str
    conversation: list[dict[str, str]] = Field(default_factory=list)


class ParetoAskReq(BaseModel):
    task_id: Optional[str] = None
    solution_id: str
    question: str
    heuristic_rule_names: list[str] = Field(default_factory=list)
    conversation: list[dict[str, str]] = Field(default_factory=list)


class HeuristicReferenceReq(BaseModel):
    rule_names: list[str] = Field(default_factory=list)
    objective_keys: list[str] = Field(default_factory=list)


class ExactReferenceReq(BaseModel):
    task_id: Optional[str] = None
    mode: str = "single"
    objective_key: Optional[str] = "makespan"
    objective_weights: dict[str, float] = Field(default_factory=dict)
    time_limit_s: int = 60


class ExportSolutionReq(BaseModel):
    task_id: Optional[str] = None
    solution_id: str


class ReviewProgressReq(BaseModel):
    selection: list[str] = Field(default_factory=list)
    detail_id: Optional[str] = None
    ai_recommended_id: Optional[str] = None


def _plan_start_ref():
    if shop is not None:
        return shop.plan_start_at
    if inst_store.has_data():
        return inst_store.get_plan_start_at()
    return datetime.now().astimezone()


def _coerce_offset(value):
    return datetime_to_offset_hours(_plan_start_ref(), value)


def _coerce_excel_offset(plan_start_at, value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return datetime_to_offset_hours(plan_start_at, text) or 0.0


def _time_payload(current_shop: ShopFloor, value: float | None, key: str) -> dict:
    return {
        key: round(value, 3) if value is not None else None,
        f"{key}_at": current_shop.time_label(value) if value is not None else None,
    }


def _ensure_shop_calendar_capacity(current_shop: ShopFloor) -> dict:
    return current_shop.ensure_calendar_capacity(
        min_days=max(current_shop.calendar_days(), 14),
        safety_factor=1.45,
        max_days=720,
    )


def _serialize_schedule_entry(current_shop: ShopFloor, entry: dict) -> dict:
    payload = dict(entry)
    start_value = entry.get("start")
    end_value = entry.get("end")
    # 仿真不可行时，部分工序的 start/end 可能是 inf（资源日历容量不足等），
    # 这里做有限性校验，避免 time_label -> timedelta(hours=inf) 抛 OverflowError，
    # 同时把 inf 数值也置为 None，防止 JSON 输出 Infinity 污染前端。
    if start_value is not None and math.isfinite(float(start_value)):
        payload["start_at"] = current_shop.time_label(start_value)
    else:
        payload["start"] = None
        payload["start_at"] = None
    if end_value is not None and math.isfinite(float(end_value)):
        payload["end_at"] = current_shop.time_label(end_value)
    else:
        payload["end"] = None
        payload["end_at"] = None
    return payload


def _serialize_downtime_row(current_shop: ShopFloor, row: dict) -> dict:
    payload = dict(row)
    payload["start_at"] = current_shop.time_label(row["start_time"])
    payload["end_at"] = current_shop.time_label(row["end_time"])
    return payload


def _shift_payload(current_shop: ShopFloor, shift) -> dict:
    start = shift.day * 24.0 + shift.start_hour
    end = start + shift.hours
    return {
        "day": shift.day,
        "start_hour": shift.start_hour,
        "hours": shift.hours,
        "start": round(start, 3),
        "end": round(end, 3),
        "start_at": current_shop.time_label(start),
        "end_at": current_shop.time_label(end),
    }


def _resource_calendar_payload(current_shop: ShopFloor, resource) -> dict:
    return {
        "shifts": [_shift_payload(current_shop, shift) for shift in getattr(resource, "shifts", [])],
        "downtimes": [
            {
                "id": downtime.id,
                "machine_id": downtime.machine_id,
                "downtime_type": downtime.downtime_type,
                "start": round(downtime.start_time, 3),
                "end": round(downtime.end_time, 3),
                "start_at": current_shop.time_label(downtime.start_time),
                "end_at": current_shop.time_label(downtime.end_time),
            }
            for downtime in getattr(resource, "downtimes", [])
        ],
    }


def _normalize_initial_status(value: str | None) -> str:
    status = str(value or "").strip().lower()
    return status if status in {"pending", "ready", "processing", "completed"} else ""


def _has_initial_wip(current_shop: ShopFloor) -> bool:
    return any(op.status.value in {"ready", "processing", "completed"} for op in current_shop.operations.values())


def _rule_reference_solution(
    current_shop: ShopFloor,
    rule_name: str,
    objective_keys: list[str] | None = None,
    schedule_limit: int | None = 120,
) -> dict:
    if rule_name not in BUILTIN_RULES:
        raise HTTPException(400, f"未知规则: {rule_name}")
    sim = Simulator(current_shop, BUILTIN_RULES[rule_name])
    result = sim.run()
    analytics = build_schedule_analytics(current_shop, result)
    keys = objective_keys or list(OBJECTIVE_SPECS.keys())
    objective_payload = {
        key: round(float(analytics.objective_values.get(key, result.to_dict().get(key, 0.0))), 4)
        for key in keys
        if key in OBJECTIVE_SPECS
    }
    metrics = result.to_dict()
    summary = {
        "completed_operations": analytics.summary.get("completed_operations", analytics.completed_operations),
        "total_operations": analytics.summary.get("total_operations", len(current_shop.operations)),
        "tardy_order_ids": analytics.tardy_order_ids,
        "tardy_task_ids": analytics.tardy_task_ids,
        "bottleneck_machine_ids": analytics.bottleneck_machine_ids,
        "avg_utilization": round(float(analytics.objective_values.get("avg_utilization", metrics.get("avg_utilization", 0.0))), 4),
        "critical_utilization": round(float(analytics.objective_values.get("critical_utilization", metrics.get("critical_utilization", 0.0))), 4),
        "avg_active_window_utilization": round(float(analytics.objective_values.get("avg_active_window_utilization", 0.0)), 4),
        "critical_active_window_utilization": round(float(analytics.objective_values.get("critical_active_window_utilization", 0.0)), 4),
        "avg_net_available_utilization": round(float(analytics.objective_values.get("avg_net_available_utilization", 0.0)), 4),
        "critical_net_available_utilization": round(float(analytics.objective_values.get("critical_net_available_utilization", 0.0)), 4),
        "tooling_utilization": round(float(analytics.objective_values.get("tooling_utilization", 0.0)), 4),
        "personnel_utilization": round(float(analytics.objective_values.get("personnel_utilization", 0.0)), 4),
        "evaluation_mode": "exact",
    }
    schedule = []
    schedule_entries = list(result.schedule or [])
    if schedule_limit is not None:
        schedule_entries = schedule_entries[:schedule_limit]
    for entry in schedule_entries:
        task = current_shop.tasks.get(entry["task_id"])
        order = current_shop.orders.get(task.order_id) if task else None
        schedule.append(
            _serialize_schedule_entry(
                current_shop,
                {
                    **entry,
                    "order_id": order.id if order else "",
                    "order_name": order.name if order else "",
                    "priority": order.priority if order else 1,
                    "due_date": round(order.due_date, 3) if order else 0,
                    "due_at": current_shop.time_label(order.due_date) if order else None,
                    "is_tardy": (entry["end"] > order.due_date) if order else False,
                    "is_main": task.is_main if task else False,
                },
            )
        )
    return {
        "solution_id": f"RULE:{rule_name}",
        "rule_name": rule_name,
        "source": "heuristic_rule",
        "generation": 0,
        "rank": None,
        "feasible": bool(analytics.feasible),
        "evaluation_mode": "exact",
        "objectives": objective_payload,
        "metrics": metrics,
        "summary": summary,
        "candidate": {
            "seed_rule_name": rule_name,
            "graph_profile": "dispatch_rule_reference",
            "destroy_fraction": 0.0,
            "feature_weights": {},
            "bias_count": 0,
        },
        "schedule": schedule,
    }


def _task_reference_solution_index(task: dict) -> dict[str, dict]:
    return {
        item.get("solution_id"): item
        for item in task.get("reference_solutions", []) or []
        if item.get("solution_id")
    }


def _requested_exact_objective_keys(base_objective_keys: list[str], req: ExactReferenceReq) -> list[str]:
    keys = list(base_objective_keys or [])
    requested = []
    if (req.mode or "single").lower() == "single":
        if req.objective_key:
            requested.append(req.objective_key)
    else:
        requested.extend((req.objective_weights or {}).keys())
    for key in requested:
        if key and key not in keys:
            keys.append(key)
    return keys


def _ensure_solution_objectives(current_shop: ShopFloor, solution: dict, objective_keys: list[str]) -> dict:
    existing = dict(solution.get("objectives") or {})
    missing = [key for key in objective_keys if key in OBJECTIVE_SPECS and key not in existing]
    if not missing or not solution.get("schedule"):
        return existing
    analytics = build_schedule_analytics(current_shop, SimResult(schedule=solution["schedule"]))
    for key in missing:
        existing[key] = round(float(analytics.objective_values.get(key, 0.0)), 4)
    solution["objectives"] = existing
    return existing


def _build_exact_reference_fallback(
    current_shop: ShopFloor,
    task: dict,
    objective_keys: list[str],
    req: ExactReferenceReq,
    schedule_limit: int | None,
    reason: str,
) -> dict:
    export_result = task.get("export_result") or task.get("result") or {}
    candidates: dict[str, dict] = {}
    baseline = export_result.get("baseline")
    if baseline and baseline.get("solution_id"):
        candidates[baseline["solution_id"]] = baseline
    for item in export_result.get("solutions", []) or []:
        if item.get("solution_id"):
            candidates[item["solution_id"]] = item
    for item in export_result.get("reference_solutions", []) or []:
        if item.get("solution_id") and not str(item.get("source", "")).startswith("exact_reference"):
            candidates[item["solution_id"]] = item
    for rule_name in BUILTIN_RULES:
        rule_solution = _rule_reference_solution(current_shop, rule_name, objective_keys, schedule_limit=None)
        candidates[rule_solution["solution_id"]] = rule_solution

    if not candidates:
        raise HTTPException(400, f"精确求解未得到可用方案，且没有可回退的候选: {reason}")

    requested_mode = (req.mode or "single").lower()
    if requested_mode == "single":
        objective_key = req.objective_key or "makespan"
        spec = EXACT_OBJECTIVES.get(objective_key)
        if spec is None:
            raise HTTPException(400, f"不支持的精确目标: {objective_key}")

        def single_score(item: dict):
            objectives = _ensure_solution_objectives(current_shop, item, [objective_key])
            return float(objectives.get(objective_key, 0.0))

        reverse = spec.direction == "max"
        winner = sorted(candidates.values(), key=single_score, reverse=reverse)[0]
        request_payload = {
            "mode": "single",
            "objective_key": objective_key,
            "objective_weights": {objective_key: 1.0},
            "support": {objective_key: "fallback_exact_simulation"},
        }
    else:
        weights = {
            key: abs(float(value))
            for key, value in (req.objective_weights or {}).items()
            if key in EXACT_OBJECTIVES and abs(float(value)) > 1e-9
        }
        total = sum(weights.values())
        if total <= 0:
            raise HTTPException(400, "加权精确方案至少需要提供一个非零权重目标")
        normalized = {key: value / total for key, value in weights.items()}
        requested_keys = list(normalized.keys())
        ranges = {}
        for key in requested_keys:
            values = [float(_ensure_solution_objectives(current_shop, item, [key]).get(key, 0.0)) for item in candidates.values()]
            ranges[key] = (min(values), max(values))

        def weighted_score(item: dict):
            objectives = _ensure_solution_objectives(current_shop, item, requested_keys)
            score = 0.0
            for key, weight in normalized.items():
                raw = float(objectives.get(key, 0.0))
                lower, upper = ranges[key]
                span = upper - lower
                if span <= 1e-9:
                    normalized_value = 0.0
                else:
                    normalized_value = (raw - lower) / span
                if EXACT_OBJECTIVES[key].direction == "max":
                    normalized_value = 1.0 - normalized_value
                score += weight * normalized_value
            item.setdefault("objectives", {})["weighted_score"] = round(score, 6)
            return score

        winner = sorted(candidates.values(), key=weighted_score)[0]
        request_payload = {
            "mode": "weighted",
            "objective_key": None,
            "objective_weights": normalized,
            "support": {key: "fallback_exact_simulation" for key in normalized},
        }

    chosen = dict(winner)
    schedule_entries = list(chosen.get("schedule") or [])
    if schedule_limit is not None:
        chosen["schedule"] = schedule_entries[:schedule_limit]
    analytics = build_schedule_analytics(current_shop, SimResult(schedule=schedule_entries))
    objective_payload = {
        key: round(float(analytics.objective_values.get(key, 0.0)), 4)
        for key in objective_keys
        if key in OBJECTIVE_SPECS
    }
    chosen["solution_id"] = (
        f"EXACT:SINGLE:{request_payload.get('objective_key', 'makespan')}"
        if request_payload["mode"] == "single"
        else f"EXACT:WEIGHTED:{uuid.uuid5(uuid.NAMESPACE_DNS, json.dumps(request_payload, ensure_ascii=False, sort_keys=True)).hex[:8]}"
    )
    chosen["source"] = "exact_reference_fallback"
    chosen["rank"] = None
    chosen["feasible"] = bool(analytics.feasible)
    chosen["evaluation_mode"] = "exact_simulation_selection"
    if request_payload["mode"] == "weighted":
        weighted_score = 0.0
        requested_keys = list(request_payload.get("objective_weights", {}).keys())
        ranges = {}
        for key in requested_keys:
            values = [float(_ensure_solution_objectives(current_shop, item, [key]).get(key, 0.0)) for item in candidates.values()]
            ranges[key] = (min(values), max(values))
        for key, weight in request_payload.get("objective_weights", {}).items():
            raw = float(objective_payload.get(key, analytics.objective_values.get(key, 0.0)))
            lower, upper = ranges[key]
            span = upper - lower
            if span <= 1e-9:
                normalized_value = 0.0
            else:
                normalized_value = (raw - lower) / span
            if EXACT_OBJECTIVES[key].direction == "max":
                normalized_value = 1.0 - normalized_value
            weighted_score += weight * normalized_value
        objective_payload["weighted_score"] = round(weighted_score, 6)
    chosen["objectives"] = objective_payload
    chosen["metrics"] = {
        **{key: round(float(value), 4) for key, value in analytics.objective_values.items()},
        **{
            key: round(float(value), 4)
            for key, value in winner.get("metrics", {}).items()
            if isinstance(value, (int, float)) and key not in analytics.objective_values
        },
        "evaluation_mode": "exact_simulation_selection",
        "status": "FALLBACK_SIMULATION",
        "solve_time_s": 0.0,
    }
    chosen["exact_info"] = {
        "mode": request_payload["mode"],
        "objective_key": request_payload.get("objective_key"),
        "objective_weights": request_payload.get("objective_weights", {}),
        "support": request_payload.get("support", {}),
        "time_limit_s": req.time_limit_s,
        "solve_time_s": 0.0,
        "status": "FALLBACK_SIMULATION",
        "fallback_reason": reason,
        "selected_from": winner.get("solution_id"),
    }
    chosen["summary"] = {
        "completed_operations": analytics.summary.get("completed_operations", analytics.completed_operations),
        "total_operations": analytics.summary.get("total_operations", len(current_shop.operations)),
        "tardy_order_ids": analytics.tardy_order_ids,
        "tardy_task_ids": analytics.tardy_task_ids,
        "bottleneck_machine_ids": analytics.bottleneck_machine_ids,
        "avg_utilization": round(float(analytics.objective_values.get("avg_utilization", 0.0)), 4),
        "critical_utilization": round(float(analytics.objective_values.get("critical_utilization", 0.0)), 4),
        "tooling_utilization": round(float(analytics.objective_values.get("tooling_utilization", 0.0)), 4),
        "personnel_utilization": round(float(analytics.objective_values.get("personnel_utilization", 0.0)), 4),
        "evaluation_mode": "exact_simulation_selection",
        "exact_mode": request_payload["mode"],
    }
    return chosen


def _build_exact_reference_solution(
    current_shop: ShopFloor,
    task: dict,
    objective_keys: list[str],
    req: ExactReferenceReq,
    schedule_limit: int | None = 120,
) -> dict:
    mode = (req.mode or "single").lower()
    if mode not in {"single", "weighted"}:
        raise HTTPException(400, "精确参考方案模式仅支持 single 或 weighted")

    if mode == "single":
        objective_key = req.objective_key or "makespan"
        if objective_key not in EXACT_OBJECTIVES:
            raise HTTPException(400, f"不支持的精确目标: {objective_key}")
        solver = ExactSolver(current_shop, objectives=[objective_key], time_limit_s=req.time_limit_s)
    else:
        weights = {
            key: float(value)
            for key, value in (req.objective_weights or {}).items()
            if key in EXACT_OBJECTIVES and abs(float(value)) > 1e-9
        }
        if not weights:
            raise HTTPException(400, "加权精确方案至少需要提供一个非零权重目标")
        solver = ExactSolver(current_shop, objectives=list(weights.keys()), objective_weights=weights, time_limit_s=req.time_limit_s)

    result = solver.solve()
    if result.status not in {"OPTIMAL", "FEASIBLE"}:
        return _build_exact_reference_fallback(
            current_shop=current_shop,
            task=task,
            objective_keys=objective_keys,
            req=req,
            schedule_limit=schedule_limit,
            reason=result.status,
        )

    normalized_request = result.request or {}
    key_material = json.dumps(normalized_request, ensure_ascii=False, sort_keys=True)
    suffix = uuid.uuid5(uuid.NAMESPACE_DNS, key_material).hex[:8]
    solution_id = (
        f"EXACT:SINGLE:{normalized_request.get('objective_key', 'makespan')}"
        if normalized_request.get("mode") == "single"
        else f"EXACT:WEIGHTED:{suffix}"
    )
    analytics = build_schedule_analytics(current_shop, SimResult(schedule=result.schedule))
    metrics = {
        **{key: round(float(value), 4) for key, value in analytics.objective_values.items()},
        **{key: round(float(value), 4) for key, value in result.objectives.items()},
        "solve_time_s": round(result.solve_time_s, 3),
        "status": result.status,
        "evaluation_mode": "exact_solver",
    }
    summary = {
        "completed_operations": analytics.summary.get("completed_operations", analytics.completed_operations),
        "total_operations": analytics.summary.get("total_operations", len(current_shop.operations)),
        "tardy_order_ids": analytics.tardy_order_ids,
        "tardy_task_ids": analytics.tardy_task_ids,
        "bottleneck_machine_ids": analytics.bottleneck_machine_ids,
        "avg_utilization": round(float(analytics.objective_values.get("avg_utilization", 0.0)), 4),
        "critical_utilization": round(float(analytics.objective_values.get("critical_utilization", 0.0)), 4),
        "tooling_utilization": round(float(analytics.objective_values.get("tooling_utilization", 0.0)), 4),
        "personnel_utilization": round(float(analytics.objective_values.get("personnel_utilization", 0.0)), 4),
        "evaluation_mode": "exact_solver",
        "exact_mode": normalized_request.get("mode"),
    }
    payload_objectives = {
        key: round(float(analytics.objective_values.get(key, result.objectives.get(key, 0.0))), 4)
        for key in objective_keys
        if key in OBJECTIVE_SPECS
    }
    if normalized_request.get("mode") == "weighted" and "weighted_score" in result.objectives:
        payload_objectives["weighted_score"] = round(float(result.objectives["weighted_score"]), 6)

    schedule = []
    schedule_entries = list(result.schedule or [])
    if schedule_limit is not None:
        schedule_entries = schedule_entries[:schedule_limit]
    for entry in schedule_entries:
        task = current_shop.tasks.get(entry["task_id"])
        order = current_shop.orders.get(task.order_id) if task else None
        schedule.append(
            _serialize_schedule_entry(
                current_shop,
                {
                    **entry,
                    "order_id": order.id if order else "",
                    "order_name": order.name if order else "",
                    "priority": order.priority if order else 1,
                    "due_date": round(order.due_date, 3) if order else 0,
                    "due_at": current_shop.time_label(order.due_date) if order else None,
                    "is_tardy": (entry["end"] > order.due_date) if order else False,
                    "is_main": task.is_main if task else False,
                },
            )
        )

    exact_info = {
        "mode": normalized_request.get("mode"),
        "objective_key": normalized_request.get("objective_key"),
        "objective_weights": normalized_request.get("objective_weights", {}),
        "support": normalized_request.get("support", {}),
        "time_limit_s": req.time_limit_s,
        "solve_time_s": round(result.solve_time_s, 3),
        "status": result.status,
    }
    return {
        "solution_id": solution_id,
        "rule_name": None,
        "source": "exact_reference",
        "generation": 0,
        "rank": None,
        "feasible": bool(analytics.feasible),
        "evaluation_mode": "exact_solver",
        "objectives": payload_objectives,
        "metrics": metrics,
        "summary": summary,
        "candidate": {
            "seed_rule_name": None,
            "graph_profile": "exact_reference",
            "destroy_fraction": 0.0,
            "feature_weights": {},
            "bias_count": 0,
        },
        "exact_info": exact_info,
        "schedule": schedule,
    }


def _json_safe(value: Any) -> Any:
    """递归地将 inf/-inf/nan 等非 JSON 合规的浮点值替换为 None。

    不可行排程会产生 makespan=inf、total_tardiness=inf 等指标，
    starlette 默认的 json.dumps 无法序列化这些值，会抛出
    "Out of range float values are not JSON compliant: inf" 并返回 500。
    前端已将 null 渲染为 "-"，因此此处统一转换为 None。
    """
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, dict):
        return {key: _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    return value


def _active_shop() -> Optional[ShopFloor]:
    """返回当前实例；库内容未变时复用上次构建的对象。

    build_shopfloor() 要把整库行数据重建成对象图（生产规模实例约 2s，其中 SQL 读取
    只占 70ms，其余是对象构造和 build_indexes），而此前每个接口每次调用都重建一遍。
    改为按 inst_version 版本号缓存：读版本号约 0.9ms，任何写库操作都会让它递增
    （见 data.db._bump_instance_version）。

    共享同一个 ShopFloor 是安全的，因为 Simulator 只读传入的实例、排程写在
    SimulationRuntime 的深拷贝上（见 core.sim_runtime）；ensure_calendar_capacity
    也是幂等的“不足才补”。
    """
    global _active_shop_cache
    if inst_store.has_data():
        version = get_instance_version(inst_store.db_path)
        cached = _active_shop_cache
        if cached is not None and cached[0] == version:
            return cached[1]
        current_shop = inst_store.build_shopfloor()
        _active_shop_cache = (version, current_shop)
    else:
        current_shop = shop
    if current_shop is not None:
        _ensure_shop_calendar_capacity(current_shop)
    return current_shop


def _objective_catalog_map() -> dict[str, dict]:
    return {item["key"]: item for item in objective_summary_payload()}


def _resolve_hybrid_task(task_id: Optional[str]) -> tuple[str, dict]:
    global _latest_hybrid_task_id
    resolved_id = task_id or _latest_hybrid_task_id
    if not resolved_id and _hybrid_tasks:
        done_ids = [tid for tid, item in _hybrid_tasks.items() if item.get("status") == "done" and item.get("result")]
        if done_ids:
            resolved_id = done_ids[-1]
    if not resolved_id or resolved_id not in _hybrid_tasks:
        raise HTTPException(404, "未找到可用的帕累托优化结果，请先完成一次混合优化")
    task = _hybrid_tasks[resolved_id]
    if task.get("status") == "error":
        raise HTTPException(500, task.get("error", "优化任务失败"))
    if task.get("status") != "done" or not task.get("result"):
        raise HTTPException(400, "该优化任务尚未完成，请等待混合优化结束")
    return resolved_id, task


def _safe_json_loads(raw_text: str) -> dict | None:
    if not raw_text:
        return None
    text = raw_text.strip()
    if "```json" in text:
        text = text.split("```json", 1)[1]
        text = text.split("```", 1)[0]
    elif "```" in text:
        text = text.split("```", 1)[1]
        text = text.split("```", 1)[0]
    text = text.strip()
    if not text:
        return None
    try:
        return json.loads(text)
    except Exception:
        start = text.find("{")
        end = text.rfind("}")
        if start >= 0 and end > start:
            try:
                return json.loads(text[start : end + 1])
            except Exception:
                return None
    return None


def _top_feature_weights(candidate: dict, limit: int = 6) -> list[tuple[str, float]]:
    weights = candidate.get("feature_weights", {}) or {}
    return sorted(
        ((key, float(value)) for key, value in weights.items() if abs(float(value)) > 1e-6),
        key=lambda item: abs(item[1]),
        reverse=True,
    )[:limit]


AI_PRIORITY_KPI_KEYS = [
    "total_tardiness",
    "makespan",
    "main_order_tardy_total_time",
    "main_order_tardy_count",
    "main_order_tardy_ratio",
    "tardy_job_count",
    "max_tardiness",
    "avg_tardiness",
    "avg_flowtime",
    "max_flowtime",
    "total_wait_time",
    "total_completion_time",
    "avg_utilization",
    "critical_utilization",
    "avg_active_window_utilization",
    "critical_active_window_utilization",
    "avg_net_available_utilization",
    "critical_net_available_utilization",
    "tooling_utilization",
    "personnel_utilization",
    "bottleneck_load_balance",
    "assembly_sync_penalty",
]


def _solution_metric_value(solution: dict, key: str):
    objectives = solution.get("objectives", {}) or {}
    metrics = solution.get("metrics", {}) or {}
    summary = solution.get("summary", {}) or {}
    if key in objectives:
        return objectives.get(key)
    if key in metrics:
        return metrics.get(key)
    if key in summary:
        return summary.get(key)
    return None


def _collect_ai_metric_keys(result: dict, selected: list[dict]) -> tuple[list[str], list[str]]:
    primary_keys = [key for key in (result.get("objective_keys") or []) if key in OBJECTIVE_SPECS]
    baseline = result.get("baseline", {}) or {}
    all_solutions = [baseline, *selected]
    full_keys: list[str] = []
    for key in [*primary_keys, *AI_PRIORITY_KPI_KEYS]:
        if key not in OBJECTIVE_SPECS or key in full_keys:
            continue
        if any(_solution_metric_value(solution, key) is not None for solution in all_solutions):
            full_keys.append(key)
    if any(_solution_metric_value(solution, "weighted_score") is not None for solution in all_solutions):
        full_keys.append("weighted_score")
    return primary_keys, full_keys


def _ai_metric_label(key: str, catalog_map: dict[str, dict]) -> str:
    if key == "weighted_score":
        return "加权综合得分"
    return catalog_map.get(key, {}).get("label", key)


def _ai_metric_direction(key: str, catalog_map: dict[str, dict]) -> str:
    if key == "weighted_score":
        return "min"
    return catalog_map.get(key, {}).get("direction", "min")


def _ai_metric_description(key: str, catalog_map: dict[str, dict]) -> str:
    if key == "weighted_score":
        return "按业务偏好对多个指标归一化后的综合得分，越小越优。"
    return catalog_map.get(key, {}).get("description", "")


def _solution_prompt_block(
    current_shop: Optional[ShopFloor],
    solution: dict,
    primary_objective_keys: list[str],
    full_kpi_keys: list[str],
    catalog_map: dict[str, dict],
    schedule_limit: int = 12,
) -> str:
    summary = solution.get("summary", {}) or {}
    candidate = solution.get("candidate", {}) or {}
    objectives = solution.get("objectives", {}) or {}
    deltas = solution.get("delta_vs_baseline", {}) or {}
    top_weights = _top_feature_weights(candidate)
    weight_text = ", ".join(f"{name}={value:.3f}" for name, value in top_weights) or "-"
    schedule_lines: list[str] = []
    for entry in (solution.get("schedule") or [])[:schedule_limit]:
        op_id = entry.get("op_id", "-")
        op_name = entry.get("op_name")
        machine_name = entry.get("machine_name") or entry.get("machine_id") or "-"
        start_value = entry.get("start")
        end_value = entry.get("end")
        start_at = entry.get("start_at") or (current_shop.time_label(start_value) if current_shop and start_value is not None else start_value)
        end_at = entry.get("end_at") or (current_shop.time_label(end_value) if current_shop and end_value is not None else end_value)
        tooling_text = ",".join(entry.get("tooling_ids", []) or []) or "-"
        personnel_text = ",".join(entry.get("personnel_ids", []) or []) or "-"
        schedule_lines.append(
            f"- {op_id}{f'({op_name})' if op_name else ''}: machine={machine_name}, "
            f"start={start_at}, end={end_at}, toolings={tooling_text}, personnel={personnel_text}"
        )
    primary_lines = []
    for key in primary_objective_keys:
        label = _ai_metric_label(key, catalog_map)
        direction = _ai_metric_direction(key, catalog_map)
        primary_lines.append(f"- {label}({key}, {direction}): value={_solution_metric_value(solution, key)}, delta_vs_baseline={deltas.get(key)}")
    holistic_lines = []
    for key in full_kpi_keys:
        label = _ai_metric_label(key, catalog_map)
        direction = _ai_metric_direction(key, catalog_map)
        holistic_lines.append(f"- {label}({key}, {direction}): value={_solution_metric_value(solution, key)}")
    return "\n".join(
        [
            f"SolutionID={solution.get('solution_id')}, source={solution.get('source')}, generation={solution.get('generation')}, rank={solution.get('rank')}, feasible={solution.get('feasible')}, evaluation_mode={solution.get('evaluation_mode')}",
            "PrimaryObjectives:",
            *(primary_lines or ["- no_primary_objectives"]),
            "HolisticKPIs:",
            *(holistic_lines or ["- no_holistic_kpis"]),
            (
                "RiskSummary: "
                f"completed_ops={summary.get('completed_operations')}/{summary.get('total_operations')}, "
                f"bottlenecks={summary.get('bottleneck_machine_ids', [])}, "
                f"tardy_orders={summary.get('tardy_order_ids', [])[:8]}, "
                f"tardy_tasks={summary.get('tardy_task_ids', [])[:8]}, "
                f"avg_utilization={summary.get('avg_utilization')}, "
                f"critical_utilization={summary.get('critical_utilization')}, "
                f"tooling_utilization={summary.get('tooling_utilization')}, "
                f"personnel_utilization={summary.get('personnel_utilization')}"
            ),
            (
                "RuleProfile: "
                f"seed_rule={candidate.get('seed_rule_name')}, graph_profile={candidate.get('graph_profile')}, "
                f"destroy_fraction={candidate.get('destroy_fraction')}, top_feature_weights=[{weight_text}]"
            ),
            "ScheduleExcerpt:",
            *(schedule_lines or ["- no_schedule_excerpt"]),
        ]
    )


def _baseline_prompt_block(
    current_shop: Optional[ShopFloor],
    baseline: dict,
    primary_objective_keys: list[str],
    full_kpi_keys: list[str],
    catalog_map: dict[str, dict],
    schedule_limit: int = 10,
) -> str:
    objectives = baseline.get("objectives", {}) or {}
    metrics = baseline.get("metrics", {}) or {}
    summary = baseline.get("summary", {}) or {}
    schedule_lines: list[str] = []
    for entry in (baseline.get("schedule") or [])[:schedule_limit]:
        start_value = entry.get("start")
        end_value = entry.get("end")
        schedule_lines.append(
            f"- {entry.get('op_id', '-')}: machine={entry.get('machine_name') or entry.get('machine_id') or '-'}, "
            f"start={entry.get('start_at') or (current_shop.time_label(start_value) if current_shop and start_value is not None else start_value)}, "
            f"end={entry.get('end_at') or (current_shop.time_label(end_value) if current_shop and end_value is not None else end_value)}"
        )
    primary_lines = []
    for key in primary_objective_keys:
        label = _ai_metric_label(key, catalog_map)
        direction = _ai_metric_direction(key, catalog_map)
        primary_lines.append(f"- {label}({key}, {direction})={_solution_metric_value(baseline, key)}")
    holistic_lines = []
    for key in full_kpi_keys:
        label = _ai_metric_label(key, catalog_map)
        direction = _ai_metric_direction(key, catalog_map)
        holistic_lines.append(f"- {label}({key}, {direction})={_solution_metric_value(baseline, key)}")
    return "\n".join(
        [
            f"BaselineRule={baseline.get('rule_name')}, solution_id={baseline.get('solution_id')}, evaluation_mode={baseline.get('evaluation_mode')}",
            "PrimaryObjectives:",
            *(primary_lines or ["- no_primary_objectives"]),
            "HolisticKPIs:",
            *(holistic_lines or ["- no_holistic_kpis"]),
            "MetricsRaw: " + ", ".join(f"{key}={value}" for key, value in metrics.items()),
            (
                "RiskSummary: "
                f"bottlenecks={summary.get('bottleneck_machine_ids', [])}, "
                f"tardy_orders={summary.get('tardy_order_ids', [])[:8]}, "
                f"tardy_tasks={summary.get('tardy_task_ids', [])[:8]}"
            ),
            "ScheduleExcerpt:",
            *(schedule_lines or ["- no_schedule_excerpt"]),
        ]
    )


def _history_prompt_block(conversation: list[dict[str, str]], limit: int = 6) -> str:
    if not conversation:
        return "No previous conversation."
    lines = []
    for item in conversation[-limit:]:
        role = item.get("role", "user")
        content = (item.get("content") or "").strip()
        if content:
            lines.append(f"{role.upper()}: {content[:1200]}")
    return "\n".join(lines) if lines else "No previous conversation."


def _normalize_objective_score(values: list[float], target: float, direction: str) -> float:
    if not values:
        return 0.5
    lower = min(values)
    upper = max(values)
    if upper - lower <= 1e-9:
        return 0.5
    if direction == "max":
        return (target - lower) / (upper - lower)
    return (upper - target) / (upper - lower)


def _infer_requirement_weights(requirement: str, objective_keys: list[str]) -> dict[str, float]:
    text = (requirement or "").strip().lower()
    weights = {key: 0.1 for key in objective_keys}
    if not text:
        return weights
    keyword_groups = {
        "total_tardiness": ["交期", "准时", "延迟", "延误", "逾期", "拖期", "按时"],
        "tardy_job_count": ["超期单", "超期数量", "延误单"],
        "max_tardiness": ["最坏", "最大延误", "最大拖期"],
        "makespan": ["完工", "总周期", "工期", "尽快结束", "尽快完成", "产出"],
        "avg_flowtime": ["流转", "在制", "周转"],
        "total_wait_time": ["等待", "排队", "在制", "停滞"],
        "avg_utilization": ["利用率", "稼动率", "设备效率"],
        "critical_utilization": ["瓶颈", "关键设备", "关键资源"],
        "avg_active_window_utilization": ["连续利用", "活跃窗口", "开工到完工", "连续稼动"],
        "critical_active_window_utilization": ["关键设备连续利用", "关键资源连续利用"],
        "avg_net_available_utilization": ["净可用", "可用时间利用", "班次内利用"],
        "critical_net_available_utilization": ["关键设备净可用", "关键资源净可用"],
        "assembly_sync_penalty": ["装配", "同步", "齐套", "配套", "主任务"],
        "tooling_utilization": ["工装", "夹具"],
        "personnel_utilization": ["人员", "班组", "操作工"],
    }
    for key, keywords in keyword_groups.items():
        if key not in weights:
            continue
        hits = sum(1 for word in keywords if word in text)
        if hits:
            weights[key] += 1.2 + 0.25 * hits
    if any(word in text for word in ["平衡", "综合", "折中", "兼顾"]):
        for key in weights:
            weights[key] += 0.25
    total = sum(weights.values()) or 1.0
    return {key: value / total for key, value in weights.items()}


def _combined_requirement_weights(requirement: str, primary_keys: list[str], full_keys: list[str]) -> dict[str, float]:
    keys = [key for key in full_keys if key != "weighted_score"]
    weights = _infer_requirement_weights(requirement, keys)
    if not (requirement or "").strip():
        for key in primary_keys:
            if key in weights:
                weights[key] += 0.7
        for key in keys:
            if key not in primary_keys:
                weights[key] += 0.18
    else:
        for key in primary_keys:
            if key in weights:
                weights[key] += 0.25
    total = sum(weights.values()) or 1.0
    return {key: value / total for key, value in weights.items()}


def _heuristic_solution_fit(solution: dict, solutions: list[dict], primary_keys: list[str], full_keys: list[str], catalog_map: dict[str, dict], requirement: str) -> float:
    weights = _combined_requirement_weights(requirement, primary_keys, full_keys)
    score = 0.0
    for key in [key for key in full_keys if key != "weighted_score"]:
        series = [
            float(_solution_metric_value(item, key) or 0.0)
            for item in solutions
        ]
        direction = _ai_metric_direction(key, catalog_map)
        target = float(_solution_metric_value(solution, key) or 0.0)
        score += weights.get(key, 0.0) * _normalize_objective_score(series, target, direction)
    summary = solution.get("summary", {}) or {}
    if any(word in (requirement or "") for word in ["稳定", "稳健", "风险"]):
        score += 0.08 * (1.0 - min(1.0, len(summary.get("tardy_order_ids", [])) / max(1, summary.get("total_operations", 1))))
    if any(word in (requirement or "") for word in ["可解释", "规则", "容易理解"]):
        score += 0.05 if solution.get("candidate", {}).get("seed_rule_name") else 0.0
    return score


def _format_solution_brief_text(solution: dict, objective_keys: list[str], catalog_map: dict[str, dict]) -> str:
    parts = [f"{solution.get('solution_id')} ({solution.get('source')})"]
    for key in objective_keys[:5]:
        label = _ai_metric_label(key, catalog_map)
        parts.append(f"{label}={_solution_metric_value(solution, key)}")
    candidate = solution.get("candidate", {}) or {}
    parts.append(f"seed_rule={candidate.get('seed_rule_name') or '-'}")
    parts.append(f"graph_profile={candidate.get('graph_profile') or '-'}")
    return " | ".join(parts)


def _heuristic_compare_payload(result: dict, selected: list[dict], requirement: str, task_id: str) -> dict:
    catalog_map = _objective_catalog_map()
    primary_keys, full_kpi_keys = _collect_ai_metric_keys(result, selected)
    best_by_objective = []
    for key in primary_keys:
        direction = _ai_metric_direction(key, catalog_map)
        if direction == "max":
            winner = max(selected, key=lambda item: float(_solution_metric_value(item, key) or 0.0))
        else:
            winner = min(selected, key=lambda item: float(_solution_metric_value(item, key) or 0.0))
        best_by_objective.append(
            f"{_ai_metric_label(key, catalog_map)} 最优: {winner.get('solution_id')} -> {_solution_metric_value(winner, key)}"
        )
    holistic_focus = []
    for key in [key for key in full_kpi_keys if key not in primary_keys][:4]:
        valid_values = [item for item in selected if _solution_metric_value(item, key) is not None]
        if not valid_values:
            continue
        direction = _ai_metric_direction(key, catalog_map)
        if direction == "max":
            winner = max(valid_values, key=lambda item: float(_solution_metric_value(item, key) or 0.0))
        else:
            winner = min(valid_values, key=lambda item: float(_solution_metric_value(item, key) or 0.0))
        holistic_focus.append(f"{_ai_metric_label(key, catalog_map)} 领先: {winner.get('solution_id')} -> {_solution_metric_value(winner, key)}")
    recommendation = max(selected, key=lambda item: _heuristic_solution_fit(item, selected, primary_keys, full_kpi_keys, catalog_map, requirement or ""))
    paragraphs = [
        f"本次比较共纳入 {len(selected)} 个候选方案（含帕累托解、精确参考方案与启发式规则），先按主目标，再结合全量 KPI 与风险信息做综合判断。",
        f"如果按照当前输入的业务偏好来做快速筛选，更贴近要求的方案是 {recommendation.get('solution_id')}。",
        "主目标上的局部优势如下：",
        *best_by_objective,
        "全量 KPI 与风险补充观察：",
        *(holistic_focus or ["- 当前候选之间的全量 KPI 差异不明显。"]),
        "候选方案摘要：",
        *[_format_solution_brief_text(item, full_kpi_keys, catalog_map) for item in selected],
    ]
    return {
        "mode": "compare",
        "task_id": task_id,
        "used_solution_ids": [item.get("solution_id") for item in selected],
        "analysis": {
            "headline": f"共比较 {len(selected)} 个候选方案",
            "recommended_solution_id": recommendation.get("solution_id"),
            "summary_points": best_by_objective,
            "solution_briefs": [
                {
                    "solution_id": item.get("solution_id"),
                    "source": item.get("source"),
                    "seed_rule_name": (item.get("candidate") or {}).get("seed_rule_name"),
                    "graph_profile": (item.get("candidate") or {}).get("graph_profile"),
                    "bottleneck_machine_ids": (item.get("summary") or {}).get("bottleneck_machine_ids", []),
                }
                for item in selected
            ],
        },
        "display_text": "\n".join(paragraphs),
        "used_model": "heuristic-fallback",
    }


def _heuristic_recommend_payload(result: dict, selected: list[dict], requirement: str, task_id: str) -> dict:
    catalog_map = _objective_catalog_map()
    primary_keys, full_kpi_keys = _collect_ai_metric_keys(result, selected)
    requirement_weights = _combined_requirement_weights(requirement, primary_keys, full_kpi_keys)
    ranked = sorted(
        selected,
        key=lambda item: _heuristic_solution_fit(item, selected, primary_keys, full_kpi_keys, catalog_map, requirement),
        reverse=True,
    )
    winner = ranked[0]
    fit_points = []
    for key, weight in sorted(requirement_weights.items(), key=lambda item: item[1], reverse=True):
        value = _solution_metric_value(winner, key)
        if value is None:
            continue
        fit_points.append(f"{_ai_metric_label(key, catalog_map)} = {value}")
        if len(fit_points) >= 3:
            break
    tradeoffs = []
    for key in full_kpi_keys:
        value = _solution_metric_value(winner, key)
        if value is None:
            continue
        tradeoffs.append(f"{_ai_metric_label(key, catalog_map)} 相对值 = {value}")
        if len(tradeoffs) >= 4:
            break
    display_text = "\n".join(
        [
            f"基于当前业务要求，推荐优先关注方案 {winner.get('solution_id')}。",
            (
                "推荐原因："
                + ("已结合你强调的诉求，并按主目标、全量 KPI 与风险三层信息做综合推荐。" if requirement else "未提供附加要求，因此先以主目标表现为主，再综合全量 KPI 与风险做平衡推荐。")
            ),
            "最匹配的几个点：",
            *fit_points,
            "需要同步确认的取舍：",
            *tradeoffs,
            "可进一步追问该方案的规则来源、瓶颈资源、超期订单和前几道关键工序。",
        ]
    )
    return {
        "mode": "recommend",
        "task_id": task_id,
        "used_solution_ids": [item.get("solution_id") for item in selected],
        "analysis": {
            "recommended_solution_id": winner.get("solution_id"),
            "reason": display_text,
            "fit_points": fit_points,
            "tradeoffs": tradeoffs,
        },
        "display_text": display_text,
        "used_model": "heuristic-fallback",
    }


def _heuristic_ask_payload(solution: dict, question: str, task_id: str) -> dict:
    candidate = solution.get("candidate", {}) or {}
    summary = solution.get("summary", {}) or {}
    weights = _top_feature_weights(candidate)
    weight_text = ", ".join(f"{key}={value:.3f}" for key, value in weights) or "-"
    question_text = question or ""
    answer_lines = [f"当前回答围绕方案 {solution.get('solution_id')} 展开。"]
    if any(word in question_text for word in ["规则", "rule", "逻辑"]):
        answer_lines.append(
            f"该方案的规则画像主要来自 seed_rule={candidate.get('seed_rule_name') or '-'}，graph_profile={candidate.get('graph_profile') or '-'}，重点权重为 {weight_text}。"
        )
    if any(word in question_text for word in ["过程", "排程", "步骤", "中间"]):
        answer_lines.append("这个方案的前几道关键工序如下：")
        for entry in (solution.get("schedule") or [])[:8]:
            answer_lines.append(
                f"- {entry.get('op_id')} 在 {entry.get('machine_name') or entry.get('machine_id') or '-'} 上加工，开始 {entry.get('start_at') or entry.get('start')}, 结束 {entry.get('end_at') or entry.get('end')}"
            )
    if any(word in question_text for word in ["瓶颈", "超期", "风险", "问题"]):
        answer_lines.append(
            f"该方案当前识别出的瓶颈设备为 {summary.get('bottleneck_machine_ids', []) or '-'}，超期订单为 {summary.get('tardy_order_ids', []) or '-'}。"
        )
        answer_lines.append(
            f"同时还要关注总等待={_solution_metric_value(solution, 'total_wait_time')}、平均流程时间={_solution_metric_value(solution, 'avg_flowtime')}、平均利用率={_solution_metric_value(solution, 'avg_utilization')}。"
        )
    if len(answer_lines) == 1:
        answer_lines.append(
            f"该方案来自 {solution.get('source')}，可行性={solution.get('feasible')}，平均利用率={summary.get('avg_utilization')}，关键资源利用率={summary.get('critical_utilization')}，总等待={_solution_metric_value(solution, 'total_wait_time')}，平均流程时间={_solution_metric_value(solution, 'avg_flowtime')}。"
        )
        answer_lines.append(f"如果你关心规则、瓶颈、超期订单或前几道工序过程，可以继续具体追问。")
    display_text = "\n".join(answer_lines)
    return {
        "mode": "ask",
        "task_id": task_id,
        "used_solution_ids": [solution.get("solution_id")],
        "analysis": {
            "solution_id": solution.get("solution_id"),
            "answer": display_text,
            "seed_rule_name": candidate.get("seed_rule_name"),
            "graph_profile": candidate.get("graph_profile"),
        },
        "display_text": display_text,
        "used_model": "heuristic-fallback",
    }


def _invoke_pareto_llm(prompt: str, role: str, action: str) -> tuple[dict | None, str, str]:
    config = get_config().llm
    if not config.api_key:
        return None, "", "heuristic-fallback"
    llm = LLMInterface(config.api_key, config.base_url, config.model)
    raw = llm.call(prompt, role, action, temp=0.25)
    return _safe_json_loads(raw), raw, config.model or "configured-llm"

# === Instance ===
@app.post("/api/instance/generate")
async def gen_instance(req: GenReq):
    global shop, last_result
    gen = InstanceGenerator(seed=req.seed)
    shop = gen.generate(
        num_orders=req.num_orders,
        tasks_per_order=(req.tasks_per_order_min, req.tasks_per_order_max),
        ops_per_task=(req.ops_per_task_min, req.ops_per_task_max),
        machines_per_type=req.machines_per_type,
        due_date_factor=req.due_date_factor,
        arrival_spread=req.arrival_spread,
        day_shift_hours=req.day_shift_hours,
        night_shift_hours=req.night_shift_hours,
        schedule_days=req.schedule_days,
        maintenance_prob=req.maintenance_prob,
        toolings_per_type=req.toolings_per_type,
        personnel_per_skill=req.personnel_per_skill,
        plan_start_at=req.plan_start_at,
    )
    calendar_info = _ensure_shop_calendar_capacity(shop)
    last_result = None
    # 保存到数据库
    inst_store.save_from_shopfloor(shop)
    _invalidate_graph_context("instance_generated")
    return {"status": "ok", "summary": shop.summary(), "calendar": calendar_info, "details": _instance_details(shop), "validation": _validate_instance(shop)}

@app.get("/api/instance/details")
async def inst_details():
    global shop
    if not shop and inst_store.has_data():
        shop = inst_store.build_shopfloor()
    if not shop:
        raise HTTPException(400, "请先生成实例或导入CSV")
    return _instance_details(shop)

@app.get("/api/instance/csv")
async def export_csv():
    if not shop: raise HTTPException(400, "请先生成实例")
    return PlainTextResponse(shop.to_csv(), media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=instance.csv"})

@app.get("/api/instance/db")
async def get_instance_from_db():
    """从数据库获取实例原始数据(可编辑), 含概览"""
    global shop
    if not inst_store.has_data():
        raise HTTPException(400, "数据库中无实例数据")
    data = inst_store.load_all()
    # 班次在库里是紧凑字符串（"day/start/hours;..."），前端"机器维护"页按 JSON 数组渲染，
    # 这里统一解析成结构化数组，避免文本框显示成空的 []。
    for resource_key in ("machines", "toolings", "personnel"):
        for row in data.get(resource_key, []):
            row["shifts"] = shifts_to_payload(row.get("shifts"))
    # 构建 shop 并返回概览
    if not shop:
        shop = inst_store.build_shopfloor()
    data["summary"] = shop.summary()
    data["details"] = _instance_details(shop)
    return data

@app.put("/api/instance/order/{order_id}")
async def update_order(order_id: str, data: dict):
    global shop
    inst_store.update_order(order_id, data)
    shop = inst_store.build_shopfloor()
    _invalidate_graph_context("order_updated")
    return {"status": "ok"}

@app.put("/api/instance/task/{task_id}")
async def update_task(task_id: str, data: dict):
    global shop
    inst_store.update_task(task_id, data)
    shop = inst_store.build_shopfloor()
    _invalidate_graph_context("task_updated")
    return {"status": "ok"}

@app.put("/api/instance/operation/{op_id}")
async def update_operation(op_id: str, data: dict):
    global shop
    if str(data.get("turnover_time", "")).strip():
        try:
            turnover_value = float(data["turnover_time"])
        except (TypeError, ValueError):
            turnover_value = float("nan")
        if not (math.isfinite(turnover_value) and turnover_value >= 0):
            raise HTTPException(400, f"流转等待时长非法（{data['turnover_time']}），必须是不小于 0 的有限数")
    initial_start_time = datetime_to_offset_hours(_plan_start_ref(), data.get("initial_start_at", data.get("initial_start_time")))
    initial_end_time = datetime_to_offset_hours(_plan_start_ref(), data.get("initial_end_at", data.get("initial_end_time")))
    data = {
        **data,
        "initial_status": _normalize_initial_status(data.get("initial_status")),
        "initial_start_time": initial_start_time if initial_start_time is not None else None,
        "initial_end_time": initial_end_time if initial_end_time is not None else None,
        "turnover_time": float(data["turnover_time"]) if str(data.get("turnover_time", "")).strip() else 0.0,
        "initial_remaining_processing_time": float(data["initial_remaining_processing_time"]) if str(data.get("initial_remaining_processing_time", "")).strip() else None,
        "initial_assigned_machine_id": data.get("initial_assigned_machine_id", ""),
        "initial_assigned_tooling_ids": data.get("initial_assigned_tooling_ids", ""),
        "initial_assigned_personnel_ids": data.get("initial_assigned_personnel_ids", ""),
    }
    inst_store.update_operation(op_id, data)
    shop = inst_store.build_shopfloor()
    _invalidate_graph_context("operation_updated")
    return {"status": "ok"}

@app.put("/api/instance/machine/{machine_id}")
async def update_machine(machine_id: str, data: dict):
    global shop
    inst_store.update_machine(machine_id, data)
    shop = inst_store.build_shopfloor()
    _invalidate_graph_context("machine_updated")
    return {"status": "ok"}

@app.post("/api/instance/import-excel")
async def import_excel(file: UploadFile = File(...)):
    """导入Excel文件(含5个sheet), 覆盖数据库"""
    global shop
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        def sheet_to_dicts(sheet_name, required=True):
            if sheet_name not in wb.sheetnames:
                if required:
                    raise ValueError(f"缺少sheet: {sheet_name}")
                return []
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 1:
                return []
            headers = [str(h or '').strip() for h in rows[0]]
            return [{headers[i]: (str(v) if v is not None else '') for i, v in enumerate(row)}
                    for row in rows[1:] if any(v is not None for v in row)]

        planning_rows = sheet_to_dicts("planning_context", required=False)
        orders_rows = sheet_to_dicts("orders")
        tasks_rows = sheet_to_dicts("tasks")
        ops_rows = sheet_to_dicts("operations")
        mt_rows = sheet_to_dicts("machine_types")
        machines_rows = sheet_to_dicts("machines")
        tooling_type_rows = sheet_to_dicts("tooling_types", required=False)
        tooling_rows = sheet_to_dicts("toolings", required=False)
        personnel_rows = sheet_to_dicts("personnel", required=False)
        downtime_rows = sheet_to_dicts("downtimes", required=False)
        initial_state_rows = sheet_to_dicts("initial_state", required=False)
        wb.close()

        plan_start_at = None
        if planning_rows:
            plan_start_raw = planning_rows[0].get("plan_start_at", "")
            if plan_start_raw:
                plan_start_at = datetime.fromisoformat(str(plan_start_raw).replace("Z", "+00:00"))

        converted_initial_state = []
        plan_ref = plan_start_at or inst_store.get_plan_start_at()
        for row in initial_state_rows:
            remaining_raw = str(row.get("initial_remaining_processing_time", "")).strip()
            converted_initial_state.append(
                {
                    "op_id": row.get("op_id", ""),
                    "initial_status": _normalize_initial_status(row.get("initial_status", "")),
                    "initial_start_time": _coerce_excel_offset(plan_ref, row.get("initial_start_time", None)) if str(row.get("initial_start_time", "")).strip() else None,
                    "initial_end_time": _coerce_excel_offset(plan_ref, row.get("initial_end_time", None)) if str(row.get("initial_end_time", "")).strip() else None,
                    "initial_remaining_processing_time": float(remaining_raw) if remaining_raw else None,
                    "initial_assigned_machine_id": row.get("initial_assigned_machine_id", ""),
                    "initial_assigned_tooling_ids": row.get("initial_assigned_tooling_ids", ""),
                    "initial_assigned_personnel_ids": row.get("initial_assigned_personnel_ids", ""),
                }
            )

        inst_store.save_from_csv(
            orders_rows,
            tasks_rows,
            ops_rows,
            mt_rows,
            machines_rows,
            tooling_type_rows=tooling_type_rows,
            tooling_rows=tooling_rows,
            personnel_rows=personnel_rows,
            initial_state_rows=converted_initial_state,
            plan_start_at=plan_start_at,
        )
        if downtime_rows:
            converted_downtimes = []
            plan_ref = plan_start_at or inst_store.get_plan_start_at()
            for row in downtime_rows:
                converted_downtimes.append(
                    {
                        "machine_id": row["machine_id"],
                        "downtime_type": row.get("downtime_type", "planned") or "planned",
                        "start_time": _coerce_excel_offset(plan_ref, row.get("start_time", 0)),
                        "end_time": _coerce_excel_offset(plan_ref, row.get("end_time", 0)),
                    }
                )
            downtime_store.replace_all(converted_downtimes)
        else:
            downtime_store.clear_all()
        shop = inst_store.build_shopfloor()
        _invalidate_graph_context("instance_imported")

        validation = _validate_instance(shop)
        logging.info(
            "excel import: orders=%d tasks=%d operations=%d validation=%s errors=%d",
            len(shop.orders), len(shop.tasks), len(shop.operations),
            validation["status"], validation["error_count"],
        )
        return {"status": "ok", "summary": shop.summary(), "details": _instance_details(shop), "validation": validation}
    except HTTPException:
        raise
    except Exception as e:
        logging.exception("Excel import failed")
        raise HTTPException(400, f"Excel导入失败: {e}")

@app.get("/api/instance/template")
async def download_template():
    """下载Excel模板文件"""
    payload = build_instance_template_bytes()
    return Response(
        content=payload,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="instance_template_v20260326_1.xlsx"',
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )

def _instance_details(s: ShopFloor):
    orders = []
    for order_id, order in s.orders.items():
        tasks_info = []
        for task_id in order.task_ids:
            task = s.tasks.get(task_id)
            if not task:
                continue
            ops_info = []
            for op in task.operations:
                ops_info.append(
                    {
                        "id": op.id,
                        "name": op.name,
                        "type": op.process_type,
                        "time": op.processing_time,
                        "turnover_time": op.turnover_time,
                        "predecessors": op.predecessor_ops + op.predecessor_tasks,
                        "required_tooling_types": op.required_tooling_types,
                        "required_personnel_skills": op.required_personnel_skills,
                        "derived_due_date": op.derived_due_date,
                        "derived_due_at": s.time_label(op.derived_due_date),
                        "derived_start_time": op.derived_start_time,
                        "derived_start_at": s.time_label(op.derived_start_time),
                        "critical_slack": round(op.critical_slack, 3) if op.critical_slack < float("inf") else None,
                        "initial_status": op.status.value,
                        "initial_start_time": op.start_time,
                        "initial_start_at": s.time_label(op.start_time) if op.start_time is not None else None,
                        "initial_end_time": op.end_time,
                        "initial_end_at": s.time_label(op.end_time) if op.end_time is not None else None,
                        "initial_remaining_processing_time": op.remaining_processing_time,
                        "initial_assigned_machine_id": op.assigned_machine_id,
                        "initial_assigned_tooling_ids": op.assigned_tooling_ids,
                        "initial_assigned_personnel_ids": op.assigned_personnel_ids,
                    }
                )
            tasks_info.append(
                {
                    "id": task_id,
                    "name": task.name,
                    "is_main": task.is_main,
                    "predecessors": task.predecessor_task_ids,
                    "release": task.release_time,
                    "release_at": s.time_label(task.release_time),
                    "due_date": task.due_date,
                    "due_at": s.time_label(task.due_date),
                    "derived_due_date": task.derived_due_date,
                    "derived_due_at": s.time_label(task.derived_due_date),
                    "derived_start_time": task.derived_start_time,
                    "derived_start_at": s.time_label(task.derived_start_time),
                    "critical_path_time": round(task.critical_path_time, 3),
                    "critical_slack": round(task.critical_slack, 3) if task.critical_slack < float("inf") else None,
                    "ops": ops_info,
                }
            )
        orders.append(
            {
                "id": order_id,
                "name": order.name,
                "due_date": order.due_date,
                "due_at": s.time_label(order.due_date),
                "priority": order.priority,
                "release": order.release_time,
                "release_at": s.time_label(order.release_time),
                "tasks": tasks_info,
            }
        )
    machines = []
    for machine_id, machine in s.machines.items():
        machine_type = s.machine_types.get(machine.type_id)
        calendar = _resource_calendar_payload(s, machine)
        machines.append(
            {
                "id": machine_id,
                "name": machine.name,
                "type": machine.type_id,
                "type_name": machine_type.name if machine_type else "",
                "is_critical": machine_type.is_critical if machine_type else False,
                "shifts": calendar["shifts"],
                "shift_windows": calendar["shifts"],
                "downtimes": calendar["downtimes"],
            }
        )
    machine_types = [
        {
            "id": type_id,
            "name": machine_type.name,
            "is_critical": machine_type.is_critical,
            "count": len(s._machine_by_type.get(type_id, [])),
        }
        for type_id, machine_type in s.machine_types.items()
    ]
    tooling_types = [
        {"id": type_id, "name": tooling_type.name, "count": len(s._tooling_by_type.get(type_id, []))}
        for type_id, tooling_type in s.tooling_types.items()
    ]
    toolings = [
        {"id": tool.id, "name": tool.name, "type": tool.type_id, "type_name": s.tooling_types.get(tool.type_id).name if tool.type_id in s.tooling_types else ""}
        for tool in s.toolings.values()
    ]
    personnel = [
        {"id": person.id, "name": person.name, "skills": person.skills}
        for person in s.personnel.values()
    ]
    return {
        "plan_start_at": s.time_label(0.0),
        "orders": orders,
        "machines": machines,
        "machine_types": machine_types,
        "tooling_types": tooling_types,
        "toolings": toolings,
        "personnel": personnel,
        "summary": s.summary(),
    }

# === Graph ===
GRAPH_BUILD_TIMEOUT_S = int(os.environ.get("LLM4DRD_GRAPH_TIMEOUT_S", "600"))
GRAPH_WARN_EDGES = int(os.environ.get("LLM4DRD_GRAPH_WARN_EDGES", "300000"))
GRAPH_MAX_EDGES = int(os.environ.get("LLM4DRD_GRAPH_MAX_EDGES", "2000000"))
GRAPH_MAX_NODES = int(os.environ.get("LLM4DRD_GRAPH_MAX_NODES", "100000"))


def _validate_instance(current_shop: ShopFloor) -> dict:
    """对当前实例做强校验：数据完整性、关联关系、约束条件。

    错误级别问题（error）会直接导致仿真/优化静默失败（例如空排程、指标全为 0），
    必须在“实例与约束”页显式暴露给用户。
    """
    errors: list[dict] = []
    warnings: list[dict] = []

    def err(category: str, entity: str, message: str, sheet: str = "-"):
        errors.append({"severity": "error", "category": category, "entity": entity, "message": message, "sheet": sheet})

    def warn(category: str, entity: str, message: str, sheet: str = "-"):
        warnings.append({"severity": "warning", "category": category, "entity": entity, "message": message, "sheet": sheet})

    # --- 关联关系：任务 → 订单 / 前置任务 ---
    for task_id, task in current_shop.tasks.items():
        if task.order_id not in current_shop.orders:
            err("关联关系", task_id, f"任务引用了不存在的订单 {task.order_id}", sheet="tasks / orders")
        for predecessor_id in task.predecessor_task_ids:
            if predecessor_id not in current_shop.tasks:
                err("关联关系", task_id, f"前置任务 {predecessor_id} 不存在，该任务的后续工序将永远无法就绪", sheet="tasks")

    # --- 数据完整性 + 约束条件：工序 ---
    ops_without_machine = 0
    for op_id, op in current_shop.operations.items():
        if op.task_id not in current_shop.tasks:
            err("关联关系", op_id, f"工序引用了不存在的任务 {op.task_id}", sheet="operations / tasks")
        if op.processing_time is None or float(op.processing_time) <= 0:
            err("数据完整性", op_id, f"加工时长非法（{op.processing_time}），必须大于 0", sheet="operations")
        if op.turnover_time is not None and not (math.isfinite(float(op.turnover_time)) and float(op.turnover_time) >= 0):
            err("数据完整性", op_id, f"流转等待时长非法（{op.turnover_time}），必须是不小于 0 的有限数", sheet="operations")
        for predecessor_id in op.predecessor_ops:
            if predecessor_id not in current_shop.operations:
                err("关联关系", op_id, f"前置工序 {predecessor_id} 不存在，该工序将永远无法就绪（仿真会输出空排程）", sheet="operations")
        for predecessor_task_id in op.predecessor_tasks:
            if predecessor_task_id not in current_shop.tasks:
                err("关联关系", op_id, f"前置任务 {predecessor_task_id} 不存在，该工序将永远无法就绪", sheet="operations / tasks")
        for machine_field in op.eligible_machine_ids:
            # 兼容旧数据里未拆分的逗号分隔串，逐个机台号与 machines sheet 比对
            for machine_token in machine_field.replace("，", ",").split(","):
                machine_token = machine_token.strip()
                if machine_token and machine_token not in current_shop.machines:
                    err("关联关系", op_id, f"工序 {op.name or op_id} 指定的可用机台 {machine_token} 在 machines sheet 中不存在（machine_id 不匹配）", sheet="operations / machines")
        if not current_shop.get_eligible_machines(op):
            ops_without_machine += 1
            if ops_without_machine <= 20:
                err("约束条件", op_id, f"没有任何可用机器（工艺类型 {op.process_type}，指定机器 {op.eligible_machine_ids or '按类型匹配'}）", sheet="operations / machines")
        for tooling_type in op.required_tooling_types:
            if not current_shop.get_toolings_for_type(tooling_type):
                err("约束条件", op_id, f"缺少所需工装类型 {tooling_type} 的任何实例", sheet="operations / toolings")
        for skill_id in op.required_personnel_skills:
            if not current_shop.get_personnel_for_skill(skill_id):
                err("约束条件", op_id, f"缺少具备技能 {skill_id} 的任何人员", sheet="operations / personnel")
    if ops_without_machine > 20:
        err("约束条件", "operations", f"另有 {ops_without_machine - 20} 道工序同样没有可用机器（已省略明细）", sheet="operations / machines")

    # --- 资源日历可用性：某类资源有实例、但全部没有任何可用工作窗口 ---
    # （班次可能都落在计划起点之前，或被停机完全占满；仿真会因此永远排不出相关工序。
    #   "完全没有实例" 的情况已在上面的按工序检查里覆盖，这里只查 "有实例但排不了班"。）
    used_process_types = {op.process_type for op in current_shop.operations.values()}
    for process_type in sorted(used_process_types):
        machines = current_shop.get_machines_for_type(process_type)
        if machines and all(not _resource_has_calendar(machine) for machine in machines):
            err("约束条件", process_type, f"工艺类型 {process_type} 的全部机器都没有任何可用排班窗口（班次可能都落在计划起点之前或被停机占满），相关工序无法开工", sheet="machines")

    used_tooling_types: set[str] = set()
    used_skills: set[str] = set()
    for op in current_shop.operations.values():
        used_tooling_types.update(op.required_tooling_types)
        used_skills.update(op.required_personnel_skills)
    for tooling_type in sorted(used_tooling_types):
        toolings = current_shop.get_toolings_for_type(tooling_type)
        if toolings and all(not _resource_has_calendar(tooling) for tooling in toolings):
            err("约束条件", tooling_type, f"工装类型 {tooling_type} 的全部实例都没有任何可用排班窗口，需要该工装的工序无法开工", sheet="toolings")
    for skill_id in sorted(used_skills):
        people = current_shop.get_personnel_for_skill(skill_id)
        if people and all(not _resource_has_calendar(person) for person in people):
            err("约束条件", skill_id, f"技能 {skill_id} 的全部人员都没有任何可用排班窗口，需要该技能的工序无法开工", sheet="personnel")

    # --- 工序前驱环检测（有环则整条链永远无法就绪）---
    color: dict[str, int] = {}
    cycle_reported = False
    for start_id in current_shop.operations:
        if cycle_reported or color.get(start_id):
            continue
        stack = [(start_id, iter(current_shop.operations[start_id].predecessor_ops))]
        color[start_id] = 1
        while stack:
            node_id, iterator = stack[-1]
            advanced = False
            for predecessor_id in iterator:
                if predecessor_id not in current_shop.operations:
                    continue
                state = color.get(predecessor_id, 0)
                if state == 1:
                    err("关联关系", node_id, f"工序前驱存在循环依赖（涉及 {predecessor_id}），相关工序永远无法开工", sheet="operations")
                    cycle_reported = True
                    stack.clear()
                    advanced = True
                    break
                if state == 0:
                    color[predecessor_id] = 1
                    stack.append((predecessor_id, iter(current_shop.operations[predecessor_id].predecessor_ops)))
                    advanced = True
                    break
            if not advanced and stack:
                finished_id, _ = stack.pop()
                color[finished_id] = 2

    # --- 订单交期与结构 ---
    for order_id, order in current_shop.orders.items():
        if not order.task_ids:
            warn("数据完整性", order_id, "订单下没有任何任务", sheet="orders / tasks")
        if math.isfinite(order.due_date) and order.due_date < order.release_time:
            warn("约束条件", order_id, f"交期（{order.due_date:.1f}h）早于释放时间（{order.release_time:.1f}h），必然延误", sheet="orders")

    # --- 资源日历容量 ---
    calendar_info = _ensure_shop_calendar_capacity(current_shop)
    if calendar_info.get("final_days", 0) < calendar_info.get("required_days", 0):
        err("约束条件", "calendar", f"资源班次日历（{calendar_info['final_days']} 天）无法覆盖预计排产跨度（{calendar_info['required_days']} 天），后段工序将无法安排", sheet="machines / personnel")

    status = "failed" if errors else ("warning" if warnings else "passed")
    return {
        "status": status,
        "errors": errors,
        "warnings": warnings,
        "error_count": len(errors),
        "warning_count": len(warnings),
        "stats": {
            "orders": len(current_shop.orders),
            "tasks": len(current_shop.tasks),
            "operations": len(current_shop.operations),
            "machines": len(current_shop.machines),
            "toolings": len(current_shop.toolings),
            "personnel": len(current_shop.personnel),
            "calendar": calendar_info,
        },
        "checked_at": datetime.now().astimezone().isoformat(),
    }


def _validation_payload(current_shop, force: bool = False) -> dict:
    """当前实例的校验结论；库里有匹配当前实例的结论就直接取，不重算。"""
    if not force:
        cached = workflow_store.load("validation")
        if cached:
            return cached
    validation = _validate_instance(current_shop)
    _save_workflow_progress("validation", validation)
    logging.info(
        "instance validation: status=%s errors=%d warnings=%d ops=%d",
        validation["status"], validation["error_count"], validation["warning_count"],
        len(current_shop.operations),
    )
    return validation


@app.get("/api/instance/validate")
async def validate_instance(force: bool = False):
    current_shop = _active_shop()
    if current_shop is None:
        raise HTTPException(400, "当前没有可用实例，请先生成或导入")
    return _validation_payload(current_shop, force=force)


@app.get("/api/instance/validate/export")
async def export_validation_excel():
    """导出数据强校验的完整结果为 Excel（不受前端仅展示前 50 条的限制）。"""
    current_shop = _active_shop()
    if current_shop is None:
        raise HTTPException(400, "当前没有可用实例，请先生成或导入")
    validation = _validation_payload(current_shop)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "校验问题明细"
    ws.append(["级别", "问题Sheet", "类别", "实体", "问题明细"])
    for item in list(validation["errors"]) + list(validation["warnings"]):
        ws.append([
            "错误" if item.get("severity") == "error" else "警告",
            item.get("sheet", "-"),
            item.get("category", "-"),
            item.get("entity", "-"),
            item.get("message", "-"),
        ])

    ws_summary = wb.create_sheet("校验概要")
    stats = validation.get("stats", {})
    for key, value in [
        ("校验状态", validation["status"]),
        ("错误数", validation["error_count"]),
        ("警告数", validation["warning_count"]),
        ("校验时间", validation["checked_at"]),
        ("订单数", stats.get("orders", 0)),
        ("任务数", stats.get("tasks", 0)),
        ("工序数", stats.get("operations", 0)),
        ("机器数", stats.get("machines", 0)),
        ("工装数", stats.get("toolings", 0)),
        ("人员数", stats.get("personnel", 0)),
    ]:
        ws_summary.append([key, value])

    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 80)

    buffer = io.BytesIO()
    wb.save(buffer)
    filename = f"validation_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_sim_export_excel(payload: dict, shop: Optional[ShopFloor]) -> bytes:
    """把一次仿真结果（/api/simulate 返回的 payload）导出为多 sheet 的 Excel。"""
    gantt = payload.get("gantt") or []
    metrics = payload.get("metrics") or {}
    rule = payload.get("rule") or "-"
    diagnosis = payload.get("diagnosis")
    diagnosis_detail = payload.get("diagnosis_detail")
    scheduled_op_ids = {str(e.get("op_id")) for e in gantt}

    wb = openpyxl.Workbook()

    # Sheet 1: 排程结果（甘特明细）
    ws = wb.active
    ws.title = "排程结果"
    headers = [
        "工序ID", "任务ID", "订单ID", "订单名称", "机器ID", "工序类型",
        "开始(小时)", "开始时间", "结束(小时)", "结束时间", "时长(小时)",
        "状态", "优先级", "交期", "是否延误", "是否主订单",
    ]
    ws.append(headers)
    for e in sorted(gantt, key=lambda x: (x.get("start") or 0)):
        ws.append([
            e.get("op_id", ""),
            e.get("task_id", ""),
            e.get("order_id", ""),
            e.get("order_name", ""),
            e.get("machine_id", ""),
            e.get("process_type", ""),
            e.get("start"),
            e.get("start_at"),
            e.get("end"),
            e.get("end_at"),
            e.get("duration"),
            e.get("status") or "completed",
            e.get("priority", ""),
            e.get("due_at"),
            "是" if e.get("is_tardy") else "否",
            "是" if e.get("is_main") else "否",
        ])

    # Sheet 2: 关键指标
    ws_metrics = wb.create_sheet("关键指标")
    ws_metrics.append(["指标", "数值"])
    metric_order = [
        ("规则", rule),
        ("可行性 feasible", metrics.get("feasible")),
        ("完成工序 / 总工序", f'{metrics.get("completed_operations")} / {metrics.get("total_operations")}'),
        ("总周期 makespan(小时)", metrics.get("makespan")),
        ("总延误 total_tardiness(小时)", metrics.get("total_tardiness")),
        ("平均延误 avg_tardiness(小时)", metrics.get("avg_tardiness")),
        ("最大延误 max_tardiness(小时)", metrics.get("max_tardiness")),
        ("延误订单数", metrics.get("tardy_job_count")),
        ("平均流程时间 avg_flowtime(小时)", metrics.get("avg_flowtime")),
        ("总等待时间 total_wait_time(小时)", metrics.get("total_wait_time")),
        ("平均等待时间 avg_wait_time(小时)", metrics.get("avg_wait_time")),
        ("净可用利用率 avg_net_available_utilization", metrics.get("avg_net_available_utilization")),
        ("关键资源净可用利用率 critical_net_available_utilization", metrics.get("critical_net_available_utilization")),
        ("平均利用率 avg_utilization", metrics.get("avg_utilization")),
        ("关键资源利用率 critical_utilization", metrics.get("critical_utilization")),
        ("主订单延误数 main_order_tardy_count", metrics.get("main_order_tardy_count")),
        ("主订单延误总时长 main_order_tardy_total_time(小时)", metrics.get("main_order_tardy_total_time")),
        ("主订单延误比例 main_order_tardy_ratio", metrics.get("main_order_tardy_ratio")),
        ("总主订单数 total_main_orders", metrics.get("total_main_orders")),
    ]
    for key, value in metric_order:
        ws_metrics.append([key, value])
    known = {k for k, _ in metric_order}
    for key, value in metrics.items():
        if key not in known and not isinstance(value, (dict, list)):
            ws_metrics.append([key, value])

    # Sheet 3: 延误明细
    ws_tardy = wb.create_sheet("延误明细")
    ws_tardy.append(["订单ID", "订单名称", "工序ID", "机器ID", "结束时间", "交期", "超出(小时)"])
    for e in sorted(gantt, key=lambda x: (x.get("end") or 0), reverse=True):
        if e.get("is_tardy"):
            due = e.get("due_date")
            end = e.get("end")
            over = round(end - due, 3) if isinstance(end, (int, float)) and isinstance(due, (int, float)) else ""
            ws_tardy.append([
                e.get("order_id", ""), e.get("order_name", ""), e.get("op_id", ""),
                e.get("machine_id", ""), e.get("end_at"), e.get("due_at"), over,
            ])

    # Sheet 4: 诊断与未完成工序
    ws_diag = wb.create_sheet("诊断")
    if diagnosis:
        ws_diag.append(["仿真诊断", diagnosis])
        ws_diag.append([])
    if diagnosis_detail:
        if isinstance(diagnosis_detail, dict):
            for k, v in diagnosis_detail.items():
                ws_diag.append([str(k), str(v)])
        else:
            ws_diag.append([str(diagnosis_detail)])
    if shop is not None:
        unscheduled = [op for op in shop.operations.values() if str(op.id) not in scheduled_op_ids]
        if unscheduled:
            ws_diag.append([])
            ws_diag.append(["未完成工序数", len(unscheduled)])
            ws_diag.append(["工序ID", "任务ID", "工序类型", "状态"])
            for op in unscheduled:
                ws_diag.append([op.id, op.task_id, op.process_type, getattr(op, "status", "")])
    if ws_diag.max_row == 1 and ws_diag.cell(1, 1).value is None:
        ws_diag.append(["无", "本次仿真结果完整，无诊断信息"])

    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 80)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@app.post("/api/simulate/export-excel")
async def export_sim_excel():
    """导出最近一次仿真结果为 Excel（供用户下载）。"""
    current_shop = _active_shop()
    if last_sim_payload is None:
        raise HTTPException(400, "尚未运行仿真，无法导出。请先在仿真与洞察中运行一次规则仿真。")
    try:
        data = _build_sim_export_excel(last_sim_payload, current_shop)
    except Exception as exc:  # noqa: BLE001
        logging.exception("sim export excel failed")
        raise HTTPException(500, f"导出 Excel 失败：{exc}")
    rule = last_sim_payload.get("rule") or "sim"
    filename = f"sim_result_{rule}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _estimate_graph_size(current_shop: ShopFloor) -> dict:
    node_count = (
        len(current_shop.orders) + len(current_shop.tasks) + len(current_shop.operations)
        + len(current_shop.machines) + len(current_shop.toolings) + len(current_shop.personnel)
    )
    task_predecessor_edges = {
        (predecessor_id, task_id)
        for task_id, task in current_shop.tasks.items()
        for predecessor_id in task.predecessor_task_ids
    }
    task_predecessor_edges.update(
        (predecessor_id, op.task_id)
        for op in current_shop.operations.values()
        if op.task_id in current_shop.tasks
        for predecessor_id in op.predecessor_tasks
    )
    structural_edges = len(current_shop.tasks) + len(task_predecessor_edges)
    structural_edges += len(current_shop.operations)
    structural_edges += sum(len(op.predecessor_ops) + len(op.predecessor_tasks) for op in current_shop.operations.values())
    machine_edges = 0
    tooling_edges = 0
    personnel_edges = 0
    for op in current_shop.operations.values():
        machine_edges += len(op.eligible_machine_ids or current_shop._machine_by_type.get(op.process_type, []))
        tooling_edges += sum(len(current_shop._tooling_by_type.get(type_id, [])) for type_id in op.required_tooling_types)
        personnel_edges += sum(len(current_shop._personnel_by_skill.get(skill_id, [])) for skill_id in op.required_personnel_skills)
    total_edges = structural_edges + machine_edges + tooling_edges + personnel_edges
    return {
        "estimated_nodes": node_count,
        "estimated_edges": total_edges,
        "structural_edges": structural_edges,
        "machine_edges": machine_edges,
        "tooling_edges": tooling_edges,
        "personnel_edges": personnel_edges,
    }


@app.post("/api/graph/build")
async def build_graph(bg: BackgroundTasks, force: bool = False):
    """在后台构建异构图，并通过状态接口报告规模、阶段、进度与错误。

    指纹未变时默认复用已持久化的图谱产物（秒回），这样任务流中途失败重试不必重建；
    force=True 用于用户明确要求重建。
    """
    global shop, _graph_tasks
    if not shop and not inst_store.has_data():
        raise HTTPException(400, "请先生成实例")

    for existing_id, existing in _graph_tasks.items():
        if existing.get("status") in {"queued", "running"}:
            return {"task_id": existing_id, "status": existing["status"], "message": "已有图谱构建任务正在运行"}

    current_shop = _active_shop()
    task_id = str(uuid.uuid4())[:8]
    _graph_tasks = dict(list(_graph_tasks.items())[-19:])
    _graph_tasks[task_id] = {
        "status": "queued",
        "stage": "queued",
        "progress": 0,
        "message": "任务已提交，正在准备规模预检",
        "started_at": time.time(),
        "elapsed_s": 0.0,
        "timeout_s": GRAPH_BUILD_TIMEOUT_S,
        "estimate": None,
        "stats": None,
        "warning": None,
        "error": None,
    }

    def update(**values):
        task = _graph_tasks[task_id]
        task.update(values)
        task["elapsed_s"] = round(time.time() - task["started_at"], 2)

    def run():
        global shop
        deadline = time.monotonic() + GRAPH_BUILD_TIMEOUT_S
        try:
            if current_shop is None:
                update(status="running", stage="loading", progress=1, message="正在从数据库恢复车间实例；数据量较大时此步骤可能需要一些时间")
                build_shop = inst_store.build_shopfloor()
                shop = build_shop
            else:
                build_shop = current_shop
            update(status="running", stage="preflight", progress=3, message="正在估算节点、关系边和内存压力")
            estimate = _estimate_graph_size(build_shop)
            update(estimate=estimate, progress=8, message=f"规模预检完成：预计 {estimate['estimated_nodes']:,} 个节点、{estimate['estimated_edges']:,} 条边")
            if estimate["estimated_nodes"] > GRAPH_MAX_NODES:
                raise ValueError(f"数据量过大：预计节点 {estimate['estimated_nodes']:,}，超过安全上限 {GRAPH_MAX_NODES:,}")
            if estimate["estimated_edges"] > GRAPH_MAX_EDGES:
                raise ValueError(
                    f"数据量过大：预计关系边 {estimate['estimated_edges']:,}，超过安全上限 {GRAPH_MAX_EDGES:,}。"
                    "请减少每道工序的可用机器、工装或人员范围后重试。"
                )
            if estimate["estimated_edges"] > GRAPH_WARN_EDGES:
                update(warning=f"预计 {estimate['estimated_edges']:,} 条边，构建可能持续较长时间，请保持页面开启")

            def build_progress(processed, total, nodes, edges):
                ratio = processed / max(total, 1)
                if processed >= total:
                    update(
                        stage="saving",
                        progress=66,
                        message=(
                            f"内存图构建完成，正在编译并保存 {nodes:,} 个节点、"
                            f"{edges:,} 条边"
                        ),
                    )
                    return
                update(
                    stage="building",
                    progress=min(64, round(10 + ratio * 54)),
                    message=f"正在构建内存图：已处理 {processed:,}/{total:,} 个业务实体，当前 {nodes:,} 个节点、{edges:,} 条边",
                )

            update(
                stage="building",
                progress=10,
                message="正在复用已构建的图谱" if not force else "正在构建订单、任务、工序和资源关系",
            )
            _, diagnostics = graph_context_service.get_or_build(
                build_shop,
                force_rebuild=force,
                progress_callback=build_progress,
                deadline=deadline,
                current_fingerprint_provider=lambda: compute_graph_fingerprint(
                    _active_shop()
                ),
            )
            stats = graph_store.load_meta()
            if not stats:
                raise RuntimeError("图谱构建完成但展示元数据缺失")
            update(
                status="done",
                stage="done",
                progress=100,
                message="图谱已存在，直接复用" if diagnostics.cache_hit else "图谱构建并保存成功",
                stats=stats,
                graph_context=_graph_context_diagnostics_payload(diagnostics),
            )
        except TimeoutError as exc:
            update(status="error", stage="timeout", message="图谱构建超时", error=f"{exc}；已超过 {GRAPH_BUILD_TIMEOUT_S} 秒限制")
        except Exception as exc:
            update(status="error", stage="error", message="图谱构建失败", error=str(exc))

    bg.add_task(run)
    return {"task_id": task_id, "status": "queued", "message": "图谱构建任务已提交", "timeout_s": GRAPH_BUILD_TIMEOUT_S}


@app.get("/api/graph/status/{task_id}")
async def graph_build_status(task_id: str):
    task = _graph_tasks.get(task_id)
    if not task:
        raise HTTPException(404, "图谱构建任务不存在或已过期")
    if task.get("status") in {"queued", "running"}:
        task["elapsed_s"] = round(time.time() - task["started_at"], 2)
    return {"task_id": task_id, **task}

def _graph_meta_payload() -> dict:
    meta = graph_store.load_meta()
    if not meta:
        raise HTTPException(400, "暂无图数据,请先构建图")
    compute_meta = graph_artifact_store.load_context_meta()
    cache_ready = bool(
        compute_meta
        and compute_meta.get("status") == "ready"
        and compute_meta.get("instance_hash") == meta.get("instance_hash")
        and compute_meta.get("topology_hash") == meta.get("topology_hash")
        and compute_meta.get("feature_hash") == meta.get("feature_hash")
        and int(compute_meta.get("schema_version", -1))
        == int(meta.get("schema_version", -2))
        and compute_meta.get("builder_version") == meta.get("builder_version")
        and not compute_meta.get("invalid_reason")
        and not meta.get("invalid_reason")
    )
    return {
        **meta,
        "instance_hash_prefix": str(meta.get("instance_hash", ""))[:12],
        "topology_hash_prefix": str(meta.get("topology_hash", ""))[:12],
        "feature_hash_prefix": str(meta.get("feature_hash", ""))[:12],
        "cache_ready": cache_ready,
        "invalid_reason": meta.get("invalid_reason", "")
        or (compute_meta or {}).get("invalid_reason", ""),
    }


def _require_graph_cache_ready() -> None:
    meta = _graph_meta_payload()
    if not meta["cache_ready"]:
        reason = meta.get("invalid_reason") or "实例或图上下文已变化"
        raise HTTPException(409, f"图谱已失效，请重新构建：{reason}")


@app.get("/api/graph/meta")
async def graph_meta():
    return _graph_meta_payload()

@app.get("/api/graph/nodes")
async def graph_nodes(node_type: str = None, search: str = None,
                      limit: int = 200, offset: int = 0):
    _require_graph_cache_ready()
    total, nodes = graph_store.load_nodes(node_type=node_type, search=search, limit=min(max(limit, 1), 1000), offset=max(offset, 0))
    return {"total": total, "nodes": nodes}

@app.get("/api/graph/edges")
async def graph_edges(edge_type: str = None, search: str = None,
                      limit: int = 200, offset: int = 0):
    _require_graph_cache_ready()
    total, edges = graph_store.load_edges(edge_type=edge_type, search=search, limit=min(max(limit, 1), 1000), offset=max(offset, 0))
    return {"total": total, "edges": edges}

@app.get("/api/graph/order/{order_id}")
async def graph_order(order_id: str):
    _require_graph_cache_ready()
    result = graph_store.load_order_subgraph(order_id)
    if not result["order_id"]:
        raise HTTPException(404, f"图谱中不存在订单 {order_id}")
    return result

@app.get("/api/graph/orders/search")
async def graph_order_search(q: str = ""):
    """按订单号/名称在数据库做模糊解析并返回其完整关联子图（OS_ 机器在 SQL 层过滤）。"""
    query = (q or "").strip()
    if not query:
        raise HTTPException(400, "请输入订单号")
    _require_graph_cache_ready()
    result = graph_store.search_order_subgraph(query)
    if not result["order_id"]:
        raise HTTPException(404, f"没有找到该订单：{query}")
    return result

@app.get("/api/graph/node/{node_id:path}/neighbors")
async def node_neighbors(node_id: str):
    _require_graph_cache_ready()
    return graph_store.get_node_neighbors(node_id)

# === Simulate ===
# 未排程工序的根因分类：category -> (可读标签, 是否为可直接排查的根因)
_INFEASIBLE_REASON_LABELS: dict[str, tuple[str, bool]] = {
    "dangling": ("前驱工序/任务引用不存在（数据断链），永远无法就绪", True),
    "no_machine": ("没有匹配的机器类型，无法分配设备", True),
    "no_tooling": ("所需工装类型没有任何实例", True),
    "no_personnel": ("所需人员技能没有任何实例", True),
    "machine_no_calendar": ("所有匹配机器都没有可用排班日历（工作日历为空）", True),
    "tooling_no_calendar": ("所需工装全部没有可用排班日历", True),
    "personnel_no_calendar": ("所需人员全部没有可用排班日历", True),
    "release_inf": ("工序投放时间为无穷大（订单/任务 release_time 异常）", True),
    "starved": ("前驱已完成但始终抢不到资源（资源竞争 / 日历产能不足）", True),
    "calendar_exhausted": ("已就绪但资源日历覆盖不到工序所需工时（机器/工装/人员某类日历太短），被仿真器挂起", True),
    "dependency_cycle": ("依赖环：工序/任务前驱互相矛盾，环内工序永远无法就绪（需打断环）", True),
    "blocked_by_predecessor": ("被上游未完成工序阻塞（级联受阻，非根因）", False),
}


def _resource_has_calendar(resource) -> bool:
    """资源在计划起点之后是否存在任何可用工作窗口。"""
    try:
        return resource.next_available_time(0.0) != float("inf")
    except Exception:
        return True


# 足够大的上界，用于求某资源在计划期内的总可用工时
_CAL_HORIZON = 1e7


def _resource_available_hours(resource) -> Optional[float]:
    """某资源(机器/工装/人员)在 [0, _CAL_HORIZON] 内的总可用工时。

    单个资源对象没有 calendar_days 属性(那是 ShopFloor 级别方法)，
    用 CalendarResourceMixin.available_time_between 求和可用窗口。
    """
    try:
        return resource.available_time_between(0.0, _CAL_HORIZON)
    except Exception:
        return None


def _calendar_bottleneck_hint(current_shop: ShopFloor, op: Operation) -> list[str]:
    """对"已就绪但排不下(资源日历耗尽)"的工序，定位可用工时 < 所需工时的资源类型。

    返回形如 ['工装:tool_turning(单件可用 656h)', '人员:skill_turning(单人可用 656h)']
    的描述，便于在前端/日志直接看到应核查哪类资源的日历。比较口径: 取该类型中
    "单台/单件/单人最长可用工时"与工序所需工时(取 work_remaining 或 processing_time)。
    """
    try:
        need = float(op.work_remaining if op.work_remaining is not None else (op.processing_time or 0.0))
    except Exception:
        need = 0.0
    if need <= 0:
        return []
    hints: list[str] = []
    machines = current_shop.get_eligible_machines(op)
    if machines:
        avail = max((_resource_available_hours(m) or 0.0) for m in machines)
        if need > avail:
            hints.append(f"机器(单台可用{avail:.0f}h<{need:.0f}h)")
    for t in op.required_tooling_types or []:
        tools = current_shop.get_toolings_for_type(t)
        if tools:
            avail = max((_resource_available_hours(x) or 0.0) for x in tools)
            if need > avail:
                hints.append(f"工装:{t}(单件可用{avail:.0f}h<{need:.0f}h)")
    for s in op.required_personnel_skills or []:
        people = current_shop.get_personnel_for_skill(s)
        if people:
            avail = max((_resource_available_hours(x) or 0.0) for x in people)
            if need > avail:
                hints.append(f"人员:{s}(单人可用{avail:.0f}h<{need:.0f}h)")
    return hints


def _diagnose_infeasible(current_shop: ShopFloor, result: SimResult, analytics,
                         unschedulable_ops: Optional[set] = None) -> dict:
    """对不可行 / 部分排程的仿真结果逐工序做根因分析。

    返回结构化诊断：
    {
      "total": 总工序数, "scheduled": 已完成工序数, "unscheduled": 未完成工序数,
      "reasons": [{"category", "label", "is_root", "count", "examples":[...], "hint_ids":[...]}...]
    }
    每道未完成工序按优先级归入唯一一个根因分类，便于定位真正需要调整的数据。
    """
    schedule = result.schedule or []
    total_ops = len(current_shop.operations)

    completed_op_ids: set[str] = {e.get("op_id") for e in schedule if e.get("op_id")}
    completed_op_ids.update(op.id for op in current_shop.operations.values() if op.status == OpStatus.COMPLETED)

    task_op_ids: dict[str, list[str]] = {}
    for op in current_shop.operations.values():
        task_op_ids.setdefault(op.task_id, []).append(op.id)
    completed_task_ids = {
        task_id for task_id, op_ids in task_op_ids.items()
        if op_ids and all(op_id in completed_op_ids for op_id in op_ids)
    }

    buckets: dict[str, list[str]] = {}
    hint_ids: dict[str, set[str]] = {}
    category_by_op: dict[str, str] = {}

    def _add(category: str, op: Operation, missing: Optional[list[str]] = None) -> None:
        label = op.name or op.id
        buckets.setdefault(category, []).append(f"{label}({op.id})")
        category_by_op[op.id] = category
        for token in missing or []:
            hint_ids.setdefault(category, set()).add(str(token))

    # 0. 依赖环(仿真前 SCC 检测，来自 SimResult.dependency_cycles) —— 环内工序互相
    #    阻塞、永远无法就绪。必须在所有其他根因之前判定：否则环内工序的前驱"都未
    #    完成"，会被误归为"级联受阻"，掩盖真因(互相矛盾的前驱关系)。任务级环与工序
    #    级环修复方向不同(改任务令前驱 / 改 predecessor_ops)，需区分上报。
    cycle_by_op: dict[str, dict] = {}
    for _cyc in getattr(result, "dependency_cycles", []) or []:
        for _oid in _cyc.get("ops", []):
            cycle_by_op[_oid] = _cyc
    _CYCLE_KIND_LABEL = {
        "task": "任务级前驱互锁(改任务令 predecessor_tasks)",
        "op": "工序级前驱环(改 predecessor_ops)",
        "mixed": "混合前驱环(任务+工序前驱)",
    }

    for op in current_shop.operations.values():
        if op.id in completed_op_ids:
            continue

        # 0. 依赖环内工序 —— 最优先根因，永远无法就绪
        _cyc = cycle_by_op.get(op.id)
        if _cyc is not None:
            _kind = _cyc.get("kind", "op")
            _hints = [_CYCLE_KIND_LABEL.get(_kind, _kind)]
            if _kind in ("task", "mixed"):
                _hints.append("涉事任务:" + ";".join(_cyc.get("tasks", [])[:6]))
            if _kind in ("op", "mixed"):
                _hints.append("涉事工序:" + ";".join(_cyc.get("ops", [])[:6]))
            _add("dependency_cycle", op, _hints)
            continue

        # 1. 前驱引用断链：引用了不存在的工序 / 任务
        dangling = [p for p in op.predecessor_ops if p not in current_shop.operations]
        dangling += [p for p in op.predecessor_tasks if p not in current_shop.tasks]
        if dangling:
            _add("dangling", op, dangling)
            continue

        # 2. 没有任何匹配机器（工序类型 / 指定机器都不存在）
        machines = current_shop.get_eligible_machines(op)
        if not machines:
            _add("no_machine", op, [op.process_type])
            continue

        # 3. 所需工装类型没有任何实例
        missing_tool = [t for t in op.required_tooling_types if not current_shop.get_toolings_for_type(t)]
        if missing_tool:
            _add("no_tooling", op, missing_tool)
            continue

        # 4. 所需人员技能没有任何实例
        missing_skill = [s for s in op.required_personnel_skills if not current_shop.get_personnel_for_skill(s)]
        if missing_skill:
            _add("no_personnel", op, missing_skill)
            continue

        # 5. 所有匹配机器都没有可用排班日历
        if all(not _resource_has_calendar(m) for m in machines):
            _add("machine_no_calendar", op, [op.process_type])
            continue

        # 6. 某个所需工装类型的全部实例都没有可用排班日历
        tool_no_cal = [
            t for t in op.required_tooling_types
            if all(not _resource_has_calendar(x) for x in current_shop.get_toolings_for_type(t))
        ]
        if tool_no_cal:
            _add("tooling_no_calendar", op, tool_no_cal)
            continue

        # 7. 某个所需技能的全部人员都没有可用排班日历
        person_no_cal = [
            s for s in op.required_personnel_skills
            if all(not _resource_has_calendar(x) for x in current_shop.get_personnel_for_skill(s))
        ]
        if person_no_cal:
            _add("personnel_no_calendar", op, person_no_cal)
            continue

        # 8. 投放时间无穷大
        if current_shop.get_operation_release_time(op) == float("inf"):
            _add("release_inf", op)
            continue

        # 9a. 已就绪但资源日历覆盖不到所需工时 —— 仿真器将其挂起(_unschedulable_ops)。
        #     这类工序前驱都已完成，若不单列会被误归为"级联受阻/抢不到资源"，
        #     且无法指向真正该改的日历。这里明确标为根因并附带最短资源类型提示。
        if unschedulable_ops and op.id in unschedulable_ops:
            _add("calendar_exhausted", op, _calendar_bottleneck_hint(current_shop, op))
            continue

        # 9. 前驱未完成 —— 级联受阻（受害者，真正根因在上游）
        blocked_pre = [p for p in op.predecessor_ops if p not in completed_op_ids]
        blocked_pre += [p for p in op.predecessor_tasks if p not in completed_task_ids]
        if blocked_pre:
            _add("blocked_by_predecessor", op, blocked_pre)
            continue

        # 10. 前驱都完成、资源类型也齐全，但始终没排上 —— 资源竞争 / 日历产能不足
        _add("starved", op)

    reasons: list[dict] = []
    for category, examples in buckets.items():
        label, is_root = _INFEASIBLE_REASON_LABELS.get(category, (category, True))
        reasons.append({
            "category": category,
            "label": label,
            "is_root": is_root,
            "count": len(examples),
            "examples": examples[:8],
            "hint_ids": sorted(hint_ids.get(category, set()))[:12],
        })
    # 根因优先、数量多的排前面
    reasons.sort(key=lambda r: (not r["is_root"], -r["count"]))

    return {
        "total": total_ops,
        "scheduled": analytics.completed_operations,
        "unscheduled": total_ops - analytics.completed_operations,
        "reasons": reasons,
        "bottlenecks": _bottleneck_root_ops(current_shop, completed_op_ids, category_by_op),
    }


def _bottleneck_root_ops(
    current_shop: ShopFloor,
    completed_op_ids: set[str],
    category_by_op: dict[str, str],
    top_n: int = 5,
) -> list[dict]:
    """找出"卡住最多下游"的根因工序。

    在未完成工序的依赖子图上做拓扑传播：每道被阻塞工序继承其未完成前驱的
    根因标签（集合上限 6 个防爆炸），最终统计每个根因工序传递性阻塞的下游
    工序数。机器少、占用久、又卡在链条上游的工序会自然排到最前——这类工序
    是真正的排产瓶颈，应优先增加可用机台/班次或拆分工时。
    """
    unscheduled_ids = {
        op_id for op_id in current_shop.operations
        if op_id not in completed_op_ids
    }
    if not unscheduled_ids:
        return []
    root_ids = {op_id for op_id, cat in category_by_op.items()
                if cat not in ("blocked_by_predecessor", "dependency_cycle")}
    if not root_ids:
        return []

    # 未完成工序按任务分组（任务级前驱：前驱任务的每道未完成工序都会卡住后继）
    task_unscheduled: dict[str, list[str]] = {}
    for op_id in unscheduled_ids:
        op = current_shop.operations[op_id]
        task_unscheduled.setdefault(op.task_id, []).append(op_id)

    dependents: dict[str, list[str]] = {}
    indegree: dict[str, int] = {op_id: 0 for op_id in unscheduled_ids}
    for op_id in unscheduled_ids:
        op = current_shop.operations[op_id]
        blocking_preds: set[str] = {
            pred for pred in op.predecessor_ops if pred in unscheduled_ids
        }
        for task_id in op.predecessor_tasks:
            blocking_preds.update(task_unscheduled.get(task_id, []))
        blocking_preds.discard(op_id)
        for pred in blocking_preds:
            dependents.setdefault(pred, []).append(op_id)
            indegree[op_id] += 1

    # Kahn 拓扑传播根因标签（有环的残留节点直接跳过，环本身属于数据校验问题）
    MAX_LABELS = 6
    labels: dict[str, set[str]] = {op_id: ({op_id} if op_id in root_ids else set()) for op_id in unscheduled_ids}
    impact: dict[str, int] = {root: 0 for root in root_ids}
    queue = [op_id for op_id, degree in indegree.items() if degree == 0]
    while queue:
        current = queue.pop()
        current_labels = labels[current]
        for dependent in dependents.get(current, []):
            dep_labels = labels[dependent]
            if len(dep_labels) < MAX_LABELS:
                dep_labels.update(list(current_labels)[: MAX_LABELS - len(dep_labels)])
            indegree[dependent] -= 1
            if indegree[dependent] == 0:
                queue.append(dependent)
    for op_id in unscheduled_ids:
        if op_id in root_ids:
            continue
        for root in labels[op_id]:
            if root in impact:
                impact[root] += 1

    ranked = sorted(root_ids, key=lambda op_id: impact.get(op_id, 0), reverse=True)[:top_n]
    payload: list[dict] = []
    for op_id in ranked:
        op = current_shop.operations[op_id]
        machines = current_shop.get_eligible_machines(op)
        payload.append({
            "op_id": op_id,
            "op_name": op.name or op_id,
            "process_type": op.process_type,
            "category": category_by_op.get(op_id, ""),
            "eligible_machine_count": len(machines),
            "eligible_machine_ids": [machine.id for machine in machines][:6],
            "processing_time": round(float(op.processing_time or 0.0), 2),
            "blocked_downstream": impact.get(op_id, 0),
        })
    payload = [item for item in payload if item["blocked_downstream"] > 0] or payload[:1]
    return payload


def _format_infeasible_detail(diag: dict, rule_name: str) -> str:
    """把结构化诊断渲染成多行文本，供后台日志打印，方便逐条排查。"""
    lines: list[str] = []
    lines.append(
        f"仿真[{rule_name}] 不可行：仅完成 {diag['scheduled']}/{diag['total']} 道工序，"
        f"未排 {diag['unscheduled']} 道。逐工序根因分类如下（★=可直接排查的根因）："
    )
    for idx, reason in enumerate(diag["reasons"], 1):
        marker = "★" if reason["is_root"] else "·"
        line = f"  {marker} [{idx}] {reason['label']}：{reason['count']} 道工序"
        if reason["hint_ids"]:
            line += f"；涉及: {', '.join(reason['hint_ids'])}"
        if reason["examples"]:
            line += f"；示例工序: {', '.join(reason['examples'])}"
        lines.append(line)
    if not diag["reasons"]:
        lines.append("  （未识别到明确的结构性根因，请到“实例与约束”页运行数据校验查看完整错误明细）")
    else:
        lines.append("  提示：先处理带 ★ 的根因；“级联受阻”类工序会在上游根因修复后自动恢复。")
    bottlenecks = diag.get("bottlenecks") or []
    if bottlenecks:
        lines.append("  瓶颈工序 TOP（其未排导致最多下游工序受阻，优先增加机台/班次或核查约束）：")
        for rank, item in enumerate(bottlenecks, 1):
            label, _ = _INFEASIBLE_REASON_LABELS.get(item.get("category", ""), (item.get("category", ""), True))
            lines.append(
                f"    {rank}. {item['op_name']}({item['op_id']})：工艺 {item['process_type']}，"
                f"可用机器 {item['eligible_machine_count']} 台，加工 {item['processing_time']}h，"
                f"阻塞下游 {item['blocked_downstream']} 道工序；根因：{label}"
            )
    return "\n".join(lines)


def _diagnosis_oneline(diag: dict, scheduled_entries: int) -> str:
    """从结构化诊断生成一行摘要（供前端 banner / toast 展示）。"""
    hints: list[str] = []
    if scheduled_entries == 0:
        hints.append("仿真没有排出任何工序，所有指标会显示为 0")
    else:
        hints.append(f"仅排出 {diag['scheduled']}/{diag['total']} 道工序，指标只反映部分排程")
    # 优先展示排名靠前的 2 个根因，附数量
    for reason in [r for r in diag["reasons"] if r["is_root"]][:2]:
        detail = reason["label"]
        if reason["hint_ids"]:
            detail += f"（{', '.join(reason['hint_ids'][:5])}）"
        hints.append(f"{reason['count']} 道工序：{detail}")
    hints.append("完整根因分类见后台日志；或到“实例与约束”页运行数据校验查看明细")
    return "；".join(hints)


def _simulation_diagnosis(current_shop: ShopFloor, result: SimResult, analytics) -> Optional[str]:
    """当仿真结果异常（空排程 / 未覆盖全部工序）时，给出可读的一行原因诊断（供前端展示）。"""
    if analytics.feasible:
        return None
    diag = _diagnose_infeasible(current_shop, result, analytics)
    return _diagnosis_oneline(diag, len(result.schedule or []))


@app.post("/api/simulate")
def simulate(req: SimReq):
    # def 端点由 FastAPI 放进线程池执行：仿真大实例可能跑数十秒，
    # 若在事件循环里同步跑会让整个后端（含前端所有请求）失去响应。
    with _sim_lock:
        return _simulate_locked(req)


def _simulate_locked(req: SimReq):
    global last_result, last_sim_payload, shop
    current_shop = _active_shop()
    if current_shop is None:
        raise HTTPException(400, "请先生成实例")
    shop = current_shop
    func = BUILTIN_RULES.get(req.rule_name, BUILTIN_RULES["ATC"])
    sim = Simulator(current_shop, func, runtime=_cached_sim_runtime(current_shop))
    r = sim.run()
    analytics = build_schedule_analytics(current_shop, r)
    # 调试日志：确认计算真实执行，并输出关键中间变量
    logging.info(
        "simulate[%s]: scheduled_entries=%d completed_ops=%d/%d makespan=%.2f "
        "total_tardiness=%.2f avg_net_available_utilization=%.4f feasible=%s events=%d wall=%.0fms",
        req.rule_name, len(r.schedule), analytics.completed_operations, len(current_shop.operations),
        analytics.objective_values.get("makespan", 0.0),
        analytics.objective_values.get("total_tardiness", 0.0),
        analytics.objective_values.get("avg_net_available_utilization", 0.0),
        analytics.feasible, r.event_count, r.wall_time_ms,
    )
    diagnosis = None
    diagnosis_detail = None
    if not analytics.feasible:
        diagnosis_detail = _diagnose_infeasible(
            current_shop, r, analytics,
            unschedulable_ops=getattr(sim, "_unschedulable_ops", None),
        )
        diagnosis = _diagnosis_oneline(diagnosis_detail, len(r.schedule or []))
        # 后台打印逐工序根因分类，方便定位具体的不可行原因
        logging.warning(
            "simulate[%s] infeasible:\n%s",
            req.rule_name, _format_infeasible_detail(diagnosis_detail, req.rule_name),
        )
    r._rule_name = req.rule_name
    last_result = r
    # Enrich gantt with order info
    gantt = []
    for e in r.schedule:
        task = current_shop.tasks.get(e["task_id"])
        order = current_shop.orders.get(task.order_id) if task else None
        gantt.append(
            _serialize_schedule_entry(
                current_shop,
                {
                    **e,
                    "order_id": order.id if order else "",
                    "order_name": order.name if order else "",
                    "priority": order.priority if order else 1,
                    "due_date": round(order.due_date, 3) if order else 0,
                    "due_at": current_shop.time_label(order.due_date) if order else None,
                    "is_tardy": (e["end"] > order.due_date) if order else False,
                    "is_main": task.is_main if task else False,
                },
            )
        )
    metrics = r.to_dict()
    metrics.update({key: round(float(value), 4) for key, value in analytics.objective_values.items()})
    metrics["completed_operations"] = analytics.completed_operations
    metrics["total_operations"] = len(current_shop.operations)
    metrics["feasible"] = analytics.feasible
    payload = _json_safe({
        "metrics": metrics,
        "gantt": gantt,
        "rule": req.rule_name,
        "diagnosis": diagnosis,
        "diagnosis_detail": diagnosis_detail,
    })
    last_sim_payload = payload
    _save_workflow_progress("simulation", payload)
    return payload

@app.post("/api/simulate/compare")
def compare(rule_names: list[str] = None):
    # 同 /api/simulate：def 端点走线程池，避免多规则串行仿真阻塞事件循环。
    current_shop = _active_shop()
    if current_shop is None: raise HTTPException(400, "请先生成实例")
    names = rule_names or get_all_rule_names()
    results = []
    # 逐规则复用同一个 runtime：静态数据(深拷贝/日历/派生时刻/环检测)只建一次。
    with _sim_lock:
        runtime = _cached_sim_runtime(current_shop)
        for n in names:
            if n not in BUILTIN_RULES: continue
            r = Simulator(current_shop, BUILTIN_RULES[n], runtime=runtime).run()
            analytics = build_schedule_analytics(current_shop, r)
            metrics = r.to_dict()
            metrics.update({key: round(float(value), 4) for key, value in analytics.objective_values.items()})
            results.append({"rule": n, "metrics": metrics})
    results.sort(key=lambda x: x["metrics"]["total_tardiness"])
    return _json_safe({"comparison": results})


@app.post("/api/simulate/reference-solutions")
async def simulate_reference_solutions(req: HeuristicReferenceReq):
    current_shop = _active_shop()
    if not current_shop:
        raise HTTPException(400, "请先生成实例")
    rules = [rule for rule in req.rule_names if rule in BUILTIN_RULES]
    payload = [_rule_reference_solution(current_shop, rule, req.objective_keys) for rule in rules]
    return {"solutions": payload}

# === Exact Reference ===
@app.get("/api/exact/objectives")
async def exact_objectives():
    return {"objectives": exact_objective_catalog_payload()}


@app.post("/api/optimize/exact-reference")
async def optimize_exact_reference(req: ExactReferenceReq):
    task_id, task = _resolve_hybrid_task(req.task_id)
    current_shop = _active_shop()
    if current_shop is None:
        raise HTTPException(400, "当前没有可用实例")
    objective_keys = _requested_exact_objective_keys((task.get("result") or {}).get("objective_keys", []), req)
    solution = _build_exact_reference_solution(current_shop, task, objective_keys, req, schedule_limit=120)
    existing = _task_reference_solution_index(task)
    existing[solution["solution_id"]] = solution
    task["reference_solutions"] = list(existing.values())
    export_result = task.get("export_result")
    if export_result is not None:
        export_refs = {item.get("solution_id"): item for item in export_result.get("reference_solutions", []) or [] if item.get("solution_id")}
        full_solution = _build_exact_reference_solution(current_shop, task, objective_keys, req, schedule_limit=None)
        export_refs[full_solution["solution_id"]] = full_solution
        export_result["reference_solutions"] = list(export_refs.values())
    _save_workflow_progress("optimization", {"task_id": task_id, "task": task})
    return {"task_id": task_id, "solution": solution, "reference_solution_count": len(task.get("reference_solutions", []))}

# === Hybrid Optimization ===
@app.get("/api/optimize/objectives")
async def optimize_objectives():
    return {"objectives": objective_summary_payload()}

@app.post("/api/optimize/hybrid")
async def optimize_hybrid(req: HybridOptimizeReq, bg: BackgroundTasks):
    global _hybrid_tasks, _latest_hybrid_task_id
    if not shop and not inst_store.has_data():
        raise HTTPException(400, "请先生成实例")

    current_shop = _active_shop()
    graph_context_mode = resolve_graph_context_mode()
    task_id = str(uuid.uuid4())[:8]
    _latest_hybrid_task_id = task_id
    _hybrid_tasks[task_id] = {
        "status": "running",
        "phase": "initializing",
        "message": "正在初始化优化器并校验实例数据",
        "created_at": time.time(),
        "updated_at": time.time(),
        "config": req.model_dump(),
        "current_generation": 0,
        "archive_size": 0,
        "population_size": req.population_size,
        "coarse_pool_size": 0,
        "feasible_ratio": 0.0,
        "hypervolume": 0.0,
        "elapsed_s": 0.0,
        "total_evaluations": 0,
        "approximate_evaluations": 0,
        "exact_evaluations": 0,
        "bottleneck_machine_ids": [],
        "history": [],
        "result": None,
        "reference_solutions": [],
        "graph_context_mode": graph_context_mode.value,
        "graph_context": None,
    }

    def _run():
        global _latest_hybrid_task_id
        heartbeat_stop = threading.Event()
        heartbeat_started_at = time.time()

        def _heartbeat():
            while not heartbeat_stop.wait(OPTIMIZE_HEARTBEAT_INTERVAL_S):
                task = _hybrid_tasks.get(task_id)
                if not task or task.get("status") != "running":
                    return
                now = time.time()
                task["elapsed_s"] = round(now - heartbeat_started_at, 2)
                task["updated_at"] = now

        heartbeat_thread = threading.Thread(
            target=_heartbeat,
            name=f"hybrid-heartbeat-{task_id}",
            daemon=True,
        )
        heartbeat_thread.start()
        try:
            graph_context = None
            graph_context_diagnostics = None
            if graph_context_mode != GraphContextMode.LEGACY:
                task = _hybrid_tasks[task_id]
                task["phase"] = "graph_context_loading"
                task["message"] = "正在加载或构建统一图上下文"
                task["updated_at"] = time.time()
                graph_context, graph_context_diagnostics = (
                    graph_context_service.get_or_build(
                        current_shop,
                        current_fingerprint_provider=lambda: compute_graph_fingerprint(
                            _active_shop()
                        ),
                    )
                )
                task["graph_context"] = _graph_context_diagnostics_payload(
                    graph_context_diagnostics
                )
                if graph_context_diagnostics.cache_level == "built":
                    task["phase"] = "graph_context_building"
                    task["message"] = "统一图上下文已构建并完成完整性校验"
                task["updated_at"] = time.time()

            optimizer = HybridNSGA3ALNSOptimizer(
                current_shop,
                HybridConfig(
                    objective_keys=req.objective_keys,
                    target_solution_count=req.target_solution_count,
                    time_limit_s=req.time_limit_s,
                    population_size=req.population_size,
                    generations=req.generations,
                    alns_iterations_per_candidate=req.alns_iterations_per_candidate,
                    candidate_filter_multiplier=req.candidate_filter_multiplier,
                    coarse_pool_multiplier=req.coarse_pool_multiplier,
                    elite_refine_ratio=req.elite_refine_ratio,
                    elite_refine_min=req.elite_refine_min,
                    coarse_time_ratio=req.coarse_time_ratio,
                    promotion_pool_multiplier=req.promotion_pool_multiplier,
                    random_promotion_ratio=req.random_promotion_ratio,
                    refine_rounds=req.refine_rounds,
                    alns_aggression=req.alns_aggression,
                    stagnation_generations=req.stagnation_generations,
                    parallel_workers=req.parallel_workers,
                    seed=req.seed,
                    baseline_rule_name=req.baseline_rule_name,
                ),
                graph_context,
                graph_context_mode,
            )
            if optimizer.graph_context_diff is not None:
                _hybrid_tasks[task_id]["graph_context_diff"] = {
                    "total_differences": optimizer.graph_context_diff.total_differences,
                    "relation_differences": list(
                        optimizer.graph_context_diff.relation_differences
                    ),
                    "feature_differences": list(
                        optimizer.graph_context_diff.feature_differences
                    ),
                    "group_differences": list(
                        optimizer.graph_context_diff.group_differences
                    ),
                }

            task = _hybrid_tasks[task_id]
            task["phase"] = "coarse"
            task["message"] = "正在建立基线并初始化候选池"
            task["updated_at"] = time.time()

            def _progress(snapshot: dict):
                task = _hybrid_tasks[task_id]
                phase = snapshot.get("phase", "coarse")
                phase_messages = {
                    "coarse": "正在进行近似评估与候选广搜",
                    "exact_promotion": "正在对优质候选进行精确评估",
                    "elite_refine": "正在使用 ALNS 精修精英方案",
                    "finalize": "正在整理 Pareto 前沿与最终方案",
                }
                task["phase"] = phase
                task["message"] = phase_messages.get(phase, "优化任务正在运行")
                task["updated_at"] = time.time()
                task["current_generation"] = snapshot.get("generation", 0)
                task["archive_size"] = snapshot.get("archive_size", 0)
                task["population_size"] = snapshot.get("population_size", req.population_size)
                task["coarse_pool_size"] = snapshot.get("coarse_pool_size", 0)
                task["feasible_ratio"] = snapshot.get("feasible_ratio", 0.0)
                task["hypervolume"] = snapshot.get("hypervolume", 0.0)
                task["elapsed_s"] = snapshot.get("elapsed_s", 0.0)
                task["total_evaluations"] = snapshot.get("total_evaluations", 0)
                task["approximate_evaluations"] = snapshot.get("approximate_evaluations", 0)
                task["exact_evaluations"] = snapshot.get("exact_evaluations", 0)
                task["bottleneck_machine_ids"] = snapshot.get("bottleneck_machine_ids", [])
                task["history"].append(snapshot)
                task["history"] = task["history"][-50:]

            result = optimizer.run(progress_callback=_progress)
            payload = result.to_dict()
            export_payload = result.to_export_dict()
            if graph_context_diagnostics is not None:
                graph_context_payload = _graph_context_diagnostics_payload(
                    graph_context_diagnostics
                )
                payload["graph_context"] = graph_context_payload
                export_payload["graph_context"] = graph_context_payload
            payload["graph_context_mode"] = graph_context_mode.value
            export_payload["graph_context_mode"] = graph_context_mode.value
            _hybrid_tasks[task_id]["status"] = "done"
            _hybrid_tasks[task_id]["phase"] = "done"
            _hybrid_tasks[task_id]["message"] = "优化完成，方案已可用于评审"
            _hybrid_tasks[task_id]["updated_at"] = time.time()
            _hybrid_tasks[task_id]["result"] = payload
            _hybrid_tasks[task_id]["export_result"] = export_payload
            _hybrid_tasks[task_id]["archive_size"] = payload["archive_size"]
            _hybrid_tasks[task_id]["current_generation"] = payload["generations_completed"]
            _hybrid_tasks[task_id]["elapsed_s"] = payload["elapsed_s"]
            _hybrid_tasks[task_id]["total_evaluations"] = payload["total_evaluations"]
            _hybrid_tasks[task_id]["preview"] = payload["solutions"][:3]
            _latest_hybrid_task_id = task_id
            _save_workflow_progress(
                "optimization", {"task_id": task_id, "task": _hybrid_tasks[task_id]},
            )
        except Exception as exc:
            error_message = str(exc).strip() or "优化器抛出了未提供说明的异常"
            error_trace = traceback.format_exc()
            logging.exception("Hybrid optimization task %s failed", task_id)
            _hybrid_tasks[task_id].update({
                "status": "error",
                "phase": "error",
                "message": "优化任务执行失败",
                "error": error_message,
                "error_type": type(exc).__name__,
                "technical_detail": error_trace[-4000:],
                "updated_at": time.time(),
            })
        finally:
            heartbeat_stop.set()
            heartbeat_thread.join(
                timeout=max(0.1, OPTIMIZE_HEARTBEAT_INTERVAL_S * 2)
            )

    bg.add_task(_run)
    return {"task_id": task_id, "status": "started", "config": req.model_dump()}

@app.get("/api/optimize/hybrid/status/{task_id}")
async def optimize_hybrid_status(task_id: str):
    task = _hybrid_tasks.get(task_id)
    if not task:
        raise HTTPException(404, "任务不存在")
    return {
        "task_id": task_id,
        "status": task["status"],
        "phase": task.get("phase", "initializing"),
        "message": task.get("message", "优化任务正在运行"),
        "created_at": task.get("created_at"),
        "updated_at": task.get("updated_at"),
        "config": task.get("config"),
        "current_generation": task.get("current_generation", 0),
        "archive_size": task.get("archive_size", 0),
        "population_size": task.get("population_size", 0),
        "coarse_pool_size": task.get("coarse_pool_size", 0),
        "feasible_ratio": task.get("feasible_ratio", 0.0),
        "hypervolume": task.get("hypervolume", 0.0),
        "elapsed_s": task.get("elapsed_s", 0.0),
        "total_evaluations": task.get("total_evaluations", 0),
        "approximate_evaluations": task.get("approximate_evaluations", 0),
        "exact_evaluations": task.get("exact_evaluations", 0),
        "bottleneck_machine_ids": task.get("bottleneck_machine_ids", []),
        "history": task.get("history", []),
        "preview": task.get("preview", []),
        "error": task.get("error"),
        "error_type": task.get("error_type"),
        "technical_detail": task.get("technical_detail"),
        "graph_context_mode": task.get("graph_context_mode", "legacy"),
        "graph_context": task.get("graph_context"),
        "graph_context_diff": task.get("graph_context_diff"),
    }

def _hybrid_result_payload(task_id: str, task: dict) -> dict:
    return {
        "task_id": task_id,
        "status": "done",
        **task["result"],
        "reference_solutions": task.get("reference_solutions", []),
    }


@app.get("/api/optimize/hybrid/result/{task_id}")
async def optimize_hybrid_result(task_id: str):
    task = _hybrid_tasks.get(task_id)
    if not task:
        raise HTTPException(404, "任务不存在")
    if task["status"] == "error":
        raise HTTPException(500, task.get("error", "优化任务失败"))
    if task["status"] != "done" or not task.get("result"):
        return {
            "task_id": task_id,
            "status": task["status"],
            "message": "优化仍在进行中",
            "progress": {
                "current_generation": task.get("current_generation", 0),
                "archive_size": task.get("archive_size", 0),
                "feasible_ratio": task.get("feasible_ratio", 0.0),
                "hypervolume": task.get("hypervolume", 0.0),
                "elapsed_s": task.get("elapsed_s", 0.0),
                "graph_context": task.get("graph_context"),
            },
        }
    return _hybrid_result_payload(task_id, task)


@app.get("/api/workflow/progress")
async def workflow_progress():
    """已完成步骤的结果快照，供前端启动时恢复，而不必重跑整条流程。

    只返回与当前库内实例匹配的快照：实例一改，旧的校验和排程结论自动消失，
    前端据此把对应步骤退回“待开始”。
    """
    snapshot = workflow_store.load_all()
    optimization = snapshot.get("optimization") or {}
    task_id, task = optimization.get("task_id"), optimization.get("task") or {}
    return {
        "validation": snapshot.get("validation"),
        "simulation": snapshot.get("simulation"),
        "optimization": (
            _hybrid_result_payload(task_id, task)
            if task_id and task.get("result") else None
        ),
        "review": snapshot.get("review"),
    }


@app.put("/api/workflow/review")
async def save_review_progress(req: ReviewProgressReq):
    """记住评审页选中/对比的方案，重启后回到同样的视图。"""
    workflow_store.save("review", req.model_dump())
    return {"status": "saved"}


@app.post("/api/optimize/hybrid/export-solution")
async def optimize_hybrid_export_solution(req: ExportSolutionReq):
    task_id, task = _resolve_hybrid_task(req.task_id)
    current_shop = _active_shop()
    if current_shop is None:
        raise HTTPException(400, "当前没有可用实例")
    solution = _resolve_export_solution(current_shop, task, req.solution_id)
    payload = _build_solution_export_bytes(current_shop, task_id, task.get("export_result") or task.get("result") or {}, solution)
    safe_solution_id = str(solution.get("solution_id", "solution")).replace(":", "_")
    filename = f"solution_export_{safe_solution_id}.xlsx"
    return Response(
        content=payload,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/ai/pareto/compare")
async def ai_compare_pareto(req: ParetoCompareReq):
    task_id, task = _resolve_hybrid_task(req.task_id)
    result = task["result"]
    current_shop = _active_shop()
    selected = _resolve_candidate_set(current_shop, task, result, req.solution_ids, req.heuristic_rule_names, minimum=2, default_limit=3)
    prompt = _build_compare_prompt(current_shop, task_id, result, selected, req.requirement or "", req.conversation)
    parsed, raw, model_name = _invoke_pareto_llm(prompt, "AI-Pareto", "pareto_compare")
    if not parsed:
        return _heuristic_compare_payload(result, selected, req.requirement or "", task_id)
    return {
        "mode": "compare",
        "task_id": task_id,
        "used_solution_ids": [item.get("solution_id") for item in selected],
        "analysis": parsed,
        "display_text": _format_pareto_display("compare", parsed),
        "llm_raw": raw[:4000],
        "used_model": model_name,
    }


@app.post("/api/ai/pareto/recommend")
async def ai_recommend_pareto(req: ParetoRecommendReq):
    task_id, task = _resolve_hybrid_task(req.task_id)
    result = task["result"]
    current_shop = _active_shop()
    selected = _resolve_candidate_set(current_shop, task, result, req.solution_ids, req.heuristic_rule_names, minimum=1, default_limit=6)
    prompt = _build_recommend_prompt(current_shop, task_id, result, selected, req.requirement, req.conversation)
    parsed, raw, model_name = _invoke_pareto_llm(prompt, "AI-Pareto", "pareto_recommend")
    if not parsed:
        return _heuristic_recommend_payload(result, selected, req.requirement, task_id)
    recommended_id = parsed.get("recommended_solution_id")
    available_ids = {item.get("solution_id") for item in selected}
    if recommended_id not in available_ids and selected:
        parsed["recommended_solution_id"] = selected[0].get("solution_id")
    return {
        "mode": "recommend",
        "task_id": task_id,
        "used_solution_ids": [item.get("solution_id") for item in selected],
        "analysis": parsed,
        "display_text": _format_pareto_display("recommend", parsed),
        "llm_raw": raw[:4000],
        "used_model": model_name,
    }


@app.post("/api/ai/pareto/ask")
async def ai_ask_pareto(req: ParetoAskReq):
    task_id, task = _resolve_hybrid_task(req.task_id)
    result = task["result"]
    current_shop = _active_shop()
    solution = None
    if req.solution_id.startswith("RULE:") and current_shop is not None:
        solution = _rule_reference_solution(current_shop, req.solution_id.split(":", 1)[1], result.get("objective_keys", []))
    else:
        selected = _select_pareto_solutions(result, [req.solution_id], minimum=1, default_limit=1, reference_solutions=task.get("reference_solutions", []))
        solution = selected[0]
    prompt = _build_ask_prompt(current_shop, task_id, result, solution, req.question, req.conversation)
    parsed, raw, model_name = _invoke_pareto_llm(prompt, "AI-Pareto", "pareto_ask")
    if not parsed:
        return _heuristic_ask_payload(solution, req.question, task_id)
    if not parsed.get("solution_id"):
        parsed["solution_id"] = solution.get("solution_id")
    return {
        "mode": "ask",
        "task_id": task_id,
        "used_solution_ids": [solution.get("solution_id")],
        "analysis": parsed,
        "display_text": _format_pareto_display("ask", parsed),
        "llm_raw": raw[:4000],
        "used_model": model_name,
    }

# === Pareto ===
@app.get("/api/pareto/objectives")
async def pareto_objs():
    return {k: {"label": v.label, "direction": v.direction} for k, v in OBJECTIVES.items()}

@app.post("/api/pareto/optimize")
async def pareto(req: ParetoReq):
    if not shop: raise HTTPException(400, "请先生成实例")
    opt = ParetoOptimizer(shop, req.objectives)
    rules = {n: BUILTIN_RULES[n] for n in (req.rule_names or BUILTIN_RULES.keys()) if n in BUILTIN_RULES}
    sols = opt.evaluate(rules)
    return opt.to_frontend(sols)

# === Train ===
@app.post("/api/train")
async def train(req: TrainReq, bg: BackgroundTasks):
    global last_engine
    def _run():
        global last_engine
        gen = InstanceGenerator()
        ts = gen.generate_training_set(req.num_train_instances)
        c = get_config()
        llm = LLMInterface(api_key=c.llm.api_key, base_url=c.llm.base_url, model=c.llm.model)
        engine = EvolutionEngine(EvolutionConfig(population_size=req.population_size, max_generations=req.max_generations), llm)
        engine.initialize(req.seed_rules)
        engine.evolve(ts)
        last_engine = engine
    bg.add_task(_run)
    return {"status": "training_started"}

@app.get("/api/train/logs")
async def train_logs():
    if not last_engine: return {"logs": [], "history": []}
    return {"logs": last_engine.get_llm_logs(), "history": last_engine.history,
            "best": {"fitness": last_engine.best.fitness, "code": last_engine.best.code} if last_engine.best else None}

# === LLM Config ===
@app.get("/api/config/llm")
async def get_llm_cfg():
    c = get_config().llm
    return {"base_url": c.base_url, "model": c.model, "has_key": bool(c.api_key),
            "preview": (c.api_key[:6]+"...") if len(c.api_key)>6 else ""}

@app.put("/api/config/llm")
async def set_llm_cfg(req: LLMCfg):
    cp = os.path.join(os.path.dirname(__file__), "..", "config.json")
    try:
        with open(cp) as f: raw = json.load(f)
    except: raw = {"llm": {}}
    llm = raw.setdefault("llm", {})
    if req.base_url is not None: llm["base_url"] = req.base_url
    if req.api_key is not None: llm["api_key"] = req.api_key
    if req.model is not None: llm["model"] = req.model
    with open(cp, "w") as f: json.dump(raw, f, indent=2, ensure_ascii=False)
    reload_config()
    return {"status": "ok"}

@app.post("/api/config/llm/test")
async def test_llm():
    c = get_config().llm
    if not c.api_key: return {"status": "error", "msg": "未配置API Key"}
    try:
        llm = LLMInterface(c.api_key, c.base_url, c.model)
        r = llm.call("Reply OK.", "test", "test", temp=0.1)
        return {"status": "ok", "response": r[:200], "model": c.model}
    except Exception as e:
        return {"status": "error", "msg": str(e)}

@app.get("/api/health")
async def health():
    return {"status": "ok", "has_instance": shop is not None, "version": "3.0"}


# === Downtime ===
@app.get("/api/downtime")
async def list_downtimes():
    current_shop = shop if shop is not None else (inst_store.build_shopfloor() if inst_store.has_data() else None)
    rows = downtime_store.list_all()
    if current_shop is None:
        return {"downtimes": rows}
    return {"downtimes": [_serialize_downtime_row(current_shop, row) for row in rows]}

@app.post("/api/downtime")
async def add_downtime(req: DowntimeReq):
    global shop
    start_time = _coerce_offset(req.start_time)
    end_time = _coerce_offset(req.end_time)
    if start_time >= end_time:
        raise HTTPException(400, "start_time must be less than end_time")
    new_id = downtime_store.save(req.machine_id, req.downtime_type, start_time, end_time)
    # Refresh shop to pick up new downtime
    if inst_store.has_data():
        shop = inst_store.build_shopfloor()
    return {"status": "ok", "id": new_id}

@app.put("/api/downtime/{dt_id}")
async def update_downtime(dt_id: int, data: dict):
    global shop
    payload = dict(data)
    payload["start_time"] = _coerce_offset(payload["start_time"])
    payload["end_time"] = _coerce_offset(payload["end_time"])
    downtime_store.update(dt_id, payload)
    if inst_store.has_data():
        shop = inst_store.build_shopfloor()
    return {"status": "ok"}

@app.delete("/api/downtime/{dt_id}")
async def delete_downtime(dt_id: int):
    global shop
    downtime_store.delete(dt_id)
    if inst_store.has_data():
        shop = inst_store.build_shopfloor()
    return {"status": "ok"}


# === NSGA-II ===
@app.post("/api/pareto/nsga2")
async def nsga2_start(req: NSGA2Req, bg: BackgroundTasks):
    if not shop:
        raise HTTPException(400, "请先生成实例")
    task_id = str(uuid.uuid4())[:8]
    _nsga2_tasks[task_id] = {"status": "running", "progress": 0, "total": 1, "gen": 0, "result": None}

    def _run():
        def cb(done, total, gen):
            _nsga2_tasks[task_id]["progress"] = done
            _nsga2_tasks[task_id]["total"] = total
            _nsga2_tasks[task_id]["gen"] = gen
        try:
            opt = NSGA2Optimizer(shop, req.objectives, pop_size=req.pop_size,
                                  generations=req.generations, seed=req.seed)
            sols = opt.run(callback=cb)
            # Convert to frontend format
            obj_keys = req.objectives
            pts = []
            for s in sols:
                pt = {"rule": s.rule_name, "rank": s.rank, "crowding": round(s.crowding, 3),
                      "is_pareto": s.rank == 0}
                for k in obj_keys:
                    pt[k] = round(s.objectives.get(k, 0), 4)
                if hasattr(s, 'weights'):
                    pt["weights"] = [round(w, 3) for w in s.weights]
                pts.append(pt)
            _nsga2_tasks[task_id]["result"] = {
                "objectives": obj_keys,
                "solutions": pts,
                "pareto_front": [p for p in pts if p["is_pareto"]],
            }
            _nsga2_tasks[task_id]["status"] = "done"
        except Exception as e:
            _nsga2_tasks[task_id]["status"] = "error"
            _nsga2_tasks[task_id]["error"] = str(e)

    bg.add_task(_run)
    return {"task_id": task_id, "status": "started"}

@app.get("/api/pareto/nsga2/status/{task_id}")
async def nsga2_status(task_id: str):
    t = _nsga2_tasks.get(task_id)
    if not t:
        raise HTTPException(404, "任务不存在")
    return t


# === Exact Solver ===
@app.post("/api/exact/solve")
async def exact_solve(req: ExactReq, bg: BackgroundTasks):
    if not shop:
        raise HTTPException(400, "请先生成实例")
    task_id = str(uuid.uuid4())[:8]
    _exact_tasks[task_id] = {"status": "running", "result": None}

    def _run():
        try:
            solver = ExactSolver(shop, req.objectives, req.time_limit_s)
            result = solver.solve()
            _exact_tasks[task_id]["result"] = result.to_dict()
            _exact_tasks[task_id]["result"]["schedule"] = [_serialize_schedule_entry(shop, entry) for entry in result.schedule[:100]]
            _exact_tasks[task_id]["status"] = "done"
        except Exception as e:
            _exact_tasks[task_id]["status"] = "error"
            _exact_tasks[task_id]["error"] = str(e)

    bg.add_task(_run)
    return {"task_id": task_id, "status": "started"}

@app.get("/api/exact/status/{task_id}")
async def exact_status(task_id: str):
    t = _exact_tasks.get(task_id)
    if not t:
        raise HTTPException(404, "任务不存在")
    return t


# === Online Scheduling ===
@app.post("/api/online/start")
async def online_start(req: OnlineStartReq):
    global online_scheduler_v3, shop
    if not shop and inst_store.has_data():
        shop = inst_store.build_shopfloor()
    if not shop:
        raise HTTPException(400, "请先生成实例")
    online_scheduler_v3 = OnlineSchedulerV3(shop, req.rule_name)
    return {"status": "ok", **online_scheduler_v3.get_status()}

@app.post("/api/online/advance")
async def online_advance(req: OnlineAdvanceReq):
    if not online_scheduler_v3:
        raise HTTPException(400, "请先启动在线调度")
    return online_scheduler_v3.advance(req.delta)

@app.post("/api/online/breakdown")
async def online_breakdown(req: OnlineBreakdownReq):
    if not online_scheduler_v3:
        raise HTTPException(400, "请先启动在线调度")
    return online_scheduler_v3.on_breakdown(req.machine_id, _coerce_offset(req.repair_at))

@app.post("/api/online/repair")
async def online_repair(req: OnlineRepairReq):
    if not online_scheduler_v3:
        raise HTTPException(400, "请先启动在线调度")
    return online_scheduler_v3.on_repair(req.machine_id)

@app.post("/api/online/reschedule")
async def online_reschedule(req: OnlineRescheduleReq):
    if not online_scheduler_v3:
        raise HTTPException(400, "请先启动在线调度")
    return online_scheduler_v3.reschedule(req.rule_name)

@app.get("/api/online/status")
async def online_status():
    if not online_scheduler_v3:
        return {"status": "not_started"}
    return online_scheduler_v3.get_status()


# === Scenario Analysis ===
@app.post("/api/scenario/compare")
async def scenario_compare(req: ScenarioCompareReq):
    if not shop:
        raise HTTPException(400, "请先生成实例")
    rule_names = req.rule_names or list(BUILTIN_RULES.keys())
    results = []
    import copy, random as _rand
    for rule_name in rule_names:
        if rule_name not in BUILTIN_RULES:
            continue
        rule_fn = BUILTIN_RULES[rule_name]
        rep_results = []
        for rep in range(max(1, req.num_replications)):
            sim_shop = copy.deepcopy(shop)
            # Apply pt_variance if specified
            if req.pt_variance > 0:
                rng = _rand.Random(rep * 7919 + hash(rule_name) % 1000)
                for op in sim_shop.operations.values():
                    factor = 1.0 + rng.gauss(0, req.pt_variance)
                    op.processing_time = max(0.1, op.processing_time * factor)
            sim = Simulator(sim_shop, rule_fn)
            r = sim.run()
            rep_results.append(r.to_dict())
        # Aggregate
        keys = ["total_tardiness", "makespan", "avg_utilization", "critical_utilization",
                "main_order_tardy_count", "main_order_tardy_ratio"]
        agg = {}
        for k in keys:
            vals = [rr[k] for rr in rep_results if k in rr]
            if vals:
                agg[k + "_mean"] = round(sum(vals)/len(vals), 3)
                agg[k + "_min"] = round(min(vals), 3)
                agg[k + "_max"] = round(max(vals), 3)
        results.append({"rule": rule_name, "replications": len(rep_results), **agg})
    results.sort(key=lambda x: x.get("total_tardiness_mean", 9e9))
    return {"comparison": results, "num_replications": req.num_replications}


# ============================================================
#  AI 自然语言排产调整 + 排产结果分析
# ============================================================

# 规则描述映射, 用于让LLM理解每条规则的含义
RULE_DESCRIPTIONS = {
    "EDD": "最早交期优先 — 优先调度交期最紧迫的工件, 减少延迟数量",
    "SPT": "最短加工时间 — 优先调度加工时间最短的工件, 减少平均等待时间",
    "LPT": "最长加工时间 — 优先调度加工时间最长的工件, 平衡机器负载",
    "CR": "关键比率 — 综合考虑交期松紧和剩余加工时间, 紧急度高的优先",
    "ATC": "表观延迟成本 — 综合加工时间和松弛度, 平衡延迟和效率",
    "FIFO": "先到先服务 — 按到达顺序处理, 公平但不考虑紧急度",
    "MST": "最小松弛时间 — 优先调度松弛时间最少的工件, 减少延迟风险",
    "PRIORITY": "订单优先级 — 优先调度高优先级订单, 兼顾紧急度",
    "KIT_AWARE": "配套感知 — 前置任务完成度高的优先, 减少装配等待",
    "BOTTLENECK": "瓶颈感知 — 主任务和紧急工件优先, 面向瓶颈资源优化",
    "COMPOSITE": "综合规则 — 加权组合紧急度、优先级、配套比、主任务等多因素",
}

NL_RULE_SELECT_PROMPT = """你是一个工业排产调度专家。用户用自然语言描述了排产目标, 你需要:

1. 分析用户的排产需求和目标
2. 从可用的11条调度规则中, 选择最合适的规则或组合策略
3. 如果需要自定义规则, 生成规则代码

## 可用规则库:
{rules_desc}

## 当前实例概况:
{instance_summary}

## 用户的排产需求:
{user_prompt}

请用以下JSON格式回复:
{{
  "analysis": "对用户需求的分析(中文, 150字内)",
  "thinking_process": "你的推理过程: 为什么选择这个规则, 你是如何将用户的自然语言需求映射到调度规则的(中文, 200字内)",
  "selected_rule": "选择的规则名(如EDD, ATC等, 如果需要自定义则填CUSTOM)",
  "rule_reason": "选择该规则的原因(中文, 100字内)",
  "custom_code": "仅当selected_rule=CUSTOM时填写自定义规则代码, 否则为空",
  "custom_description": "仅当selected_rule=CUSTOM时填写自定义规则的中文描述",
  "parameter_adjustments": "建议的参数调整说明(如有)",
  "tradeoffs": "该方案的权衡说明(中文, 100字内) — 选择此方案会牺牲什么, 获得什么"
}}"""

NL_ANALYSIS_PROMPT = """你是一个工业排产调度分析专家。用户询问了排产结果的问题, 请基于以下数据进行深入分析。

## 当前排产使用的规则:
名称: {rule_name}
描述: {rule_desc}

## 排产KPI汇总:
{kpi_summary}

## 排产明细数据 (按时间排序, 前100条):
{schedule_detail}

## 当前实例订单信息:
{order_info}

## 用户问题:
{user_question}

请进行深入分析并回答用户问题。回复要求:
1. 用中文回答
2. 结构化回答, 使用 Markdown 格式
3. 如果用户问的是某个订单的排产计划, 列出该订单所有工序的排产详情
4. 如果用户问的是为什么某个机器先加工A再加工B, 请从规则逻辑角度解释:
   - 该规则的优先级计算公式是什么
   - 在那个时间点, A和B各自的特征值(交期、优先级、松弛时间等)是什么
   - 根据规则公式, A的得分为什么比B高
5. 如果用户问整体排产情况, 请从多个维度分析
6. 一定要解释清楚大模型是怎么将调度规则转化为具体的排产决策的

请用以下JSON格式回复:
{{
  "answer": "对用户问题的完整回答(Markdown格式, 中文)",
  "rule_logic_explanation": "调度规则如何影响排产决策的解释(中文)",
  "key_insights": ["关键发现1", "关键发现2", "..."],
  "suggestions": ["优化建议1", "优化建议2", "..."]
}}"""


PARETO_COMPARE_PROMPT = """你是一个离散制造调度方案评审专家。请基于给定的帕累托前沿候选方案，帮助业务人员用自然语言比较不同方案。

你的任务：
1. 先围绕 PrimaryObjectives 比较这些方案在本次优化主目标上的表现。
2. 再结合 HolisticKPIs 对全部关键 KPI 做全面比较，不能只盯住主目标。
3. 同时结合 RiskSummary 解释各方案的优势、风险和明显取舍。
4. 如果用户给了额外要求，请结合主目标、全量 KPI 和风险三层信息说明哪个方案更匹配。
5. 只能在给定方案内做判断，不能杜撰不存在的方案。

请只返回 JSON，格式如下：
{{
  "headline": "一句话总结",
  "recommended_solution_id": "如果无法明确推荐则留空字符串",
  "summary_points": ["要点1", "要点2", "要点3"],
  "solution_briefs": [
    {{
      "solution_id": "方案ID",
      "positioning": "该方案适合什么场景",
      "strengths": ["优势1", "优势2"],
      "risks": ["风险1", "风险2"],
      "fit_for": ["适合诉求1", "适合诉求2"]
    }}
  ],
  "final_advice": "给业务的最终建议"
}}
"""


PARETO_RECOMMEND_PROMPT = """你是一个工业排产决策顾问。请根据用户的自然语言要求，从给定的候选方案中选出最匹配的一个方案。

要求：
1. 只能从提供的方案中选择一个 recommended_solution_id。
2. 推荐时必须先看本次 PrimaryObjectives，再全面考虑 HolisticKPIs，最后结合 RiskSummary 做综合判断。
3. 解释为什么这个方案最符合用户要求。
4. 说明该选择的主要收益和主要代价。
5. 如果用户的要求有歧义，请指出需要确认的问题，但仍然给出当前最佳推荐。

请只返回 JSON，格式如下：
{{
  "recommended_solution_id": "方案ID",
  "reason": "推荐原因",
  "fit_points": ["匹配点1", "匹配点2", "匹配点3"],
  "tradeoffs": ["代价1", "代价2"],
  "followup_questions": ["建议确认的问题1", "建议确认的问题2"]
}}
"""


PARETO_ASK_PROMPT = """你是一个工业排产方案解释专家。请围绕指定的一个候选方案，回答用户的问题。

要求：
1. 回答必须结合给定方案的主目标、全量 KPI、规则画像、风险信息和排程片段。
2. 如果用户问规则，就解释 seed_rule、graph_profile 和主要特征权重。
3. 如果用户问过程，就解释关键工序排序和资源分配过程。
4. 如果用户问风险，就指出超期订单、瓶颈设备、等待/流程时间、利用率和潜在代价。
5. 不要回答超出给定数据范围的内容。

请只返回 JSON，格式如下：
{{
  "solution_id": "方案ID",
  "answer": "直接回答用户问题",
  "rule_explanation": "规则与权重的解释",
  "process_explanation": "关键排程过程解释",
  "risk_points": ["风险1", "风险2"],
  "suggestions": ["建议1", "建议2"]
}}
"""


def _format_pareto_display(mode: str, analysis: dict) -> str:
    if mode == "compare":
        lines = [
            analysis.get("headline") or "方案比较完成。",
            "",
            "比较要点：",
            *[f"- {item}" for item in (analysis.get("summary_points") or [])],
        ]
        for item in analysis.get("solution_briefs") or []:
            lines.extend(
                [
                    "",
                    f"方案 {item.get('solution_id')}: {item.get('positioning') or '-'}",
                    f"优势: {', '.join(item.get('strengths') or []) or '-'}",
                    f"风险: {', '.join(item.get('risks') or []) or '-'}",
                    f"适用诉求: {', '.join(item.get('fit_for') or []) or '-'}",
                ]
            )
        if analysis.get("recommended_solution_id"):
            lines.extend(["", f"更贴合当前诉求的方案: {analysis.get('recommended_solution_id')}"])
        if analysis.get("final_advice"):
            lines.extend(["", f"建议: {analysis.get('final_advice')}"])
        return "\n".join(lines)
    if mode == "recommend":
        lines = [
            f"推荐方案: {analysis.get('recommended_solution_id') or '-'}",
            analysis.get("reason") or "",
            "",
            "匹配点：",
            *[f"- {item}" for item in (analysis.get("fit_points") or [])],
            "",
            "主要取舍：",
            *[f"- {item}" for item in (analysis.get("tradeoffs") or [])],
        ]
        followups = analysis.get("followup_questions") or []
        if followups:
            lines.extend(["", "建议进一步确认：", *[f"- {item}" for item in followups]])
        return "\n".join(line for line in lines if line is not None)
    lines = [
        analysis.get("answer") or "方案问答完成。",
    ]
    if analysis.get("rule_explanation"):
        lines.extend(["", "规则解释：", analysis.get("rule_explanation")])
    if analysis.get("process_explanation"):
        lines.extend(["", "过程解释：", analysis.get("process_explanation")])
    risk_points = analysis.get("risk_points") or []
    if risk_points:
        lines.extend(["", "风险点：", *[f"- {item}" for item in risk_points]])
    suggestions = analysis.get("suggestions") or []
    if suggestions:
        lines.extend(["", "建议：", *[f"- {item}" for item in suggestions]])
    return "\n".join(lines)


def _select_pareto_solutions(result: dict, solution_ids: list[str], minimum: int = 1, default_limit: int = 4, reference_solutions: list[dict] | None = None) -> list[dict]:
    solution_index = {item["solution_id"]: item for item in result.get("solutions", [])}
    for item in reference_solutions or []:
        if item.get("solution_id"):
            solution_index[item["solution_id"]] = item
    selected = [solution_index[sid] for sid in solution_ids if sid in solution_index]
    if not selected:
        selected = list(result.get("solutions", []))[:default_limit]
    if len(selected) < minimum:
        raise HTTPException(400, f"至少需要选择 {minimum} 个方案")
    return selected


def _resolve_candidate_set(current_shop: Optional[ShopFloor], task: dict, result: dict, solution_ids: list[str], heuristic_rule_names: list[str], minimum: int = 1, default_limit: int = 4) -> list[dict]:
    reference_solutions = task.get("reference_solutions", []) or []
    candidates: list[dict] = list(_select_pareto_solutions(result, solution_ids, minimum=0, default_limit=default_limit, reference_solutions=reference_solutions))
    if current_shop is not None:
        for rule_name in heuristic_rule_names:
            if rule_name in BUILTIN_RULES:
                candidates.append(_rule_reference_solution(current_shop, rule_name, result.get("objective_keys", [])))
    unique: dict[str, dict] = {}
    for item in candidates:
        unique[item["solution_id"]] = item
    resolved = list(unique.values())
    if not resolved and minimum > 0:
        resolved = list(_select_pareto_solutions(result, [], minimum=minimum, default_limit=default_limit, reference_solutions=reference_solutions))
    if len(resolved) < minimum:
        raise HTTPException(400, f"至少需要选择 {minimum} 个候选方案或启发式规则")
    return resolved


def _resolve_export_solution(current_shop: Optional[ShopFloor], task: dict, solution_id: str) -> dict:
    export_result = task.get("export_result") or task.get("result") or {}
    if solution_id == "BASELINE":
        baseline = export_result.get("baseline")
        if not baseline:
            raise HTTPException(404, "未找到基线方案")
        return baseline
    if solution_id.startswith("RULE:"):
        if current_shop is None:
            raise HTTPException(400, "当前没有可用实例")
        return _rule_reference_solution(
            current_shop,
            solution_id.split(":", 1)[1],
            export_result.get("objective_keys", []),
            schedule_limit=None,
        )
    for item in task.get("reference_solutions", []) or []:
        if item.get("solution_id") == solution_id:
            return item
    for item in export_result.get("reference_solutions", []) or []:
        if item.get("solution_id") == solution_id:
            return item
    for item in export_result.get("solutions", []):
        if item.get("solution_id") == solution_id:
            return item
    raise HTTPException(404, f"未找到方案 {solution_id}")


def _calendar_export_rows(current_shop: ShopFloor, schedule: list[dict]) -> list[dict]:
    touched_machine_ids = {entry.get("machine_id") for entry in schedule if entry.get("machine_id") in current_shop.machines}
    rows: list[dict] = []
    for machine_id in sorted(touched_machine_ids):
        machine = current_shop.machines[machine_id]
        for shift in machine.shifts:
            payload = _shift_payload(current_shop, shift)
            rows.append(
                {
                    "machine_id": machine_id,
                    "machine_name": machine.name,
                    "window_type": "shift_available",
                    "start": payload["start"],
                    "end": payload["end"],
                    "start_at": payload["start_at"],
                    "end_at": payload["end_at"],
                }
            )
        for downtime in machine.downtimes:
            rows.append(
                {
                    "machine_id": machine_id,
                    "machine_name": machine.name,
                    "window_type": f"{downtime.downtime_type}_downtime",
                    "start": round(downtime.start_time, 3),
                    "end": round(downtime.end_time, 3),
                    "start_at": current_shop.time_label(downtime.start_time),
                    "end_at": current_shop.time_label(downtime.end_time),
                }
            )
    return rows


def _schedule_status_label(status: str | None) -> str:
    normalized = str(status or "").strip().lower()
    if normalized == "completed":
        return "已完成"
    if normalized == "in_progress":
        return "进行中"
    return "未来排产"


def _schedule_status_color(status: str | None) -> str:
    normalized = str(status or "").strip().lower()
    if normalized == "completed":
        return "#4b7bec"
    if normalized == "in_progress":
        return "#c7681f"
    return "#1c7c54"


def _completed_initial_export_entries(current_shop: ShopFloor, existing_schedule: list[dict]) -> list[dict]:
    existing_ids = {entry.get("op_id") for entry in existing_schedule if entry.get("op_id")}
    rows: list[dict] = []
    for op in current_shop.operations.values():
        if op.id in existing_ids or op.status != OpStatus.COMPLETED or op.end_time is None:
            continue
        task = current_shop.tasks.get(op.task_id)
        order = current_shop.orders.get(task.order_id) if task else None
        start = op.start_time if op.start_time is not None else max(0.0, float(op.end_time) - float(op.processing_time))
        machine = current_shop.machines.get(op.assigned_machine_id) if op.assigned_machine_id else None
        rows.append(
            {
                "order_id": order.id if order else None,
                "order_name": order.name if order else None,
                "task_id": op.task_id,
                "op_id": op.id,
                "op_name": op.name,
                "machine_id": op.assigned_machine_id or None,
                "machine_name": machine.name if machine else op.assigned_machine_id,
                "tooling_ids": list(op.assigned_tooling_ids or []),
                "personnel_ids": list(op.assigned_personnel_ids or []),
                "start": round(float(start), 3),
                "end": round(float(op.end_time), 3),
                "start_at": current_shop.time_label(start),
                "end_at": current_shop.time_label(op.end_time),
                "duration": round(float(op.processing_time), 3),
                "elapsed_duration": round(max(0.0, float(op.end_time) - float(start)), 3),
                "is_main": bool(task.is_main) if task else False,
                "due_at": current_shop.time_label(task.due_date) if task and task.due_date is not None else None,
                "status": "completed",
                "status_label": "已完成",
            }
        )
    rows.sort(key=lambda entry: (entry.get("start", 0.0), entry.get("machine_id") or "", entry.get("op_id") or ""))
    return rows


def _solution_schedule_with_initial_history(current_shop: ShopFloor, solution: dict) -> list[dict]:
    schedule_entries = [dict(entry) for entry in (solution.get("schedule", []) or [])]
    for entry in schedule_entries:
        status = entry.get("status") or "scheduled"
        entry["status"] = status
        entry["status_label"] = _schedule_status_label(status)
    completed_entries = _completed_initial_export_entries(current_shop, schedule_entries)
    merged = [*completed_entries, *schedule_entries]
    merged.sort(key=lambda entry: (entry.get("start", 0.0), entry.get("machine_id") or "", entry.get("op_id") or ""))
    return merged


def _build_solution_export_bytes(current_shop: ShopFloor, task_id: str, export_result: dict, solution: dict) -> bytes:
    wb = openpyxl.Workbook()
    export_schedule = _solution_schedule_with_initial_history(current_shop, solution)
    ws_summary = wb.active
    ws_summary.title = "summary"
    ws_summary.append(["task_id", task_id])
    ws_summary.append(["solution_id", solution.get("solution_id")])
    ws_summary.append(["source", solution.get("source") or solution.get("rule_name") or "baseline"])
    ws_summary.append(["evaluation_mode", solution.get("evaluation_mode", "exact")])
    ws_summary.append(["plan_start_at", current_shop.time_label(0.0)])
    ws_summary.append(["objective_count", len(export_result.get("objective_keys", []))])
    ws_summary.append([])
    ws_summary.append(["metric", "value"])
    for key, value in (solution.get("objectives") or {}).items():
        ws_summary.append([key, value])
    for key, value in (solution.get("summary") or {}).items():
        if isinstance(value, list):
            ws_summary.append([key, ", ".join(str(item) for item in value)])
        else:
            ws_summary.append([key, value])
    ws_summary.append([])
    ws_summary.append(["status_legend", "meaning"])
    ws_summary.append(["已完成", "计划起点前已经完成的历史工序"])
    ws_summary.append(["进行中", "计划起点时已开工、仍占用资源的工序"])
    ws_summary.append(["未来排产", "本次方案中新安排的未来工序"])

    ws_schedule = wb.create_sheet("schedule")
    ws_schedule.append([
        "order_id", "order_name", "task_id", "op_id", "op_name", "machine_id", "machine_name",
        "tooling_ids", "personnel_ids", "start", "end", "start_at", "end_at", "duration",
        "elapsed_duration", "is_main", "due_at", "status", "status_label",
    ])
    for entry in export_schedule:
        ws_schedule.append([
            entry.get("order_id"),
            entry.get("order_name"),
            entry.get("task_id"),
            entry.get("op_id"),
            entry.get("op_name"),
            entry.get("machine_id"),
            entry.get("machine_name"),
            ",".join(entry.get("tooling_ids", []) or []),
            ",".join(entry.get("personnel_ids", []) or []),
            entry.get("start"),
            entry.get("end"),
            entry.get("start_at"),
            entry.get("end_at"),
            entry.get("duration"),
            entry.get("elapsed_duration"),
            entry.get("is_main"),
            entry.get("due_at"),
            entry.get("status"),
            entry.get("status_label") or _schedule_status_label(entry.get("status")),
        ])

    ws_calendar = wb.create_sheet("machine_calendar")
    ws_calendar.append(["machine_id", "machine_name", "window_type", "start", "end", "start_at", "end_at"])
    for row in _calendar_export_rows(current_shop, export_schedule):
        ws_calendar.append([
            row["machine_id"],
            row["machine_name"],
            row["window_type"],
            row["start"],
            row["end"],
            row["start_at"],
            row["end_at"],
        ])

    ws_rules = wb.create_sheet("rule_profile")
    ws_rules.append(["field", "value"])
    for key, value in (solution.get("candidate") or {}).items():
        if isinstance(value, dict):
            for sub_key, sub_value in value.items():
                ws_rules.append([f"{key}.{sub_key}", sub_value])
        else:
            ws_rules.append([key, value])

    ws_legend = wb.create_sheet("status_legend")
    ws_legend.append(["status", "label", "color", "description"])
    ws_legend.append(["completed", "已完成", _schedule_status_color("completed"), "计划起点前已经完成的历史工序"])
    ws_legend.append(["in_progress", "进行中", _schedule_status_color("in_progress"), "计划起点时已开工、仍占用资源的工序"])
    ws_legend.append(["scheduled", "未来排产", _schedule_status_color("scheduled"), "本次方案中新安排的未来工序"])

    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 36)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _build_compare_prompt(current_shop: Optional[ShopFloor], task_id: str, result: dict, selected: list[dict], requirement: str, conversation: list[dict[str, str]]) -> str:
    catalog_map = _objective_catalog_map()
    primary_objective_keys, full_kpi_keys = _collect_ai_metric_keys(result, selected)
    primary_desc = "\n".join(
        f"- {_ai_metric_label(key, catalog_map)} ({key}): direction={_ai_metric_direction(key, catalog_map)}, description={_ai_metric_description(key, catalog_map)}"
        for key in primary_objective_keys
    )
    full_kpi_desc = "\n".join(
        f"- {_ai_metric_label(key, catalog_map)} ({key}): direction={_ai_metric_direction(key, catalog_map)}, description={_ai_metric_description(key, catalog_map)}"
        for key in full_kpi_keys
    )
    solution_blocks = "\n\n".join(
        _solution_prompt_block(current_shop, solution, primary_objective_keys, full_kpi_keys, catalog_map, schedule_limit=10)
        for solution in selected
    )
    return (
        f"{PARETO_COMPARE_PROMPT}\n\n"
        f"TaskID={task_id}\n"
        f"UserRequirement={requirement or '未提供额外要求，请先做客观对比。'}\n\n"
        f"PrimaryObjectiveCatalog:\n{primary_desc or '- no_primary_objectives'}\n\n"
        f"HolisticKPICatalog:\n{full_kpi_desc or '- no_holistic_kpis'}\n\n"
        f"Baseline:\n{_baseline_prompt_block(current_shop, result.get('baseline', {}), primary_objective_keys, full_kpi_keys, catalog_map)}\n\n"
        f"Solutions:\n{solution_blocks}\n\n"
        f"ConversationHistory:\n{_history_prompt_block(conversation)}"
    )


def _build_recommend_prompt(current_shop: Optional[ShopFloor], task_id: str, result: dict, selected: list[dict], requirement: str, conversation: list[dict[str, str]]) -> str:
    catalog_map = _objective_catalog_map()
    primary_objective_keys, full_kpi_keys = _collect_ai_metric_keys(result, selected)
    primary_desc = "\n".join(
        f"- {_ai_metric_label(key, catalog_map)} ({key}): direction={_ai_metric_direction(key, catalog_map)}, description={_ai_metric_description(key, catalog_map)}"
        for key in primary_objective_keys
    )
    full_kpi_desc = "\n".join(
        f"- {_ai_metric_label(key, catalog_map)} ({key}): direction={_ai_metric_direction(key, catalog_map)}, description={_ai_metric_description(key, catalog_map)}"
        for key in full_kpi_keys
    )
    solution_blocks = "\n\n".join(
        _solution_prompt_block(current_shop, solution, primary_objective_keys, full_kpi_keys, catalog_map, schedule_limit=10)
        for solution in selected
    )
    return (
        f"{PARETO_RECOMMEND_PROMPT}\n\n"
        f"TaskID={task_id}\n"
        f"UserRequirement={requirement}\n\n"
        f"PrimaryObjectiveCatalog:\n{primary_desc or '- no_primary_objectives'}\n\n"
        f"HolisticKPICatalog:\n{full_kpi_desc or '- no_holistic_kpis'}\n\n"
        f"Baseline:\n{_baseline_prompt_block(current_shop, result.get('baseline', {}), primary_objective_keys, full_kpi_keys, catalog_map)}\n\n"
        f"CandidateSolutions:\n{solution_blocks}\n\n"
        f"ConversationHistory:\n{_history_prompt_block(conversation)}"
    )


def _build_ask_prompt(current_shop: Optional[ShopFloor], task_id: str, result: dict, solution: dict, question: str, conversation: list[dict[str, str]]) -> str:
    catalog_map = _objective_catalog_map()
    primary_objective_keys, full_kpi_keys = _collect_ai_metric_keys(result, [solution])
    return (
        f"{PARETO_ASK_PROMPT}\n\n"
        f"TaskID={task_id}\n"
        f"UserQuestion={question}\n\n"
        f"Baseline:\n{_baseline_prompt_block(current_shop, result.get('baseline', {}), primary_objective_keys, full_kpi_keys, catalog_map, schedule_limit=6)}\n\n"
        f"TargetSolution:\n{_solution_prompt_block(current_shop, solution, primary_objective_keys, full_kpi_keys, catalog_map, schedule_limit=18)}\n\n"
        f"ConversationHistory:\n{_history_prompt_block(conversation)}"
    )


def _get_instance_summary():
    """获取当前实例的概要信息"""
    if not shop:
        return "无实例"
    s = shop.summary()
    order_details = []
    for oid, o in list(shop.orders.items())[:20]:
        order_details.append(f"  {oid}: {o.name}, 优先级P{o.priority}, 交期{o.due_date:.1f}h, 释放{o.release_time:.1f}h, 任务数{len(o.task_ids)}")
    return f"""订单数: {s.get('orders',0)}, 任务数: {s.get('tasks',0)}, 工序数: {s.get('operations',0)}, 机器数: {s.get('machines',0)}
订单详情(前20):
""" + "\n".join(order_details)


def _get_order_info():
    """获取订单详细信息"""
    if not shop:
        return "无实例"
    lines = []
    for oid, o in shop.orders.items():
        tasks_info = []
        for tid in o.task_ids:
            t = shop.tasks.get(tid)
            if t:
                ops_str = ", ".join(f"{op.name}({op.process_type},{op.processing_time:.1f}h)" for op in t.operations)
                tasks_info.append(f"    任务{t.name}({'主' if t.is_main else '子'}): [{ops_str}]")
        lines.append(f"  {o.name}(ID:{oid}): P{o.priority}, 交期{o.due_date:.1f}h, 释放{o.release_time:.1f}h")
        lines.extend(tasks_info[:5])
    return "\n".join(lines[:60])


def _get_schedule_detail():
    """获取排产明细文本"""
    if not last_result or not last_result.schedule:
        return "暂无排产数据"
    lines = []
    for e in last_result.schedule[:100]:
        task = shop.tasks.get(e["task_id"]) if shop else None
        order = shop.orders.get(task.order_id) if task and shop else None
        lines.append(
            f"  {e['op_name']} | 机器:{e['machine_name']} | "
            f"时间:{e['start']:.1f}→{e['end']:.1f}h({e['duration']:.1f}h) | "
            f"任务:{e.get('task_id','')} | "
            f"订单:{order.name if order else ''}(P{order.priority if order else '?'}, 交期{order.due_date:.1f}h)"
            if order else
            f"  {e['op_name']} | 机器:{e['machine_name']} | 时间:{e['start']:.1f}→{e['end']:.1f}h"
        )
    return "\n".join(lines)


def _get_instance_summary():
    if not shop:
        return "No active instance."
    summary = shop.summary()
    lines = [
        (
            f"Orders={summary.get('orders', 0)}, Tasks={summary.get('tasks', 0)}, "
            f"Operations={summary.get('operations', 0)}, Machines={summary.get('machines', 0)}, "
            f"Toolings={summary.get('toolings', 0)}, Personnel={summary.get('personnel', 0)}, "
            f"PlanStart={summary.get('plan_start_at')}"
        )
    ]
    for order_id, order in list(shop.orders.items())[:20]:
        lines.append(
            f"{order_id}: {order.name}, priority=P{order.priority}, "
            f"release={shop.time_label(order.release_time)}, due={shop.time_label(order.due_date)}, "
            f"tasks={len(order.task_ids)}"
        )
    return "\n".join(lines)


def _get_order_info():
    if not shop:
        return "No active instance."
    lines = []
    for order_id, order in shop.orders.items():
        lines.append(
            f"{order.name} (ID:{order_id}): priority=P{order.priority}, "
            f"release={shop.time_label(order.release_time)}, due={shop.time_label(order.due_date)}"
        )
        for task_id in order.task_ids[:5]:
            task = shop.tasks.get(task_id)
            if not task:
                continue
            ops_text = ", ".join(
                (
                    f"{op.name}[{op.process_type}, pt={op.processing_time:.1f}h, "
                    f"tool={','.join(op.required_tooling_types) or '-'}, "
                    f"person={','.join(op.required_personnel_skills) or '-'}]"
                )
                for op in task.operations
            )
            task_kind = "main" if task.is_main else "sub"
            lines.append(
                f"  Task {task.name} ({task_kind}): release={shop.time_label(task.release_time)}, "
                f"due={shop.time_label(task.due_date)}, ops=[{ops_text}]"
            )
    return "\n".join(lines[:80])


def _get_schedule_detail():
    if not last_result or not last_result.schedule:
        return "No schedule available."
    lines = []
    for entry in last_result.schedule[:100]:
        task = shop.tasks.get(entry["task_id"]) if shop else None
        order = shop.orders.get(task.order_id) if task and shop else None
        lines.append(
            (
                f"{entry['op_name']} | machine={entry['machine_name']} | "
                f"start={shop.time_label(entry['start']) if shop else entry['start']} | "
                f"end={shop.time_label(entry['end']) if shop else entry['end']} | "
                f"duration={entry['duration']:.1f}h | "
                f"toolings={','.join(entry.get('tooling_ids', [])) or '-'} | "
                f"personnel={','.join(entry.get('personnel_ids', [])) or '-'} | "
                f"order={order.name if order else '-'} | "
                f"priority=P{order.priority if order else '?'} | "
                f"due={shop.time_label(order.due_date) if order and shop else '-'}"
            )
        )
    return "\n".join(lines)


def _get_kpi_summary():
    """获取KPI汇总"""
    if not last_result:
        return "暂无KPI数据"
    d = last_result.to_dict()
    return f"""总延迟: {d.get('total_tardiness', 0)}
Makespan: {d.get('makespan', 0)}
平均延迟: {d.get('avg_tardiness', 0)}
最大延迟: {d.get('max_tardiness', 0)}
延迟工序数: {d.get('tardy_job_count', 0)} / {d.get('total_jobs', 0)}
平均流程时间: {d.get('avg_flowtime', 0)}
主订单延误数: {d.get('main_order_tardy_count', 0)} / {d.get('total_main_orders', 0)}
主订单延误比例: {d.get('main_order_tardy_ratio', 0):.1%}
平均利用率: {d.get('avg_utilization', 0):.1%}
关键资源利用率: {d.get('critical_utilization', 0):.1%}
总等待时间: {d.get('total_wait_time', 0)}
仿真事件数: {d.get('event_count', 0)}"""


@app.post("/api/ai/adjust-schedule")
async def ai_adjust_schedule(req: NLScheduleReq):
    """AI自然语言排产目标调整: 用户用自然语言描述目标, AI选择/生成最优规则并执行排产"""
    global last_result
    if not shop:
        raise HTTPException(400, "请先生成实例")

    c = get_config().llm
    if not c.api_key:
        raise HTTPException(400, "请先配置大模型API Key (⚙️ 大模型配置)")

    llm = LLMInterface(c.api_key, c.base_url, c.model)

    # 构建规则描述
    rules_desc = "\n".join(f"- {k}: {v}" for k, v in RULE_DESCRIPTIONS.items())
    instance_summary = _get_instance_summary()

    prompt = NL_RULE_SELECT_PROMPT.format(
        rules_desc=rules_desc,
        instance_summary=instance_summary,
        user_prompt=req.prompt,
    )

    # 调用LLM进行规则选择
    raw_resp = llm.call(prompt, "AI-Scheduler", "rule_select", temp=0.3)

    # 解析LLM回复
    try:
        # 尝试从回复中提取JSON
        resp_text = raw_resp
        if "```json" in resp_text:
            resp_text = resp_text[resp_text.index("```json") + 7:]
            resp_text = resp_text[:resp_text.index("```")]
        elif "```" in resp_text:
            resp_text = resp_text[resp_text.index("```") + 3:]
            resp_text = resp_text[:resp_text.index("```")]
        ai_result = json.loads(resp_text.strip())
    except (json.JSONDecodeError, ValueError):
        ai_result = {
            "analysis": raw_resp[:300],
            "thinking_process": "LLM返回非标准格式, 使用默认规则ATC",
            "selected_rule": "ATC",
            "rule_reason": "默认选择ATC规则",
            "custom_code": "",
            "custom_description": "",
            "parameter_adjustments": "",
            "tradeoffs": "",
        }

    selected_rule = ai_result.get("selected_rule", "ATC")

    # 执行排产
    if selected_rule == "CUSTOM" and ai_result.get("custom_code"):
        try:
            from ..core.rules import compile_rule_from_code
            func = compile_rule_from_code(ai_result["custom_code"])
            rule_display_name = "AI自定义规则"
        except Exception as e:
            # 自定义规则编译失败, 回退到ATC
            selected_rule = "ATC"
            func = BUILTIN_RULES["ATC"]
            rule_display_name = "ATC (自定义规则编译失败)"
            ai_result["tradeoffs"] = f"自定义规则编译失败({e}), 回退到ATC"
    else:
        if selected_rule not in BUILTIN_RULES:
            selected_rule = "ATC"
        func = BUILTIN_RULES[selected_rule]
        rule_display_name = selected_rule

    sim = Simulator(shop, func)
    r = sim.run()
    r._rule_name = selected_rule
    last_result = r

    # 构建甘特图数据
    gantt = []
    for e in r.schedule:
        task = shop.tasks.get(e["task_id"])
        order = shop.orders.get(task.order_id) if task else None
        gantt.append(
            _serialize_schedule_entry(
                shop,
                {
                    **e,
                    "order_id": order.id if order else "",
                    "order_name": order.name if order else "",
                    "priority": order.priority if order else 1,
                    "due_date": round(order.due_date, 3) if order else 0,
                    "due_at": shop.time_label(order.due_date) if order else None,
                    "is_tardy": (e["end"] > order.due_date) if order else False,
                    "is_main": task.is_main if task else False,
                },
            )
        )

    # 同时运行所有规则做对比, 展示为什么选这个规则更好
    compare_results = []
    for rn, rfn in BUILTIN_RULES.items():
        cr = Simulator(shop, rfn).run()
        compare_results.append({"rule": rn, "total_tardiness": round(cr.total_tardiness, 2),
                                 "makespan": round(cr.makespan, 2),
                                 "avg_utilization": round(cr.avg_utilization, 4),
                                 "main_order_tardy_count": cr.main_order_tardy_count})

    return {
        "status": "ok",
        "ai_analysis": ai_result,
        "selected_rule": rule_display_name,
        "rule_description": RULE_DESCRIPTIONS.get(selected_rule, ai_result.get("custom_description", "")),
        "metrics": r.to_dict(),
        "gantt": gantt,
        "rule_comparison": sorted(compare_results, key=lambda x: x["total_tardiness"]),
        "llm_raw": raw_resp[:2000],
    }


@app.post("/api/ai/analyze-schedule")
async def ai_analyze_schedule(req: NLAnalysisReq):
    """AI自然语言排产分析: 用户提问, AI解读排产结果和工件顺序逻辑"""
    if not shop:
        raise HTTPException(400, "请先生成实例")
    if not last_result:
        raise HTTPException(400, "请先运行仿真排产")

    c = get_config().llm
    if not c.api_key:
        raise HTTPException(400, "请先配置大模型API Key (⚙️ 大模型配置)")

    llm = LLMInterface(c.api_key, c.base_url, c.model)

    # 确定当前使用的规则名
    rule_name = "ATC"  # default
    # 从最近的排产结果猜测规则(如果有的话)
    if hasattr(last_result, '_rule_name'):
        rule_name = last_result._rule_name

    rule_desc = RULE_DESCRIPTIONS.get(rule_name, "综合调度规则")

    prompt = NL_ANALYSIS_PROMPT.format(
        rule_name=rule_name,
        rule_desc=rule_desc,
        kpi_summary=_get_kpi_summary(),
        schedule_detail=_get_schedule_detail(),
        order_info=_get_order_info(),
        user_question=req.question,
    )

    raw_resp = llm.call(prompt, "AI-Analyst", "analyze", temp=0.3)

    # 解析LLM回复
    try:
        resp_text = raw_resp
        if "```json" in resp_text:
            resp_text = resp_text[resp_text.index("```json") + 7:]
            resp_text = resp_text[:resp_text.index("```")]
        elif "```" in resp_text:
            resp_text = resp_text[resp_text.index("```") + 3:]
            resp_text = resp_text[:resp_text.index("```")]
        ai_result = json.loads(resp_text.strip())
    except (json.JSONDecodeError, ValueError):
        ai_result = {
            "answer": raw_resp,
            "rule_logic_explanation": "",
            "key_insights": [],
            "suggestions": [],
        }

    return {
        "status": "ok",
        "analysis": ai_result,
        "rule_used": rule_name,
        "rule_description": rule_desc,
        "llm_raw": raw_resp[:3000],
    }
