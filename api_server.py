"""FastAPI 服务 v3 — 完整后端"""
import os, csv, io, json, inspect, logging, time, uuid, threading
from typing import Optional
from fastapi import FastAPI, HTTPException, BackgroundTasks, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, PlainTextResponse
from pydantic import BaseModel, Field
import openpyxl

from .config import get_config, reload_config
from .models import ShopFloor
from .instance_generator import InstanceGenerator
from .dispatching_rules import BUILTIN_RULES, get_all_rule_names
from .simulator import Simulator
from .llm_evolution import EvolutionEngine, EvolutionConfig, LLMInterface
from .pareto import ParetoOptimizer, OBJECTIVES, NSGA2Optimizer
from .db_manager import init_db, InstanceStore, GraphStore, DowntimeStore
from .heterogeneous_graph import HeterogeneousGraph
from .exact_solver import ExactSolver
from .online_v3 import OnlineSchedulerV3

logging.basicConfig(level=logging.INFO)
app = FastAPI(title="LLM4DRD智能调度平台", version="3.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

FRONT = os.path.join(os.path.dirname(__file__), "frontend")
if os.path.isdir(FRONT):
    app.mount("/static", StaticFiles(directory=FRONT), name="static")

shop: Optional[ShopFloor] = None
last_result = None
last_engine: Optional[EvolutionEngine] = None
inst_store = InstanceStore()
graph_store = GraphStore()
downtime_store = DowntimeStore()
_nsga2_tasks: dict = {}
_exact_tasks: dict = {}
online_scheduler_v3: Optional[OnlineSchedulerV3] = None

@app.on_event("startup")
async def startup():
    init_db()

@app.get("/")
async def index():
    f = os.path.join(FRONT, "index.html")
    return FileResponse(f) if os.path.exists(f) else {"msg": "API running"}

# === Models ===
class GenReq(BaseModel):
    num_orders: int = 10; tasks_per_order_min: int = 2; tasks_per_order_max: int = 5
    ops_per_task_min: int = 2; ops_per_task_max: int = 5
    machines_per_type: int = 3; due_date_factor: float = 1.5
    arrival_spread: float = 0.0; seed: Optional[int] = 42
    day_shift_hours: float = 10; night_shift_hours: float = 8
    schedule_days: int = 30; maintenance_prob: float = 0.05

class SimReq(BaseModel):
    rule_name: str = "ATC"

class ParetoReq(BaseModel):
    objectives: list[str] = ["total_tardiness", "makespan"]
    rule_names: Optional[list[str]] = None

class TrainReq(BaseModel):
    population_size: int = 6; max_generations: int = 5
    num_train_instances: int = 3; seed_rules: Optional[list[str]] = None

class LLMCfg(BaseModel):
    base_url: Optional[str] = None; api_key: Optional[str] = None
    model: Optional[str] = None

class DowntimeReq(BaseModel):
    machine_id: str
    downtime_type: str = "planned"
    start_time: float
    end_time: float

class NSGA2Req(BaseModel):
    objectives: list[str] = ["total_tardiness", "makespan"]
    pop_size: int = 20
    generations: int = 10
    seed: int = 42

class ExactReq(BaseModel):
    objectives: list[str] = ["makespan", "total_tardiness"]
    time_limit_s: int = 60

class OnlineStartReq(BaseModel):
    rule_name: str = "ATC"

class OnlineAdvanceReq(BaseModel):
    delta: float = 1.0

class OnlineBreakdownReq(BaseModel):
    machine_id: str
    repair_at: float

class OnlineRepairReq(BaseModel):
    machine_id: str

class OnlineRescheduleReq(BaseModel):
    rule_name: Optional[str] = None

class ScenarioCompareReq(BaseModel):
    rule_names: Optional[list[str]] = None
    num_replications: int = 1
    breakdown_prob: float = 0.0
    pt_variance: float = 0.0

# === Instance ===
@app.post("/api/instance/generate")
async def gen_instance(req: GenReq):
    global shop, last_result
    gen = InstanceGenerator(seed=req.seed)
    shop = gen.generate(
        num_orders=req.num_orders,
        tasks_per_order=(req.tasks_per_order_min, req.tasks_per_order_max),
        ops_per_task=(req.ops_per_task_min, req.ops_per_task_max),
        machines_per_type=req.machines_per_type,
        due_date_factor=req.due_date_factor,
        arrival_spread=req.arrival_spread,
        day_shift_hours=req.day_shift_hours,
        night_shift_hours=req.night_shift_hours,
        schedule_days=req.schedule_days,
        maintenance_prob=req.maintenance_prob,
    )
    last_result = None
    # 保存到数据库
    inst_store.save_from_shopfloor(shop)
    return {"status": "ok", "summary": shop.summary(), "details": _instance_details(shop)}

@app.get("/api/instance/details")
async def inst_details():
    global shop
    if not shop and inst_store.has_data():
        shop = inst_store.build_shopfloor()
    if not shop:
        raise HTTPException(400, "请先生成实例或导入CSV")
    return _instance_details(shop)

@app.get("/api/instance/csv")
async def export_csv():
    if not shop: raise HTTPException(400, "请先生成实例")
    return PlainTextResponse(shop.to_csv(), media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=instance.csv"})

@app.get("/api/instance/db")
async def get_instance_from_db():
    """从数据库获取实例原始数据(可编辑), 含概览"""
    global shop
    if not inst_store.has_data():
        raise HTTPException(400, "数据库中无实例数据")
    data = inst_store.load_all()
    # 构建 shop 并返回概览
    if not shop:
        shop = inst_store.build_shopfloor()
    data["summary"] = shop.summary()
    data["details"] = _instance_details(shop)
    return data

@app.put("/api/instance/order/{order_id}")
async def update_order(order_id: str, data: dict):
    global shop
    inst_store.update_order(order_id, data)
    shop = inst_store.build_shopfloor()
    return {"status": "ok"}

@app.put("/api/instance/task/{task_id}")
async def update_task(task_id: str, data: dict):
    global shop
    inst_store.update_task(task_id, data)
    shop = inst_store.build_shopfloor()
    return {"status": "ok"}

@app.put("/api/instance/operation/{op_id}")
async def update_operation(op_id: str, data: dict):
    global shop
    inst_store.update_operation(op_id, data)
    shop = inst_store.build_shopfloor()
    return {"status": "ok"}

@app.put("/api/instance/machine/{machine_id}")
async def update_machine(machine_id: str, data: dict):
    global shop
    inst_store.update_machine(machine_id, data)
    shop = inst_store.build_shopfloor()
    return {"status": "ok"}

@app.post("/api/instance/import-excel")
async def import_excel(file: UploadFile = File(...)):
    """导入Excel文件(含5个sheet), 覆盖数据库"""
    global shop
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        def sheet_to_dicts(sheet_name):
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"缺少sheet: {sheet_name}")
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 1:
                return []
            headers = [str(h or '').strip() for h in rows[0]]
            return [{headers[i]: (str(v) if v is not None else '') for i, v in enumerate(row)}
                    for row in rows[1:] if any(v is not None for v in row)]

        orders_rows = sheet_to_dicts("orders")
        tasks_rows = sheet_to_dicts("tasks")
        ops_rows = sheet_to_dicts("operations")
        mt_rows = sheet_to_dicts("machine_types")
        machines_rows = sheet_to_dicts("machines")
        wb.close()

        inst_store.save_from_csv(orders_rows, tasks_rows, ops_rows, mt_rows, machines_rows)
        shop = inst_store.build_shopfloor()

        return {"status": "ok", "summary": shop.summary(), "details": _instance_details(shop)}
    except Exception as e:
        raise HTTPException(400, f"Excel导入失败: {e}")

@app.get("/api/instance/template")
async def download_template():
    """下载Excel模板文件"""
    tpl = os.path.join(os.path.dirname(__file__), "docs", "instance_template.xlsx")
    if not os.path.exists(tpl):
        raise HTTPException(404, "模板文件不存在")
    return FileResponse(tpl, filename="instance_template.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _instance_details(s: ShopFloor):
    orders = []
    for oid, o in s.orders.items():
        tasks_info = []
        for tid in o.task_ids:
            t = s.tasks.get(tid)
            if not t: continue
            ops_info = [{"id": op.id, "name": op.name, "type": op.process_type,
                         "time": op.processing_time, "predecessors": op.predecessor_ops + op.predecessor_tasks}
                        for op in t.operations]
            tasks_info.append({"id": tid, "name": t.name, "is_main": t.is_main,
                               "predecessors": t.predecessor_task_ids, "ops": ops_info})
        orders.append({"id": oid, "name": o.name, "due_date": o.due_date,
                        "priority": o.priority, "release": o.release_time, "tasks": tasks_info})
    machines = []
    for mid, m in s.machines.items():
        mt = s.machine_types.get(m.type_id)
        machines.append({"id": mid, "name": m.name, "type": m.type_id,
                          "type_name": mt.name if mt else "", "is_critical": mt.is_critical if mt else False})
    machine_types = [{"id": k, "name": v.name, "is_critical": v.is_critical,
                       "count": len(s._machine_by_type.get(k, []))}
                      for k, v in s.machine_types.items()]
    return {"orders": orders, "machines": machines, "machine_types": machine_types,
            "summary": s.summary()}

# === Graph ===
@app.post("/api/graph/build")
async def build_graph():
    """从当前实例构建异构图并保存到数据库"""
    global shop
    if not shop and inst_store.has_data():
        shop = inst_store.build_shopfloor()
    if not shop:
        raise HTTPException(400, "请先生成实例")
    hg = HeterogeneousGraph()
    hg.build_from_shopfloor(shop)
    graph_store.save_graph(hg)
    return {"status": "ok", "stats": hg.get_graph_stats()}

@app.get("/api/graph/meta")
async def graph_meta():
    meta = graph_store.load_meta()
    if not meta:
        raise HTTPException(400, "暂无图数据,请先构建图")
    return meta

@app.get("/api/graph/nodes")
async def graph_nodes(node_type: str = None, search: str = None,
                      limit: int = 200, offset: int = 0):
    nodes = graph_store.load_nodes(node_type=node_type, search=search)
    total = len(nodes)
    return {"total": total, "nodes": nodes[offset:offset + limit]}

@app.get("/api/graph/edges")
async def graph_edges(edge_type: str = None, search: str = None,
                      limit: int = 200, offset: int = 0):
    edges = graph_store.load_edges(edge_type=edge_type, search=search)
    total = len(edges)
    return {"total": total, "edges": edges[offset:offset + limit]}

@app.get("/api/graph/node/{node_id:path}/neighbors")
async def node_neighbors(node_id: str):
    return graph_store.get_node_neighbors(node_id)

# === Simulate ===
@app.post("/api/simulate")
async def simulate(req: SimReq):
    global last_result
    if not shop: raise HTTPException(400, "请先生成实例")
    func = BUILTIN_RULES.get(req.rule_name, BUILTIN_RULES["ATC"])
    sim = Simulator(shop, func)
    r = sim.run()
    last_result = r
    # Enrich gantt with order info
    gantt = []
    for e in r.schedule:
        task = shop.tasks.get(e["task_id"])
        order = shop.orders.get(task.order_id) if task else None
        gantt.append({**e,
            "order_id": order.id if order else "",
            "order_name": order.name if order else "",
            "priority": order.priority if order else 1,
            "due_date": round(order.due_date, 1) if order else 0,
            "is_tardy": (e["end"] > order.due_date) if order else False,
            "is_main": task.is_main if task else False,
        })
    return {"metrics": r.to_dict(), "gantt": gantt, "rule": req.rule_name}

@app.post("/api/simulate/compare")
async def compare(rule_names: list[str] = None):
    if not shop: raise HTTPException(400, "请先生成实例")
    names = rule_names or get_all_rule_names()
    results = []
    for n in names:
        if n not in BUILTIN_RULES: continue
        r = Simulator(shop, BUILTIN_RULES[n]).run()
        results.append({"rule": n, "metrics": r.to_dict()})
    results.sort(key=lambda x: x["metrics"]["total_tardiness"])
    return {"comparison": results}

# === Pareto ===
@app.get("/api/pareto/objectives")
async def pareto_objs():
    return {k: {"label": v.label, "direction": v.direction} for k, v in OBJECTIVES.items()}

@app.post("/api/pareto/optimize")
async def pareto(req: ParetoReq):
    if not shop: raise HTTPException(400, "请先生成实例")
    opt = ParetoOptimizer(shop, req.objectives)
    rules = {n: BUILTIN_RULES[n] for n in (req.rule_names or BUILTIN_RULES.keys()) if n in BUILTIN_RULES}
    sols = opt.evaluate(rules)
    return opt.to_frontend(sols)

# === Train ===
@app.post("/api/train")
async def train(req: TrainReq, bg: BackgroundTasks):
    global last_engine
    def _run():
        global last_engine
        gen = InstanceGenerator()
        ts = gen.generate_training_set(req.num_train_instances)
        c = get_config()
        llm = LLMInterface(api_key=c.llm.api_key, base_url=c.llm.base_url, model=c.llm.model)
        engine = EvolutionEngine(EvolutionConfig(population_size=req.population_size, max_generations=req.max_generations), llm)
        engine.initialize(req.seed_rules)
        engine.evolve(ts)
        last_engine = engine
    bg.add_task(_run)
    return {"status": "training_started"}

@app.get("/api/train/logs")
async def train_logs():
    if not last_engine: return {"logs": [], "history": []}
    return {"logs": last_engine.get_llm_logs(), "history": last_engine.history,
            "best": {"fitness": last_engine.best.fitness, "code": last_engine.best.code} if last_engine.best else None}

# === LLM Config ===
@app.get("/api/config/llm")
async def get_llm_cfg():
    c = get_config().llm
    return {"base_url": c.base_url, "model": c.model, "has_key": bool(c.api_key),
            "preview": (c.api_key[:6]+"...") if len(c.api_key)>6 else ""}

@app.put("/api/config/llm")
async def set_llm_cfg(req: LLMCfg):
    cp = os.path.join(os.path.dirname(__file__), "config.json")
    try:
        with open(cp) as f: raw = json.load(f)
    except: raw = {"llm": {}}
    llm = raw.setdefault("llm", {})
    if req.base_url is not None: llm["base_url"] = req.base_url
    if req.api_key is not None: llm["api_key"] = req.api_key
    if req.model is not None: llm["model"] = req.model
    with open(cp, "w") as f: json.dump(raw, f, indent=2, ensure_ascii=False)
    reload_config()
    return {"status": "ok"}

@app.post("/api/config/llm/test")
async def test_llm():
    c = get_config().llm
    if not c.api_key: return {"status": "error", "msg": "未配置API Key"}
    try:
        llm = LLMInterface(c.api_key, c.base_url, c.model)
        r = llm.call("Reply OK.", "test", "test", temp=0.1)
        return {"status": "ok", "response": r[:200], "model": c.model}
    except Exception as e:
        return {"status": "error", "msg": str(e)}

@app.get("/api/health")
async def health():
    return {"status": "ok", "has_instance": shop is not None, "version": "3.0"}


# === Downtime ===
@app.get("/api/downtime")
async def list_downtimes():
    return {"downtimes": downtime_store.list_all()}

@app.post("/api/downtime")
async def add_downtime(req: DowntimeReq):
    global shop
    if req.start_time >= req.end_time:
        raise HTTPException(400, "start_time must be less than end_time")
    new_id = downtime_store.save(req.machine_id, req.downtime_type, req.start_time, req.end_time)
    # Refresh shop to pick up new downtime
    if inst_store.has_data():
        shop = inst_store.build_shopfloor()
    return {"status": "ok", "id": new_id}

@app.put("/api/downtime/{dt_id}")
async def update_downtime(dt_id: int, data: dict):
    global shop
    downtime_store.update(dt_id, data)
    if inst_store.has_data():
        shop = inst_store.build_shopfloor()
    return {"status": "ok"}

@app.delete("/api/downtime/{dt_id}")
async def delete_downtime(dt_id: int):
    global shop
    downtime_store.delete(dt_id)
    if inst_store.has_data():
        shop = inst_store.build_shopfloor()
    return {"status": "ok"}


# === NSGA-II ===
@app.post("/api/pareto/nsga2")
async def nsga2_start(req: NSGA2Req, bg: BackgroundTasks):
    if not shop:
        raise HTTPException(400, "请先生成实例")
    task_id = str(uuid.uuid4())[:8]
    _nsga2_tasks[task_id] = {"status": "running", "progress": 0, "total": 1, "gen": 0, "result": None}

    def _run():
        def cb(done, total, gen):
            _nsga2_tasks[task_id]["progress"] = done
            _nsga2_tasks[task_id]["total"] = total
            _nsga2_tasks[task_id]["gen"] = gen
        try:
            opt = NSGA2Optimizer(shop, req.objectives, pop_size=req.pop_size,
                                  generations=req.generations, seed=req.seed)
            sols = opt.run(callback=cb)
            # Convert to frontend format
            obj_keys = req.objectives
            pts = []
            for s in sols:
                pt = {"rule": s.rule_name, "rank": s.rank, "crowding": round(s.crowding, 3),
                      "is_pareto": s.rank == 0}
                for k in obj_keys:
                    pt[k] = round(s.objectives.get(k, 0), 4)
                if hasattr(s, 'weights'):
                    pt["weights"] = [round(w, 3) for w in s.weights]
                pts.append(pt)
            _nsga2_tasks[task_id]["result"] = {
                "objectives": obj_keys,
                "solutions": pts,
                "pareto_front": [p for p in pts if p["is_pareto"]],
            }
            _nsga2_tasks[task_id]["status"] = "done"
        except Exception as e:
            _nsga2_tasks[task_id]["status"] = "error"
            _nsga2_tasks[task_id]["error"] = str(e)

    bg.add_task(_run)
    return {"task_id": task_id, "status": "started"}

@app.get("/api/pareto/nsga2/status/{task_id}")
async def nsga2_status(task_id: str):
    t = _nsga2_tasks.get(task_id)
    if not t:
        raise HTTPException(404, "任务不存在")
    return t


# === Exact Solver ===
@app.post("/api/exact/solve")
async def exact_solve(req: ExactReq, bg: BackgroundTasks):
    if not shop:
        raise HTTPException(400, "请先生成实例")
    task_id = str(uuid.uuid4())[:8]
    _exact_tasks[task_id] = {"status": "running", "result": None}

    def _run():
        try:
            solver = ExactSolver(shop, req.objectives, req.time_limit_s)
            result = solver.solve()
            _exact_tasks[task_id]["result"] = result.to_dict()
            _exact_tasks[task_id]["result"]["schedule"] = result.schedule[:100]
            _exact_tasks[task_id]["status"] = "done"
        except Exception as e:
            _exact_tasks[task_id]["status"] = "error"
            _exact_tasks[task_id]["error"] = str(e)

    bg.add_task(_run)
    return {"task_id": task_id, "status": "started"}

@app.get("/api/exact/status/{task_id}")
async def exact_status(task_id: str):
    t = _exact_tasks.get(task_id)
    if not t:
        raise HTTPException(404, "任务不存在")
    return t


# === Online Scheduling ===
@app.post("/api/online/start")
async def online_start(req: OnlineStartReq):
    global online_scheduler_v3, shop
    if not shop and inst_store.has_data():
        shop = inst_store.build_shopfloor()
    if not shop:
        raise HTTPException(400, "请先生成实例")
    online_scheduler_v3 = OnlineSchedulerV3(shop, req.rule_name)
    return {"status": "ok", **online_scheduler_v3.get_status()}

@app.post("/api/online/advance")
async def online_advance(req: OnlineAdvanceReq):
    if not online_scheduler_v3:
        raise HTTPException(400, "请先启动在线调度")
    return online_scheduler_v3.advance(req.delta)

@app.post("/api/online/breakdown")
async def online_breakdown(req: OnlineBreakdownReq):
    if not online_scheduler_v3:
        raise HTTPException(400, "请先启动在线调度")
    return online_scheduler_v3.on_breakdown(req.machine_id, req.repair_at)

@app.post("/api/online/repair")
async def online_repair(req: OnlineRepairReq):
    if not online_scheduler_v3:
        raise HTTPException(400, "请先启动在线调度")
    return online_scheduler_v3.on_repair(req.machine_id)

@app.post("/api/online/reschedule")
async def online_reschedule(req: OnlineRescheduleReq):
    if not online_scheduler_v3:
        raise HTTPException(400, "请先启动在线调度")
    return online_scheduler_v3.reschedule(req.rule_name)

@app.get("/api/online/status")
async def online_status():
    if not online_scheduler_v3:
        return {"status": "not_started"}
    return online_scheduler_v3.get_status()


# === Scenario Analysis ===
@app.post("/api/scenario/compare")
async def scenario_compare(req: ScenarioCompareReq):
    if not shop:
        raise HTTPException(400, "请先生成实例")
    rule_names = req.rule_names or list(BUILTIN_RULES.keys())
    results = []
    import copy, random as _rand
    for rule_name in rule_names:
        if rule_name not in BUILTIN_RULES:
            continue
        rule_fn = BUILTIN_RULES[rule_name]
        rep_results = []
        for rep in range(max(1, req.num_replications)):
            sim_shop = copy.deepcopy(shop)
            # Apply pt_variance if specified
            if req.pt_variance > 0:
                rng = _rand.Random(rep * 7919 + hash(rule_name) % 1000)
                for op in sim_shop.operations.values():
                    factor = 1.0 + rng.gauss(0, req.pt_variance)
                    op.processing_time = max(0.1, op.processing_time * factor)
            sim = Simulator(sim_shop, rule_fn)
            r = sim.run()
            rep_results.append(r.to_dict())
        # Aggregate
        keys = ["total_tardiness", "makespan", "avg_utilization", "critical_utilization",
                "main_order_tardy_count", "main_order_tardy_ratio"]
        agg = {}
        for k in keys:
            vals = [rr[k] for rr in rep_results if k in rr]
            if vals:
                agg[k + "_mean"] = round(sum(vals)/len(vals), 3)
                agg[k + "_min"] = round(min(vals), 3)
                agg[k + "_max"] = round(max(vals), 3)
        results.append({"rule": rule_name, "replications": len(rep_results), **agg})
    results.sort(key=lambda x: x.get("total_tardiness_mean", 9e9))
    return {"comparison": results, "num_replications": req.num_replications}
