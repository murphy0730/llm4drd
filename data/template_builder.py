from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Font, PatternFill


HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
HEADER_FONT = Font(bold=True)
TEMPLATE_VERSION = "2026.03.26.1"


def _add_sheet(workbook, name: str, headers: list[str], rows: list[list]):
    ws = workbook.create_sheet(title=name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for column_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width, 12), 42)
    ws.freeze_panes = "A2"
    return ws


def build_instance_template_bytes() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _add_sheet(
        wb,
        "planning_context",
        ["template_version", "plan_start_at", "note"],
        [
            [
                TEMPLATE_VERSION,
                "2026-03-25T08:00:00+08:00",
                "All time fields may be hour offsets from plan_start_at, or ISO 8601 datetimes.",
            ]
        ],
    )
    _add_sheet(
        wb,
        "orders",
        ["order_id", "order_name", "release_time", "due_date", "priority"],
        [
            ["ORD-0001", "Order-1", 0, 150, 3],
            ["ORD-0002", "Order-2", 5, 200, 2],
        ],
    )
    _add_sheet(
        wb,
        "tasks",
        ["task_id", "order_id", "task_name", "is_main", "predecessor_task_ids", "release_time", "due_date"],
        [
            ["T-0001-01", "ORD-0001", "Part-1", "N", "", "", ""],
            ["T-0001-MAIN", "ORD-0001", "Assembly-main", "Y", "T-0001-01", "", ""],
        ],
    )
    _add_sheet(
        wb,
        "operations",
        [
            "op_id",
            "task_id",
            "op_name",
            "process_type",
            "processing_time_hrs",
            "predecessor_ops",
            "predecessor_tasks",
            "eligible_machine_ids",
            "required_tooling_types",
            "required_personnel_skills",
        ],
        [
            ["OP-0001-01-01", "T-0001-01", "Turning", "turning", 5.5, "", "", "turning_1;turning_2", "tool_turning", "skill_turning"],
            ["OP-0001-01-02", "T-0001-01", "Milling", "milling", 3.2, "OP-0001-01-01", "", "milling_1", "tool_milling", "skill_milling"],
            ["OP-0001-ASM", "T-0001-MAIN", "Assembly", "assembly", 6, "", "T-0001-01", "assembly_1", "tool_assembly", "skill_assembly"],
        ],
    )
    _add_sheet(
        wb,
        "initial_state",
        [
            "op_id",
            "initial_status",
            "initial_start_time",
            "initial_end_time",
            "initial_remaining_processing_time",
            "initial_assigned_machine_id",
            "initial_assigned_tooling_ids",
            "initial_assigned_personnel_ids",
            "note",
        ],
        [
            ["OP-0001-01-01", "completed", "", 0, 0, "", "", "", "Completed before the planning anchor"],
            ["OP-0001-ASM", "processing", -2, 6, "", "assembly_1", "TL-assembly-01", "PS-assembly-01", "Currently running and occupying resources until hour 6"],
        ],
    )
    _add_sheet(
        wb,
        "machine_types",
        ["type_id", "type_name", "is_critical"],
        [
            ["turning", "Turning", "Y"],
            ["milling", "Milling", "Y"],
            ["assembly", "Assembly", "Y"],
        ],
    )
    _add_sheet(
        wb,
        "machines",
        ["machine_id", "machine_name", "type_id", "shifts"],
        [
            ["turning_1", "Turning-1", "turning", "0/8/10;0/20/8"],
            ["milling_1", "Milling-1", "milling", "0/8/10"],
            ["assembly_1", "Assembly-1", "assembly", "0/8/10"],
        ],
    )
    _add_sheet(
        wb,
        "tooling_types",
        ["type_id", "type_name"],
        [
            ["tool_turning", "Turning tooling"],
            ["tool_milling", "Milling tooling"],
            ["tool_assembly", "Assembly tooling"],
        ],
    )
    _add_sheet(
        wb,
        "toolings",
        ["tooling_id", "tooling_name", "type_id", "shifts"],
        [
            ["TL-turning-01", "Turning-tool-1", "tool_turning", "0/8/10;0/20/8"],
            ["TL-milling-01", "Milling-tool-1", "tool_milling", "0/8/10"],
            ["TL-assembly-01", "Assembly-tool-1", "tool_assembly", "0/8/10"],
        ],
    )
    _add_sheet(
        wb,
        "personnel",
        ["personnel_id", "personnel_name", "skills", "shifts"],
        [
            ["PS-turning-01", "Turning-operator-1", "skill_turning", "0/8/10;0/20/8"],
            ["PS-milling-01", "Milling-operator-1", "skill_milling", "0/8/10"],
            ["PS-assembly-01", "Assembly-operator-1", "skill_assembly", "0/8/10"],
        ],
    )
    _add_sheet(
        wb,
        "downtimes",
        ["machine_id", "downtime_type", "start_time", "end_time"],
        [
            ["turning_1", "planned", 12, 18],
            ["milling_1", "unplanned", 30, 34],
        ],
    )
    _add_sheet(
        wb,
        "instructions",
        ["item", "description"],
        [
            ["template_version", f"Current template version: {TEMPLATE_VERSION}"],
            ["time fields", "Orders, tasks, downtimes and initial_state time fields can be hour offsets from plan_start_at, or ISO 8601 datetimes."],
            ["task due_date", "Task due_date may be blank and will inherit the parent order due_date during import."],
            ["task release_time", "Task release_time may be blank and will inherit the parent order release_time during import."],
            ["derived internal targets", "After generation or import, the system derives internal latest start/finish targets from order due dates, BOM/task dependencies and operation precedence."],
            ["initial_state", "Use the initial_state sheet to describe the shop-floor state at the planning anchor: pending, ready, processing or completed."],
            ["processing state", "For processing rows, fill at least initial_assigned_machine_id. You may provide remaining productive hours, or provide initial_end_time to express that the resources are occupied until that time."],
            ["completed state", "Completed operations are treated as already done before the simulation/optimization starts, so downstream work can start immediately when other dependencies are satisfied."],
            ["shifts", "Shift format is day/start_hour/hours; multiple shifts are separated by semicolons, for example 0/8/10;0/20/8."],
            ["skills", "Personnel skills support multiple values separated by semicolons, for example skill_turning;skill_milling."],
            ["eligible_machine_ids", "Multiple machine IDs are separated by semicolons; if blank, machines are matched by process_type."],
            ["required resources", "required_tooling_types and required_personnel_skills both support multiple semicolon-separated values."],
            ["resource sheets", "tooling_types, toolings, personnel, downtimes and initial_state are the key resource/state sheets in this template."],
        ],
    )

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
