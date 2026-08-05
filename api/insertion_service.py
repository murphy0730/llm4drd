from __future__ import annotations

import copy
import csv
import io
import math
import time
import uuid
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from itertools import product
from typing import Iterable

import openpyxl
from openpyxl.styles import Font, PatternFill

from ..core.models import Operation, OpStatus, Order, ShopFloor, Task
from ..core.rules import BUILTIN_RULES
from ..core.sim_runtime import SimulationRuntime
from ..core.simulator import SimResult, Simulator, _joint_compute_effective_end, _joint_next_available_time
from ..optimization.objectives import ScheduleAnalytics, build_schedule_analytics


class InsertionError(ValueError):
    pass


@dataclass(frozen=True)
class NewOrder:
    order_id: str
    order_name: str
    release_time: float
    expected_due_date: float
    main_task_id: str


@dataclass(frozen=True)
class NewOperation:
    order_id: str
    op_id: str
    task_id: str
    op_name: str
    process_type: str
    processing_time: float
    turnover_time: float
    predecessor_ops: tuple[str, ...]
    predecessor_tasks: tuple[str, ...]
    eligible_machine_ids: tuple[str, ...]
    required_tooling_types: tuple[str, ...]
    required_personnel_skills: tuple[str, ...]


@dataclass
class _FrozenState:
    ends: dict[str, float] = field(default_factory=dict)
    entries: list[dict] = field(default_factory=list)
    added_occupancy: dict[tuple[str, str], list[tuple[float, float]]] = field(default_factory=dict)


@dataclass
class InsertionRun:
    run_id: str
    instance_version: int
    created_at: str
    payload: dict
    full_schedule: list[dict]


class InsertionRunStore:
    def __init__(self, max_runs: int = 8):
        self.max_runs = max(1, int(max_runs))
        self._runs: OrderedDict[str, InsertionRun] = OrderedDict()

    def put(self, run: InsertionRun) -> None:
        self._runs[run.run_id] = run
        self._runs.move_to_end(run.run_id)
        while len(self._runs) > self.max_runs:
            self._runs.popitem(last=False)

    def get(self, run_id: str, instance_version: int) -> InsertionRun:
        run = self._runs.get(run_id)
        if run is None:
            raise InsertionError("插单评估结果不存在或已被淘汰，请重新计算")
        if run.instance_version != instance_version:
            self._runs.pop(run_id, None)
            raise InsertionError("当前实例已发生变化，插单评估结果已失效，请重新计算")
        self._runs.move_to_end(run_id)
        return run

    def clear(self) -> None:
        self._runs.clear()


def _tokens(value) -> tuple[str, ...]:
    if isinstance(value, (list, tuple)):
        return tuple(str(item).strip() for item in value if str(item).strip())
    text = str(value or "").replace("，", ";").replace(",", ";")
    return tuple(token.strip() for token in text.split(";") if token.strip())


def _finite_number(value, label: str, *, minimum: float | None = None) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise InsertionError(f"{label}不是有效数字") from exc
    if not math.isfinite(number) or (minimum is not None and number < minimum):
        raise InsertionError(f"{label}必须是不小于 {minimum:g} 的有限数")
    return number


def _normalize_input(shop: ShopFloor, raw_orders: Iterable[dict], raw_operations: Iterable[dict]):
    order_rows = list(raw_orders or [])
    operation_rows = list(raw_operations or [])
    if not order_rows:
        raise InsertionError("请至少录入一个新订单")
    if not operation_rows:
        raise InsertionError("请至少录入一道新工序")
    if len(order_rows) > 20:
        raise InsertionError("单次最多支持 20 个新订单")
    if len(operation_rows) > 120:
        raise InsertionError("单次最多支持 120 道新工序")

    raw_by_order: dict[str, dict] = {}
    for index, raw in enumerate(order_rows, 1):
        order_id = str(raw.get("order_id") or "").strip()
        if not order_id:
            raise InsertionError(f"订单表第 {index} 行缺少 order_id")
        if order_id in raw_by_order:
            raise InsertionError(f"新订单 ID 重复：{order_id}")
        if order_id in shop.orders:
            raise InsertionError(f"订单 ID 已存在于当前实例：{order_id}")
        raw_by_order[order_id] = raw

    operations: list[NewOperation] = []
    seen_ops: set[str] = set()
    task_orders: dict[str, str] = {}
    tasks_by_order: dict[str, set[str]] = {order_id: set() for order_id in raw_by_order}
    for index, raw in enumerate(operation_rows, 1):
        order_id = str(raw.get("order_id") or "").strip()
        op_id = str(raw.get("op_id") or "").strip()
        task_id = str(raw.get("task_id") or "").strip()
        process_type = str(raw.get("process_type") or "").strip()
        if order_id not in raw_by_order:
            raise InsertionError(f"工序表第 {index} 行的 order_id 未出现在订单表：{order_id or '-'}")
        if not op_id or not task_id:
            raise InsertionError(f"工序表第 {index} 行必须填写 op_id 和 task_id")
        if op_id in seen_ops or op_id in shop.operations:
            raise InsertionError(f"工序 ID 重复或已存在：{op_id}")
        existing_task_order = task_orders.setdefault(task_id, order_id)
        if existing_task_order != order_id or task_id in shop.tasks:
            raise InsertionError(f"任务令 ID 必须全局唯一且只能属于一个订单：{task_id}")
        processing_time = _finite_number(
            raw.get("processing_time_hrs", raw.get("processing_time")),
            f"工序 {op_id} 的加工时间",
            minimum=0.001,
        )
        turnover_time = _finite_number(
            raw.get("turnover_time_hrs", raw.get("turnover_time", 0)) or 0,
            f"工序 {op_id} 的周转时间",
            minimum=0.0,
        )
        machine_ids = _tokens(raw.get("eligible_machine_ids"))
        if machine_ids:
            missing = [machine_id for machine_id in machine_ids if machine_id not in shop.machines]
            if missing:
                raise InsertionError(f"工序 {op_id} 引用了不存在的机器：{';'.join(missing)}")
        elif not process_type or not shop.get_machines_for_type(process_type):
            raise InsertionError(f"工序 {op_id} 未指定有效机器，且工艺类型 {process_type or '-'} 没有匹配机器")
        tooling_types = _tokens(raw.get("required_tooling_types"))
        personnel_skills = _tokens(raw.get("required_personnel_skills"))
        for tooling_type in tooling_types:
            if not shop.get_toolings_for_type(tooling_type):
                raise InsertionError(f"工序 {op_id} 所需工装类型没有可用工装：{tooling_type}")
        for skill in personnel_skills:
            if not shop.get_personnel_for_skill(skill):
                raise InsertionError(f"工序 {op_id} 所需人员技能没有可用人员：{skill}")
        seen_ops.add(op_id)
        tasks_by_order[order_id].add(task_id)
        operations.append(NewOperation(
            order_id=order_id,
            op_id=op_id,
            task_id=task_id,
            op_name=str(raw.get("op_name") or op_id).strip(),
            process_type=process_type,
            processing_time=processing_time,
            turnover_time=turnover_time,
            predecessor_ops=_tokens(raw.get("predecessor_ops")),
            predecessor_tasks=_tokens(raw.get("predecessor_tasks")),
            eligible_machine_ids=machine_ids,
            required_tooling_types=tooling_types,
            required_personnel_skills=personnel_skills,
        ))

    operations_by_id = {operation.op_id: operation for operation in operations}
    for operation in operations:
        for predecessor_id in operation.predecessor_ops:
            predecessor = operations_by_id.get(predecessor_id)
            if predecessor is None:
                raise InsertionError(f"工序 {operation.op_id} 的前置工序不存在：{predecessor_id}")
            if predecessor.order_id != operation.order_id:
                raise InsertionError(f"不支持跨订单前置关系：{predecessor_id} → {operation.op_id}")
        for predecessor_task in operation.predecessor_tasks:
            if predecessor_task not in tasks_by_order[operation.order_id]:
                raise InsertionError(f"工序 {operation.op_id} 的前置任务不存在或跨订单：{predecessor_task}")

    orders: list[NewOrder] = []
    for order_id, raw in raw_by_order.items():
        release = _finite_number(raw.get("release_time", 0) or 0, f"订单 {order_id} 的放行时间")
        due_raw = raw.get("expected_due_date")
        due = float("inf") if due_raw in (None, "") else _finite_number(due_raw, f"订单 {order_id} 的期望交期")
        if math.isfinite(due) and due < release:
            raise InsertionError(f"订单 {order_id} 的期望交期不能早于放行时间")
        main_task_id = str(raw.get("main_task_id") or "").strip()
        task_ids = sorted(tasks_by_order[order_id])
        if not main_task_id and len(task_ids) == 1:
            main_task_id = task_ids[0]
        if not main_task_id or main_task_id not in tasks_by_order[order_id]:
            raise InsertionError(f"订单 {order_id} 含多个任务令时必须指定有效的 main_task_id")
        orders.append(NewOrder(
            order_id=order_id,
            order_name=str(raw.get("order_name") or order_id).strip(),
            release_time=release,
            expected_due_date=due,
            main_task_id=main_task_id,
        ))
    dependencies = _dependency_map(operations)
    return orders, operations, dependencies


def _dependency_map(operations: list[NewOperation]) -> dict[str, set[str]]:
    by_task: dict[str, set[str]] = {}
    for operation in operations:
        by_task.setdefault(operation.task_id, set()).add(operation.op_id)
    dependencies: dict[str, set[str]] = {}
    for operation in operations:
        deps = set(operation.predecessor_ops)
        for task_id in operation.predecessor_tasks:
            deps.update(by_task.get(task_id, set()))
        deps.discard(operation.op_id)
        dependencies[operation.op_id] = deps
    pending = {op_id: set(deps) for op_id, deps in dependencies.items()}
    resolved: set[str] = set()
    while pending:
        ready = {op_id for op_id, deps in pending.items() if deps <= resolved}
        if not ready:
            raise InsertionError(f"新工序存在循环依赖：{', '.join(sorted(pending)[:10])}")
        resolved.update(ready)
        for op_id in ready:
            pending.pop(op_id)
    return dependencies


def _augment_shop(shop: ShopFloor, orders: list[NewOrder], operations: list[NewOperation]) -> ShopFloor:
    augmented = copy.deepcopy(shop)
    order_map = {order.order_id: order for order in orders}
    for new_order in orders:
        augmented.orders[new_order.order_id] = Order(
            id=new_order.order_id,
            name=new_order.order_name,
            release_time=new_order.release_time,
            due_date=new_order.expected_due_date,
            priority=1,
            main_task_id=new_order.main_task_id,
        )
    for operation in operations:
        task = augmented.tasks.get(operation.task_id)
        if task is None:
            order = order_map[operation.order_id]
            task = Task(
                id=operation.task_id,
                order_id=operation.order_id,
                name=operation.task_id,
                is_main=operation.task_id == order.main_task_id,
                release_time=order.release_time,
                due_date=order.expected_due_date,
            )
            augmented.tasks[task.id] = task
            augmented.orders[operation.order_id].task_ids.append(task.id)
        op = Operation(
            id=operation.op_id,
            task_id=operation.task_id,
            name=operation.op_name,
            process_type=operation.process_type,
            processing_time=operation.processing_time,
            turnover_time=operation.turnover_time,
            predecessor_ops=list(operation.predecessor_ops),
            predecessor_tasks=list(operation.predecessor_tasks),
            eligible_machine_ids=list(operation.eligible_machine_ids),
            required_tooling_types=list(operation.required_tooling_types),
            required_personnel_skills=list(operation.required_personnel_skills),
        )
        augmented.operations[op.id] = op
        task.operations.append(op)
    augmented.build_indexes()
    augmented.ensure_calendar_capacity(min_days=max(augmented.calendar_days(), 14), max_days=720)
    return augmented


def _resource_objects(shop: ShopFloor, operation: NewOperation):
    machines = [shop.machines[mid] for mid in operation.eligible_machine_ids] if operation.eligible_machine_ids else shop.get_machines_for_type(operation.process_type)
    tooling_groups = [shop.get_toolings_for_type(type_id) for type_id in operation.required_tooling_types]
    personnel_groups = [shop.get_personnel_for_skill(skill) for skill in operation.required_personnel_skills]
    combinations = []
    for machine in machines:
        tooling_choices = product(*tooling_groups) if tooling_groups else [()]
        for toolings in tooling_choices:
            if len({item.id for item in toolings}) != len(toolings):
                continue
            personnel_choices = product(*personnel_groups) if personnel_groups else [()]
            for people in personnel_choices:
                if len({item.id for item in people}) != len(people):
                    continue
                combinations.append((machine, tuple(toolings), tuple(people)))
                if len(combinations) >= 48:
                    return combinations, True
    return combinations, False


def _resource_keys(machine, toolings, people):
    return [
        ("machine", machine.id),
        *(("tooling", item.id) for item in toolings),
        *(("personnel", item.id) for item in people),
    ]


def _occupancy(schedule: Iterable[dict]) -> dict[tuple[str, str], list[tuple[float, float]]]:
    result: dict[tuple[str, str], list[tuple[float, float]]] = {}
    for entry in schedule or []:
        try:
            start, end = float(entry.get("start")), float(entry.get("end"))
        except (TypeError, ValueError):
            continue
        if not (math.isfinite(start) and math.isfinite(end) and end > start):
            continue
        keys = [("machine", str(entry.get("machine_id") or ""))]
        keys.extend(("tooling", str(item)) for item in entry.get("tooling_ids", []) or [])
        keys.extend(("personnel", str(item)) for item in entry.get("personnel_ids", []) or [])
        for key in keys:
            if key[1]:
                result.setdefault(key, []).append((start, end))
    for windows in result.values():
        windows.sort()
    return result


def _first_blocker(windows: Iterable[tuple[float, float]], start: float, end: float):
    for occupied_start, occupied_end in windows:
        if occupied_end <= start + 1e-9:
            continue
        if occupied_start >= end - 1e-9:
            return None
        return occupied_start, occupied_end
    return None


def _earliest_joint_slot(shop: ShopFloor, combination, base, added, ready: float, duration: float):
    machine, toolings, people = combination
    resources = [machine, *toolings, *people]
    keys = _resource_keys(machine, toolings, people)
    probe = max(0.0, ready)
    blockers = sum(len(base.get(key, [])) + len(added.get(key, [])) for key in keys)
    for _ in range(blockers + 10):
        start = _joint_next_available_time(resources, probe)
        if not math.isfinite(start):
            return None
        end = _joint_compute_effective_end(resources, start, duration)
        if not math.isfinite(end):
            return None
        overlapping = []
        for key in keys:
            blocker = _first_blocker([*base.get(key, []), *added.get(key, [])], start, end)
            if blocker is not None:
                overlapping.append(blocker)
        if not overlapping:
            return start, end
        probe = max(blocker[1] for blocker in overlapping)
    return None


def _frozen_search(shop: ShopFloor, base_schedule: list[dict], orders, operations, dependencies, *, beam_width=120, time_limit_s=4.0):
    by_id = {operation.op_id: operation for operation in operations}
    release_by_order = {order.order_id: order.release_time for order in orders}
    base = _occupancy(base_schedule)
    states = [_FrozenState()]
    exhaustive = True
    explored = 0
    deadline = time.perf_counter() + max(0.5, time_limit_s)

    for _ in operations:
        next_states: list[_FrozenState] = []
        for state in states:
            if time.perf_counter() >= deadline:
                exhaustive = False
                break
            ready_ops = [
                operation for operation in operations
                if operation.op_id not in state.ends and dependencies[operation.op_id] <= state.ends.keys()
            ]
            for operation in ready_ops:
                ready = release_by_order[operation.order_id]
                for predecessor_id in dependencies[operation.op_id]:
                    predecessor = by_id[predecessor_id]
                    ready = max(ready, state.ends[predecessor_id] + predecessor.turnover_time)
                combinations, capped = _resource_objects(shop, operation)
                exhaustive = exhaustive and not capped
                placements = []
                for combination in combinations:
                    slot = _earliest_joint_slot(
                        shop, combination, base, state.added_occupancy,
                        ready, operation.processing_time,
                    )
                    if slot is not None:
                        placements.append((slot[1], slot[0], combination))
                placements.sort(key=lambda item: (item[0], item[2][0].id))
                if len(placements) > 12:
                    placements = placements[:12]
                    exhaustive = False
                for end, start, combination in placements:
                    machine, toolings, people = combination
                    resource_keys = _resource_keys(machine, toolings, people)
                    resource_tail = max(
                        (occupied_end for key in resource_keys for _, occupied_end in base.get(key, [])),
                        default=0.0,
                    )
                    added = {key: list(value) for key, value in state.added_occupancy.items()}
                    for key in resource_keys:
                        added.setdefault(key, []).append((start, end))
                        added[key].sort()
                    entry = {
                        "op_id": operation.op_id,
                        "op_name": operation.op_name,
                        "task_id": operation.task_id,
                        "order_id": operation.order_id,
                        "machine_id": machine.id,
                        "machine_name": machine.name,
                        "process_type": operation.process_type,
                        "tooling_ids": [item.id for item in toolings],
                        "personnel_ids": [item.id for item in people],
                        "start": round(start, 3),
                        "end": round(end, 3),
                        "duration": round(operation.processing_time, 3),
                        "elapsed_duration": round(end - start, 3),
                        "turnover_time": round(operation.turnover_time, 3),
                        "placement": "gap" if start < resource_tail - 1e-9 else "tail",
                        "status": "inserted",
                    }
                    next_states.append(_FrozenState(
                        ends={**state.ends, operation.op_id: end},
                        entries=[*state.entries, entry],
                        added_occupancy=added,
                    ))
                    explored += 1
        if not next_states:
            if time.perf_counter() >= deadline:
                raise InsertionError("搜索达到时间预算，尚未找到完整可行排程；请减少候选资源后重试")
            raise InsertionError("当前资源日历与基准占用下无法排完全部新工序")
        next_states.sort(key=lambda state: _partial_score(state, orders, operations, dependencies))
        if len(next_states) > beam_width:
            next_states = next_states[:beam_width]
            exhaustive = False
        states = next_states
    return states, exhaustive, explored


def _partial_score(state, orders, operations, dependencies):
    due = {order.order_id: order.expected_due_date for order in orders}
    order_by_op = {operation.op_id: operation.order_id for operation in operations}
    completed_by_order: dict[str, float] = {}
    for op_id, end in state.ends.items():
        order_id = order_by_op[op_id]
        completed_by_order[order_id] = max(completed_by_order.get(order_id, 0.0), end)
    known_tardy = sum(
        math.isfinite(due[order_id]) and completion > due[order_id] + 1e-9
        for order_id, completion in completed_by_order.items()
    )
    return known_tardy, max(state.ends.values(), default=0.0), sum(state.ends.values())


def _enrich_schedule(shop: ShopFloor, schedule: Iterable[dict], new_ids: set[str]):
    result = []
    for raw in schedule or []:
        entry = dict(raw)
        task = shop.tasks.get(str(entry.get("task_id") or ""))
        order = shop.orders.get(task.order_id) if task else None
        entry.update({
            "order_id": order.id if order else entry.get("order_id", ""),
            "order_name": order.name if order else entry.get("order_name", ""),
            "due_date": order.due_date if order and math.isfinite(order.due_date) else None,
            "due_at": shop.time_label(order.due_date) if order and math.isfinite(order.due_date) else None,
            "is_inserted": str(entry.get("op_id") or "") in new_ids,
        })
        result.append(entry)
    return result


def _restore_locked_entries(shop: ShopFloor, base_schedule: list[dict], candidate_schedule: list[dict]):
    """让保护重排结果继续包含基准中的已完工/在制工序，并保持其原位置。"""
    locked_ids = {
        operation.id
        for operation in shop.operations.values()
        if operation.status in {OpStatus.COMPLETED, OpStatus.PROCESSING}
    }
    if not locked_ids:
        return candidate_schedule
    locked_entries = {
        str(entry.get("op_id")): dict(entry)
        for entry in base_schedule
        if str(entry.get("op_id") or "") in locked_ids
    }
    remaining = [
        entry for entry in candidate_schedule
        if str(entry.get("op_id") or "") not in locked_entries
    ]
    return [*locked_entries.values(), *remaining]


def _protects_existing(shop: ShopFloor, base_analytics: ScheduleAnalytics, candidate: ScheduleAnalytics, original_order_ids: set[str]):
    violations = []
    for order_id in original_order_ids:
        order = shop.orders[order_id]
        before = base_analytics.order_completion.get(order_id, 0.0)
        after = candidate.order_completion.get(order_id, float("inf"))
        limit = order.due_date if before <= order.due_date + 1e-9 else before
        if after > limit + 1e-9:
            violations.append({"order_id": order_id, "before": before, "after": after, "limit": limit})
        main_task_id = order.main_task_id
        if main_task_id:
            main_before = base_analytics.task_completion.get(main_task_id, 0.0)
            main_after = candidate.task_completion.get(main_task_id, float("inf"))
            main_limit = order.due_date if main_before <= order.due_date + 1e-9 else main_before
            if main_after > main_limit + 1e-9:
                violations.append({"order_id": order_id, "task_id": main_task_id, "before": main_before, "after": main_after, "limit": main_limit})
    return violations


def _candidate_key(candidate: dict):
    objectives = candidate["analytics"].objective_values
    return (
        objectives.get("main_order_tardy_count", math.inf),
        -objectives.get("critical_active_window_utilization", 0.0),
        objectives.get("avg_flowtime", math.inf),
        candidate.get("max_new_completion", math.inf),
        candidate.get("disruption", math.inf),
    )


def _select_candidate(candidates: list[dict]):
    minimum_tardy = min(item["analytics"].objective_values.get("main_order_tardy_count", math.inf) for item in candidates)
    finalists = [item for item in candidates if item["analytics"].objective_values.get("main_order_tardy_count", math.inf) <= minimum_tardy + 1e-9]
    if len(finalists) == 1:
        return finalists[0]
    utils = [item["analytics"].objective_values.get("critical_active_window_utilization", 0.0) for item in finalists]
    flows = [item["analytics"].objective_values.get("avg_flowtime", 0.0) for item in finalists]
    util_span = max(utils) - min(utils)
    flow_span = max(flows) - min(flows)
    for item, util, flow in zip(finalists, utils, flows):
        util_loss = (max(utils) - util) / util_span if util_span > 1e-9 else 0.0
        flow_loss = (flow - min(flows)) / flow_span if flow_span > 1e-9 else 0.0
        item["ideal_distance"] = math.sqrt(0.5 * util_loss * util_loss + 0.5 * flow_loss * flow_loss)
    return min(finalists, key=lambda item: (item["ideal_distance"], item["max_new_completion"], item["disruption"]))


def _disruption(base_schedule: list[dict], candidate_schedule: list[dict], new_ids: set[str]):
    before = {str(entry.get("op_id")): entry for entry in base_schedule if entry.get("op_id")}
    moved = machine_changed = 0
    total_shift = 0.0
    for entry in candidate_schedule:
        op_id = str(entry.get("op_id") or "")
        if op_id in new_ids or op_id not in before:
            continue
        old = before[op_id]
        shift = abs(float(entry.get("start", 0.0)) - float(old.get("start", 0.0)))
        total_shift += shift
        if shift > 1e-6:
            moved += 1
        if str(entry.get("machine_id")) != str(old.get("machine_id")):
            machine_changed += 1
    return {"moved_operations": moved, "machine_changes": machine_changed, "total_start_shift_hours": round(total_shift, 3)}


def evaluate_insertion(shop: ShopFloor, base_schedule: list[dict], raw_orders, raw_operations, *, policy: str, instance_version: int):
    started = time.perf_counter()
    orders, operations, dependencies = _normalize_input(shop, raw_orders, raw_operations)
    augmented = _augment_shop(shop, orders, operations)
    new_ids = {operation.op_id for operation in operations}
    original_order_ids = set(shop.orders)
    base_schedule = _enrich_schedule(shop, base_schedule, set())
    base_analytics = build_schedule_analytics(shop, SimResult(schedule=base_schedule))

    states, exhaustive, explored = _frozen_search(
        augmented, base_schedule, orders, operations, dependencies,
    )
    candidates = []
    for state in states[:40]:
        schedule = [*base_schedule, *state.entries]
        analytics = build_schedule_analytics(augmented, SimResult(schedule=schedule))
        candidates.append({
            "source": "frozen_gap_search",
            "schedule": schedule,
            "inserted": state.entries,
            "analytics": analytics,
            "max_new_completion": max(state.ends.values()),
            "disruption": 0.0,
        })

    accepted_reschedules = 0
    if policy == "due_protected":
        runtime = SimulationRuntime(augmented)
        for rule_name in ("ATC", "EDD", "MST", "PRIORITY", "BOTTLENECK", "COMPOSITE", "SPT"):
            result = Simulator(augmented, BUILTIN_RULES[rule_name], runtime=runtime).run()
            if not result.schedule:
                continue
            schedule = _enrich_schedule(augmented, result.schedule, new_ids)
            schedule = _restore_locked_entries(shop, base_schedule, schedule)
            analytics = build_schedule_analytics(augmented, SimResult(schedule=schedule))
            violations = _protects_existing(shop, base_analytics, analytics, original_order_ids)
            if not analytics.feasible or violations:
                continue
            change = _disruption(base_schedule, schedule, new_ids)
            inserted = [entry for entry in schedule if entry.get("is_inserted")]
            candidates.append({
                "source": f"protected_{rule_name}",
                "schedule": schedule,
                "inserted": inserted,
                "analytics": analytics,
                "max_new_completion": max((entry.get("end", 0.0) for entry in inserted), default=0.0),
                "disruption": change["total_start_shift_hours"] + 24 * change["machine_changes"],
            })
            accepted_reschedules += 1

    chosen = _select_candidate(candidates)
    merged = sorted(chosen["schedule"], key=lambda item: (float(item.get("start", 0.0)), str(item.get("machine_id", "")), str(item.get("op_id", ""))))
    inserted = sorted(chosen["inserted"], key=lambda item: (float(item.get("start", 0.0)), str(item.get("op_id", ""))))
    analytics = chosen["analytics"]
    change = _disruption(base_schedule, merged, new_ids)
    order_results = []
    for order in orders:
        completion = analytics.order_completion.get(order.order_id, 0.0)
        due = order.expected_due_date
        order_results.append({
            "order_id": order.order_id,
            "order_name": order.order_name,
            "release_time": order.release_time,
            "release_at": augmented.time_label(order.release_time),
            "expected_due_date": due if math.isfinite(due) else None,
            "expected_due_at": augmented.time_label(due) if math.isfinite(due) else None,
            "best_delivery_time": round(completion, 3),
            "best_delivery_at": augmented.time_label(completion),
            "lead_time_hours": round(max(0.0, completion - order.release_time), 3),
            "due_delta_hours": round(due - completion, 3) if math.isfinite(due) else None,
            "conclusion": "not_set" if not math.isfinite(due) else "met" if completion <= due + 1e-9 else "missed",
        })
    kpi_keys = [
        "main_order_tardy_count", "critical_active_window_utilization", "avg_flowtime",
        "total_tardiness", "makespan", "tooling_utilization", "personnel_utilization",
    ]
    kpis = []
    for key in kpi_keys:
        before = float(base_analytics.objective_values.get(key, 0.0))
        after = float(analytics.objective_values.get(key, 0.0))
        kpis.append({"key": key, "before": round(before, 4), "after": round(after, 4), "delta": round(after - before, 4)})
    impact = []
    for order_id in sorted(original_order_ids):
        before = base_analytics.order_completion.get(order_id, 0.0)
        after = analytics.order_completion.get(order_id, before)
        impact.append({
            "order_id": order_id,
            "before_completion": round(before, 3),
            "after_completion": round(after, 3),
            "delta_hours": round(after - before, 3),
            "protected": after <= (shop.orders[order_id].due_date if before <= shop.orders[order_id].due_date + 1e-9 else before) + 1e-9,
        })
    run_id = f"INS-{uuid.uuid4().hex[:12]}"
    payload = {
        "run_id": run_id,
        "status": "done",
        "policy": policy,
        "algorithm": chosen["source"],
        # 当前搜索只枚举各资源组合的最早可行位置；即使束搜索未剪枝，
        # 也不能据此证明不存在“主动等待后反而更优”的全局方案。
        "search_status": "best_found",
        "search_note": "已返回搜索预算内当前最优可行方案；未宣称全局最优",
        "elapsed_ms": round((time.perf_counter() - started) * 1000, 1),
        "explored_states": explored,
        "accepted_reschedules": accepted_reschedules,
        "order_results": order_results,
        "inserted_schedule": inserted,
        "merged_schedule": merged[:2000],
        "schedule_total": len(merged),
        "schedule_truncated": len(merged) > 2000,
        "kpis": kpis,
        "existing_order_impact": impact,
        "existing_orders_protected": all(item["protected"] for item in impact),
        "moved_operations": change["moved_operations"],
        "machine_changes": change["machine_changes"],
        "total_start_shift_hours": change["total_start_shift_hours"],
        "gap_operation_count": sum(entry.get("placement") == "gap" for entry in inserted),
        "tail_operation_count": sum(entry.get("placement") == "tail" for entry in inserted),
        "created_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    }
    return InsertionRun(run_id, instance_version, payload["created_at"], payload, merged)


ORDER_HEADERS = ["order_id", "order_name", "release_time", "expected_due_date", "main_task_id"]
OPERATION_HEADERS = [
    "order_id", "op_id", "task_id", "op_name", "process_type", "processing_time_hrs",
    "turnover_time_hrs", "predecessor_ops", "predecessor_tasks", "eligible_machine_ids",
    "required_tooling_types", "required_personnel_skills",
]


def _sheet_rows(sheet) -> list[dict]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        for row in rows[1:] if any(value not in (None, "") for value in row)
    ]


def parse_insertion_file(content: bytes, filename: str) -> dict:
    lower = str(filename or "").lower()
    if lower.endswith((".xlsx", ".xlsm")):
        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            if "orders" not in workbook.sheetnames or "operations" not in workbook.sheetnames:
                raise InsertionError("Excel 必须包含 orders 和 operations 两个 Sheet")
            orders = _sheet_rows(workbook["orders"])
            operations = _sheet_rows(workbook["operations"])
        finally:
            workbook.close()
        return {"orders": orders, "operations": operations}
    if lower.endswith((".csv", ".tsv")):
        text = content.decode("utf-8-sig")
        dialect = csv.excel_tab if lower.endswith(".tsv") else csv.excel
        rows = list(csv.DictReader(io.StringIO(text), dialect=dialect))
        if not rows:
            raise InsertionError("文件没有数据行")
        keys = set(rows[0])
        kind = "operations" if "op_id" in keys else "orders" if "order_id" in keys else ""
        if not kind:
            raise InsertionError("CSV 表头无法识别，请使用插单模板字段")
        return {"orders": rows if kind == "orders" else [], "operations": rows if kind == "operations" else []}
    raise InsertionError("仅支持 .xlsx、.xlsm、.csv 或 .tsv 文件")


def build_insertion_template() -> bytes:
    workbook = openpyxl.Workbook()
    orders = workbook.active
    orders.title = "orders"
    orders.append(ORDER_HEADERS)
    orders.append(["URGENT-001", "临时插单", 0, 72, "URG-T-MAIN"])
    operations = workbook.create_sheet("operations")
    operations.append(OPERATION_HEADERS)
    operations.append(["URGENT-001", "URG-OP-01", "URG-T-PART", "加工", "cut", 2, 0.5, "", "", "", "", ""])
    operations.append(["URGENT-001", "URG-OP-02", "URG-T-MAIN", "装配", "assembly", 3, 0, "", "URG-T-PART", "", "", ""])
    for sheet in (orders, operations):
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
        for column in sheet.columns:
            width = min(40, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_insertion_export(run: InsertionRun) -> bytes:
    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "交期结论"
    summary.append(["订单ID", "订单名称", "放行时间", "期望交期", "最佳交期", "流程时间(h)", "结论"])
    labels = {"met": "可满足", "missed": "不可满足", "not_set": "未设置期望交期"}
    for item in run.payload["order_results"]:
        summary.append([item["order_id"], item["order_name"], item["release_at"], item["expected_due_at"], item["best_delivery_at"], item["lead_time_hours"], labels[item["conclusion"]]])
    detail = workbook.create_sheet("新工序排程")
    detail.append(["订单ID", "任务令ID", "工序ID", "工序名称", "机器", "工装", "人员", "开始", "结束", "插入方式"])
    for entry in run.payload["inserted_schedule"]:
        detail.append([entry.get("order_id"), entry.get("task_id"), entry.get("op_id"), entry.get("op_name"), entry.get("machine_id"), ";".join(entry.get("tooling_ids", []) or []), ";".join(entry.get("personnel_ids", []) or []), entry.get("start"), entry.get("end"), entry.get("placement", "重排")])
    kpis = workbook.create_sheet("KPI对比")
    kpis.append(["KPI", "插单前", "插单后", "变化"])
    for item in run.payload["kpis"]:
        kpis.append([item["key"], item["before"], item["after"], item["delta"]])
    impact = workbook.create_sheet("现有订单影响")
    impact.append(["订单ID", "原完工", "新完工", "变化(h)", "保护通过"])
    for item in run.payload["existing_order_impact"]:
        impact.append([item["order_id"], item["before_completion"], item["after_completion"], item["delta_hours"], "是" if item["protected"] else "否"])
    merged = workbook.create_sheet("合并排程")
    merged.append(["订单ID", "任务令ID", "工序ID", "机器", "工装", "人员", "开始", "结束", "是否插单"])
    for entry in run.full_schedule:
        merged.append([entry.get("order_id"), entry.get("task_id"), entry.get("op_id"), entry.get("machine_id"), ";".join(entry.get("tooling_ids", []) or []), ";".join(entry.get("personnel_ids", []) or []), entry.get("start"), entry.get("end"), "是" if entry.get("is_inserted") else "否"])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
