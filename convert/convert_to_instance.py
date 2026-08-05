# -*- coding: utf-8 -*-
"""
排产数据转换脚本：源表(当前目录下日期最大的 YYYYMMDD.xlsx) + 固定基表(预排工位.xlsx)
  -> 目标表(instance_template_v2.xlsx 模板结构)

说明:
  - 源表自动选取当前目录下形如 20260729.xlsx 的文件中日期最大的一个
  - plan_start_at = 源表文件名日期 08:00:00, 所有时间字段为相对该时刻的偏移小时数
  - machines.shifts 遇周日(可用时长置0)
"""

import argparse
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# ============================================================
# 全局参数
# ============================================================

PLAN_START = None            # 运行时由源表文件名确定: 该日期 08:00:00
PLAN_START_HOUR = 8          # plan_start_at 的小时数
OS_MACHINE_COUNT = 4000      # 委外无限产能机台 OS_1 ~ OS_4000
ZZ_MACHINE_COUNT = 50        # 自制无限产能机台 ZZ_1 ~ ZZ_50
TURNOVER_TIME_HRS = 2        # 默认周转时间
OS_PRED_TURNOVER_HRS = 24    # 委外类工序的前置工序(非委外类)的周转时间
OS_TO_OS_TURNOVER_HRS = 0    # 委外类工序之间的周转时间
TIME_DECIMALS = 3
SOURCE_FILE_PATTERN = r'^(\d{8})\.xlsx$'

COL_ORDER_TIME = '接单时间'
COL_CRAFT_TIME = '工艺排配时间'   # release_time 基准: 工艺排配时间 + 5 天
COL_PLAN = '计划号'
COL_TASK = '任务令'
COL_WORK = 'WORK'
COL_STATION = '工位'
COL_LATEST_END = '最晚完工时间'
COL_PROC = '工序'
COL_REMAIN = '剩余工时'
COL_MES_DUE = '【Mes+】交期*'
COL_QUALIFY_CHAIN = '齐套需求工序'
COL_PART = '零件编号'

COL_TYPE_ID = 'type_id'
COL_TYPE_NAME = 'type_name'
COL_CRIT = 'is_critical'
COL_MACHINE_ID = 'machine_id'
COL_MACHINE_NAME = 'machine_name'

PLANNING_CONTEXT_SHEET = 'planning_context'
PLAN_START_FIELD = 'plan_start_at'

PROCESS_TYPES = [
    "AO", "AO-OS", "CG", "CNC", "CO",
    "CO-OS", "DP", "EDM", "EDM-C", "EP",
    "G", "G_C", "GL3", "GL5", "GP",
    "GP_C", "H", "HS", "IO", "JG",
    "KG", "L", "LC", "LD", "LF",
    "LW", "LW-OS", "M", "MA", "MA_C",
    "OS", "PO", "PVD", "QC", "QC_C",
    "QC-OS", "QX-OS", "QXQ-OS", "RG", "SB",
    "STO", "STP-OS", "TAP", "TO", "UC",
    "VI", "WE", "WEDM", "WEDM_C", "WEDM_O",
    "FQC",
]
_PROCESS_TYPES_SORTED = sorted(PROCESS_TYPES, key=len, reverse=True)

_CNC_ABCDE_MACHINES = (
    "CNC-01,CNC-02,CNC-03,CNC-04,CNC-05,CNC-06,CNC-07,CNC-08,CNC-09,CNC-10,"
    "CNC-18,CNC-19,CNC-HT-01,CNC-HT-02,CNC-HT-03,CNC-HT-04,CNC-HT-05,CNC-HT-06,"
    "CNC-H1,CNC-H2,CNC-H8,CNC-E800-01,CNC-E800-02,CNC-E800-03,CNC-E800-04,CNC-E800-05,"
    "CNC-P1,CNC-S1,CNC-S2,CNC-S3,CNC-E1200-01,CNC-E1200-02,CNC-E1200-03,CNC-F8,"
    "CNC-Y1,CNC-Y2,CNC-Y3,CNC-Y4,CNC-JD01,CNC-JD02,CNC-JD03,CNC-JD04,"
    "CNC-JD1200-01,CNC-JD1200-02,CNC-JD1200-03,CNC-JD1600-01"
)
_CNC_E_GROUP = (
    "CNC-H1,CNC-H2,CNC-H8,CNC-P1,"
    "CNC-E800-01,CNC-E800-02,CNC-E800-03,CNC-E800-04,CNC-E800-05,"
    "CNC-E1200-01,CNC-E1200-02,CNC-E1200-03,CNC-F8,"
    "CNC-JD1200-01,CNC-JD1200-02,CNC-JD1200-03,CNC-JD1600-01"
)

MACHINE_MAP = {
    "AO": "AO", "AO-OS": "AO-OS", "CG": "CG-01", "3D": "3D", "DDP": "DP-01",
    "CNC_ABCDE": _CNC_ABCDE_MACHINES,
    "CNC-E1200": _CNC_E_GROUP, "CNC-E800": _CNC_E_GROUP, "CNC-H8": _CNC_E_GROUP,
    "CNC-L3000": "CNC-Y1,CNC-Y2", "CNC-T30": "CNC-Y3,CNC-Y4",
    "CNC-五轴": "CNC-5A1,CNC-5A2", "CNC-F8": "CNC-F8",
    "CNC-EROWA": ("CNC-01,CNC-02,CNC-03,CNC-04,CNC-05,CNC-06,CNC-07,CNC-08,CNC-09,CNC-10,"
                  "CNC-18,CNC-19,CNC-HT-01,CNC-HT-02,CNC-HT-03,CNC-HT-04,CNC-HT-05,CNC-HT-06,"
                  "CNC-JD01,CNC-JD02,CNC-JD03,CNC-JD04"),
    "CNC-EROWA哈挺": "CNC-HT-01,CNC-HT-02,CNC-HT-03,CNC-HT-04,CNC-HT-05,CNC-HT-06",
    "CNC-JD1200T": _CNC_E_GROUP, "CNC-JD1600T": _CNC_E_GROUP,
    "CNC-JD600T": "CNC-JD01,CNC-JD02,CNC-JD03,CNC-JD04",
    "CNC-Kern": "CNC-K1,CNC-K2,CNC-K3,CNC-K4,CNC-K5",
    "CNC-P9": _CNC_E_GROUP, "CNC-高速": "CNC-S1,CNC-S2,CNC-S3",
    "CO": "CO", "CO-OS": "CO-OS", "DP": "DP-000001",
    "EDM": "EDM-01,EDM-02,EDM-03,EDM-04,EDM-05,EDM-06",
    "EDM-C": "EDM-C01,EDM-C02,EDM-C03,EDM-C04",
    "EP": "EP-01,EP-02",
    "G": "G-02,G-03,G-05", "G_C": "G-01,G-04",
    "GL3": "GL3-01,GL3-02", "GL5": "GL5-01,GL5-02",
    "GP": "GP-01,GP-02", "GP_C": "GP_C-01,GP_C-02",
    "H": "H", "HS": "HS", "IO": "IO",
    "JG": "JG-01,JG-02", "KG": "KG",
    "L": "L-01,L-02", "LC": "LC-01", "LD": "LD-01,LD-02", "LF": "LF-01",
    "LW": "LW", "LW-OS": "LW-OS",
    "M": "M-01,M-02,M-03", "MA": "MA-01,MA-02", "MA_C": "MA_C-01",
    "PO": "C9PO-01,C11PO-01,C12PO-01", "PVD": "PVD-01",
    "QC": "QC-01,QC-02,QC-03,QC-04,QC-05,QC-07,QC-08",
    "FQC": "QC-06,QC-09,QC-10,FQC-01,FQC-02,FQC-03,FQC-04,FQC-05",
    "QC_C": "QC_C-01,QC_C-02,QC_C-03,QC_C-04,QC_C-05,QC_C-06,QC_C-07",
    "RG": "RG-01", "SB": "SB", "STO": "STO", "STP-OS": "STP-OS",
    "TAP": "C9TAP-01,C12TAP-01,C12TAP-02",
    "TO": "TO-GX,TO-LG", "UC": "UC-01", "VI": "VI-01",
    "WE": "WE-01,WE-02",
    "WEDM": ("WEDM-01,WEDM-02,WEDM-03,WEDM-04,WEDM-05,WEDM-06,WEDM-07,"
             "WEDM-FANUC01,WEDM-FANUC02,WEDM-FANUC03"),
    "WEDM-C": "WEDM_C-01,WEDM_C-02,WEDM_C-03",
    "WEDM-O": "WEDM-Y1,WEDM-Y2",
}
_MACHINE_KEYS_SORTED = sorted(
    (k for k in MACHINE_MAP if k != "CNC_ABCDE"), key=len, reverse=True
)

# ---- 委外工序(OS): 无限产能, 机台池 OS_1~OS_N ----
OS_PROCESS_SET = {
    "QX-OS", "QXQ-OS", "OS", "LW-OS", "CO-OS", "QC-OS", "STP-OS", "AO-OS",
    "L|OS", "L-OS", "X-OS",
    "AO", "H", "LW", "CO",
}

# ---- 自制无限产能工序(ZZ): 机台池 ZZ_1~ZZ_N ----
ZZ_PROCESS_SET = {"BP", "HS", "KG", "STO", "IO"}

# ---- 触发"前置工序周转改24小时"的工序集合: 仅委外工序 ----
TURNOVER_24_TRIGGER_SET = set(OS_PROCESS_SET)

CNC_SUFFIXES = [
    "E1200", "E800", "H8", "L3000", "T30", "五轴", "F8",
    "EROWA哈挺", "EROWA", "JD1200T", "JD1600T", "JD600T", "Kern", "P9",
]

# ============================================================
# 工序合并规则 (operations 后处理)
# ============================================================
MERGE_PROCESS_TYPES = {"CNC", "LC", "LF"}
MERGE_SUFFIXES = {
    "CNC": ["E800", "Kern", "高速", "EROWA", "JD600T", "F8", "五轴",
            "E1200", "L3000", "H8", "P9", "JD1200T", "JD1600T"],
    "LC":  ["卡盘", "治具"],
    "LF":  ["卡盘"],
}
_MERGE_PATTERNS = {}
for _pref, _sufs in MERGE_SUFFIXES.items():
    _MERGE_PATTERNS[_pref] = [
        (s, re.compile(r'^' + re.escape(_pref) + r'(\d+)-' + re.escape(s) + r'$'))
        for s in sorted(_sufs, key=len, reverse=True)
    ]

# ============================================================
# machines.shifts (运行时依据 plan_start 计算, 周日可用时长为0)
# ============================================================
SHIFT_DAY_COUNT = 1000
_SHIFT_NORMAL_DAY0_PREFIX = "0/0/1;0/3/9.5;0/14/11"
SUNDAY_SHIFT_TPL = "{d}/0/0"          # 周日: 全天不可用
OS_TYPE_ID = "OS"

def is_sunday(plan_start, day_offset):
    return (plan_start + timedelta(days=day_offset)).weekday() == 6

def build_shifts(plan_start):
    """返回 (OS类24小时班次串, 普通两班制班次串)

    - OS 类型(委外): 每天 24 小时可用, 但遇周日整天不可用(d/0/0)
    - 非 OS 类型(自制): 两班制, 周日照常上班, 不做特殊处理
    """
    parts24 = []
    for d in range(0, SHIFT_DAY_COUNT):
        parts24.append(SUNDAY_SHIFT_TPL.format(d=d) if is_sunday(plan_start, d)
                       else f"{d}/0/24")
    shift_24h = ";".join(parts24)

    # 普通类型: day0 固定前缀 + day1..SHIFT_DAY_COUNT 循环 (周日同样上班)
    shift_normal = _SHIFT_NORMAL_DAY0_PREFIX + ";" + ";".join(
        f"{d}/3/9.5;{d}/14/11" for d in range(1, SHIFT_DAY_COUNT + 1)
    )
    return shift_24h, shift_normal

RELEASE_OFFSET_DAYS = 5   # release_time = 工艺排配时间 + 5 个自然日

def craft_release_offset(ts):
    """release_time: 工艺排配时间 + RELEASE_OFFSET_DAYS 天, 转相对 PLAN_START 的偏移小时数(负数归零)"""
    if ts is None or (isinstance(ts, float) and pd.isna(ts)) or pd.isna(ts):
        return ""
    ts = pd.Timestamp(ts) + pd.Timedelta(days=RELEASE_OFFSET_DAYS)
    return clamp_non_negative(to_offset_hours(ts))

def to_offset_hours(ts):
    if ts is None or (isinstance(ts, float) and pd.isna(ts)) or pd.isna(ts):
        return ""
    ts = pd.Timestamp(ts).to_pydatetime()
    hours = (ts - PLAN_START).total_seconds() / 3600.0
    hours = round(hours, TIME_DECIMALS)
    return int(hours) if hours == int(hours) else hours

def minutes_to_hours(v):
    if v is None or pd.isna(v):
        return ""
    try:
        hours = float(v) / 60.0
    except (TypeError, ValueError):
        return ""
    hours = round(hours, 2)
    return 0.1 if hours == 0 else hours

def clamp_non_negative(v):
    if v == "":
        return v
    return 0 if v < 0 else v

def norm_str(v):
    if v is None or pd.isna(v):
        return ""
    return str(v).strip()

def match_process_type(proc):
    for pt in _PROCESS_TYPES_SORTED:
        if proc.startswith(pt):
            return pt
    return ""

def build_os_pool():
    return ",".join(f"OS_{i}" for i in range(1, OS_MACHINE_COUNT + 1))

def build_zz_pool():
    return ",".join(f"ZZ_{i}" for i in range(1, ZZ_MACHINE_COUNT + 1))

def is_os_process(op_name):
    """是否为委外工序(精确匹配委外集合, 或以 -OS 结尾)"""
    p = norm_str(op_name)
    return bool(p) and (p in OS_PROCESS_SET or p.endswith("-OS"))

def is_zz_process(op_name):
    """是否为自制无限产能工序"""
    return norm_str(op_name) in ZZ_PROCESS_SET

def triggers_pred_turnover_24(op_name):
    """该工序是否会把其非同类前置工序的周转时间改为24小时"""
    p = norm_str(op_name)
    return bool(p) and (p in TURNOVER_24_TRIGGER_SET or p.endswith("-OS"))

def eligible_machines(proc, os_pool, zz_pool, warn_set):
    p = norm_str(proc)
    if not p:
        warn_set.add("(空工序)")
        return os_pool
    # 规则1: 委外工序(OS类) -> OS 池(无限产能)
    if p in OS_PROCESS_SET:
        return os_pool
    # 规则2: 自制无限产能工序(ZZ类) -> ZZ 池
    if p in ZZ_PROCESS_SET:
        return zz_pool
    # 规则3: 工序以 '-OS' 结尾(委外) -> OS 池
    if p.endswith("-OS"):
        return os_pool
    if p.startswith("CNC"):
        if any(p.startswith(x) for x in ("CNC_A", "CNC_B", "CNC_C", "CNC_D", "CNC_E")):
            return MACHINE_MAP["CNC_ABCDE"]
        m = re.match(r"^CNC\d*-(.+)$", p)
        if m:
            suffix = m.group(1)
            for s in CNC_SUFFIXES:
                if suffix.startswith(s):
                    key = "CNC-" + s
                    if key in MACHINE_MAP:
                        return MACHINE_MAP[key]
        return MACHINE_MAP["CNC_ABCDE"]
    if p.startswith("WEDM"):
        if p.startswith("WEDM-C") or p.startswith("WEDM_C"):
            return MACHINE_MAP["WEDM-C"]
        if p.startswith("WEDM-O") or p.startswith("WEDM_O"):
            return MACHINE_MAP["WEDM-O"]
        return MACHINE_MAP["WEDM"]
    if p.startswith("EDM"):
        if p.startswith("EDM-C"):
            return MACHINE_MAP["EDM-C"]
        return MACHINE_MAP["EDM"]
    if p.startswith("LC"):
        return MACHINE_MAP["LC"]
    if p.startswith("LF"):
        return MACHINE_MAP["LF"]
    if p.startswith("FQC"):
        return MACHINE_MAP["FQC"]
    if p.startswith("WE"):
        return MACHINE_MAP["WE"]
    if p in MACHINE_MAP:
        return MACHINE_MAP[p]
    for key in _MACHINE_KEYS_SORTED:
        if p.startswith(key):
            return MACHINE_MAP[key]
    warn_set.add(p)
    return os_pool

def plan_no_ok(plan):
    s = norm_str(plan)
    return (len(s) >= 6 and s[:6].isdigit()) or s.startswith("PM")

def task_len_ok(task):
    return len(norm_str(task)) in (11, 14)

def extract_work_number(work_str):
    m = re.search(r'WORK(\d+)', norm_str(work_str))
    return int(m.group(1)) if m else None

def get_part_prefix(part):
    if part is None or pd.isna(part):
        return ""
    p = norm_str(part)
    if '-' in p:
        return p.split('-')[0]
    return p

def parse_order_id(plan):
    s = norm_str(plan)
    s = re.sub(r'ED\d+$', '', s)
    s = re.sub(r'TL\d+$', '', s)
    s = re.sub(r'W\d*$', '', s)
    s = re.sub(r'Z\d*$', '', s)
    s = re.sub(r'-W.*$', '', s)
    s = s.rstrip('-')
    if not s:
        s = norm_str(plan)
    return s

def find_latest_source(work_dir):
    """在目录下查找形如 YYYYMMDD.xlsx 的文件, 返回日期最大的那个 (路径, 日期)"""
    candidates = []
    for p in work_dir.iterdir():
        if not p.is_file():
            continue
        m = re.match(SOURCE_FILE_PATTERN, p.name)
        if not m:
            continue
        try:
            dt = datetime.strptime(m.group(1), "%Y%m%d")
        except ValueError:
            continue
        candidates.append((dt, p))
    if not candidates:
        sys.exit(f"[错误] 在 {work_dir} 下未找到形如 20260729.xlsx 的源表文件")
    candidates.sort(key=lambda x: x[0])
    dt, path = candidates[-1]
    others = [p.name for _, p in candidates[:-1]]
    print(f"[源表] 选用 {path.name} (共发现 {len(candidates)} 个候选"
          + (f", 其余: {others}" if others else "") + ")")
    return path, dt

def read_base_table(base_path):
    xls = pd.ExcelFile(base_path)
    for sheet in xls.sheet_names:
        raw = pd.read_excel(base_path, sheet_name=sheet, header=None, nrows=20)
        for i in range(len(raw)):
            row_vals = [norm_str(v) for v in raw.iloc[i].tolist()]
            if COL_TYPE_ID in row_vals:
                df = pd.read_excel(base_path, sheet_name=sheet, header=i)
                df.columns = [norm_str(c) for c in df.columns]
                print(f"[读取] 固定基表: sheet='{sheet}', 表头在第 {i+1} 行, 数据 {len(df)} 行")
                return df
    sys.exit(f"[错误] 固定基表 {base_path} 的所有 sheet 中都未找到含'{COL_TYPE_ID}'的表头行")

# ============================================================
# 工序合并 (operations / initial_state 后处理)
# ============================================================

def parse_mergeable_op_name(op_name, process_type):
    """解析可合并的 op_name -> (前缀, 序号, 后缀); 不符合模式返回 None"""
    pt = norm_str(process_type)
    if pt not in MERGE_PROCESS_TYPES:
        return None
    name = norm_str(op_name)
    for suf, rgx in _MERGE_PATTERNS.get(pt, []):
        m = rgx.match(name)
        if m:
            return (pt, int(m.group(1)), suf)
    return None

def _sum_hours(values):
    total = 0.0
    has_val = False
    for v in values:
        if v == "" or v is None:
            continue
        try:
            total += float(v)
            has_val = True
        except (TypeError, ValueError):
            continue
    if not has_val:
        return ""
    return round(total, 2)

def _union_machines(values):
    seen, seen_set = [], set()
    for v in values:
        for m in norm_str(v).split(","):
            m = m.strip()
            if m and m not in seen_set:
                seen_set.add(m)
                seen.append(m)
    return ",".join(seen)

def merge_continuous_ops(ops_rows, init_rows, merge_report=None):
    """合并同一任务令下连续的 CNC / LC / LF 工序 (详见规则说明)"""
    if merge_report is None:
        merge_report = []
    init_by_op = {r["op_id"]: r for r in init_rows}

    def _machine_of(op_id):
        return norm_str(init_by_op.get(op_id, {}).get("initial_assigned_machine_id", ""))

    by_task = {}
    for r in ops_rows:
        by_task.setdefault(r["task_id"], []).append(r)
    for t in by_task:
        by_task[t].sort(key=lambda r: r["_work_num"])

    remap, removed_ids, merged_count = {}, set(), 0

    for t, rows in by_task.items():
        i = 0
        while i < len(rows):
            base = rows[i]
            parsed = parse_mergeable_op_name(base["op_name"], base["process_type"])
            if parsed is None:
                i += 1
                continue
            pref, idx, suf = parsed
            seg, seg_idx = [base], [idx]
            j = i + 1
            block_reason, block_rows = "", []
            while j < len(rows):
                nxt = rows[j]
                nxt_parsed = parse_mergeable_op_name(nxt["op_name"], nxt["process_type"])
                if nxt_parsed is None:
                    break
                n_pref, n_idx, n_suf = nxt_parsed
                if n_pref != pref or n_suf != suf:
                    break
                if nxt["_work_num"] != seg[-1]["_work_num"] + 1:
                    block_reason = (f"work_num不连续(前一道WORK{seg[-1]['_work_num']}, "
                                    f"本道WORK{nxt['_work_num']})")
                    block_rows = [seg[-1], nxt]
                    break
                if n_idx != seg_idx[-1] + 1:
                    block_reason = (f"工序序号不连续(前一道{pref}{seg_idx[-1]}, "
                                    f"本道{pref}{n_idx})")
                    block_rows = [seg[-1], nxt]
                    break
                seg.append(nxt)
                seg_idx.append(n_idx)
                j += 1

            if block_reason and block_rows:
                for r in block_rows:
                    merge_report.append({
                        "类型": "未合并", "task_id": t,
                        "合并后op_id": "", "合并后op_name": "",
                        "原op_id": r["op_id"], "原op_name": r["op_name"],
                        "process_type": r["process_type"], "WORK编号": r["_work_num"],
                        "原processing_time_hrs": r["processing_time_hrs"],
                        "原turnover_time_hrs": r["turnover_time_hrs"],
                        "原initial_assigned_machine_id": _machine_of(r["op_id"]),
                        "合并后processing_time_hrs": "", "合并后turnover_time_hrs": "",
                        "未合并原因": block_reason,
                    })

            if len(seg) < 2:
                i += 1
                continue

            machine_ids = [_machine_of(r["op_id"]) for r in seg]
            non_empty = [m for m in machine_ids if m]
            if len(set(non_empty)) > 1:
                reason = f"工位不同({'/'.join(m or '空' for m in machine_ids)})"
                for r, mid in zip(seg, machine_ids):
                    merge_report.append({
                        "类型": "未合并", "task_id": t,
                        "合并后op_id": "", "合并后op_name": "",
                        "原op_id": r["op_id"], "原op_name": r["op_name"],
                        "process_type": r["process_type"], "WORK编号": r["_work_num"],
                        "原processing_time_hrs": r["processing_time_hrs"],
                        "原turnover_time_hrs": r["turnover_time_hrs"],
                        "原initial_assigned_machine_id": mid,
                        "合并后processing_time_hrs": "", "合并后turnover_time_hrs": "",
                        "未合并原因": reason,
                    })
                i += 1
                continue

            keep, drop = seg[0], seg[1:]
            merged_name = f"{pref}{seg_idx[0]}-{seg_idx[-1]}-{suf}"
            merged_time = _sum_hours([r["processing_time_hrs"] for r in seg])

            for k, r in enumerate(seg):
                merge_report.append({
                    "类型": "已合并", "task_id": t,
                    "合并后op_id": keep["op_id"], "合并后op_name": merged_name,
                    "原op_id": r["op_id"], "原op_name": r["op_name"],
                    "process_type": r["process_type"], "WORK编号": r["_work_num"],
                    "原processing_time_hrs": r["processing_time_hrs"],
                    "原turnover_time_hrs": r["turnover_time_hrs"],
                    "原initial_assigned_machine_id": machine_ids[k],
                    "合并后processing_time_hrs": merged_time if k == 0 else "",
                    "合并后turnover_time_hrs": (TURNOVER_TIME_HRS if k == len(seg) - 1 else 0),
                    "未合并原因": "",
                })

            keep["op_name"] = merged_name
            keep["processing_time_hrs"] = merged_time
            keep["turnover_time_hrs"] = TURNOVER_TIME_HRS
            keep["eligible_machine_ids"] = _union_machines(
                [r["eligible_machine_ids"] for r in seg])
            ptasks = []
            for r in seg:
                for x in norm_str(r.get("predecessor_tasks", "")).split(";"):
                    x = x.strip()
                    if x and x not in ptasks:
                        ptasks.append(x)
            keep["predecessor_tasks"] = ";".join(sorted(ptasks))

            keep_init = init_by_op.get(keep["op_id"])
            if keep_init is not None:
                seg_inits = [init_by_op[r["op_id"]] for r in seg if r["op_id"] in init_by_op]
                starts = [x["initial_start_time"] for x in seg_inits if x["initial_start_time"] != ""]
                ends = [x["initial_end_time"] for x in seg_inits if x["initial_end_time"] != ""]
                if starts:
                    keep_init["initial_start_time"] = min(starts)
                if ends:
                    keep_init["initial_end_time"] = max(ends)
                keep_init["initial_remaining_processing_time"] = _sum_hours(
                    [x["initial_remaining_processing_time"] for x in seg_inits])
                keep_init["initial_assigned_machine_id"] = machine_ids[0]

            for r in drop:
                remap[r["op_id"]] = keep["op_id"]
                removed_ids.add(r["op_id"])

            merged_count += 1
            i = j

    if not removed_ids:
        print("[合并] 未发现可合并的连续 CNC/LC/LF 工序")
        return ops_rows, init_rows

    ops_rows = [r for r in ops_rows if r["op_id"] not in removed_ids]
    init_rows = [r for r in init_rows if r["op_id"] not in removed_ids]

    def resolve(op_id):
        seen = set()
        while op_id in remap and op_id not in seen:
            seen.add(op_id)
            op_id = remap[op_id]
        return op_id

    for r in ops_rows:
        pred = norm_str(r.get("predecessor_ops", ""))
        if not pred:
            continue
        new_pred = resolve(pred)
        r["predecessor_ops"] = "" if new_pred == r["op_id"] else new_pred

    print(f"[合并] 共合并 {merged_count} 段连续工序, 删除 {len(removed_ids)} 行, "
          f"operations 剩余 {len(ops_rows)} 行")

    merged_op_ids = {r["原op_id"] for r in merge_report if r["类型"] == "已合并"}
    if merged_op_ids:
        merge_report[:] = [r for r in merge_report
                           if not (r["类型"] == "未合并" and r["原op_id"] in merged_op_ids)]
    return ops_rows, init_rows


def apply_os_pred_turnover(ops_rows):
    """依据后继工序类型调整前置工序的 turnover_time_hrs。必须在工序合并之后执行
    (合并会重置段首的 turnover, 且 predecessor_ops 此时已完成重定向)。

    规则(按优先级):
      1) 委外工序 X, 其前置 P 也是委外工序   -> P.turnover = 0   (委外之间无需周转)
      2) 委外工序 X, 其前置 P 非委外工序     -> P.turnover = 24
      其余保持默认 2 小时(自制工序之间, 含 ZZ 类无限产能工序)。
    """
    by_id = {r["op_id"]: r for r in ops_rows}
    n24 = n0 = 0
    for r in ops_rows:
        pred_id = norm_str(r.get("predecessor_ops", ""))
        if not pred_id:
            continue
        pred = by_id.get(pred_id)
        if pred is None:
            continue
        # 规则1: 委外 -> 委外, 周转 0
        if is_os_process(r["op_name"]) and is_os_process(pred["op_name"]):
            if pred["turnover_time_hrs"] != OS_TO_OS_TURNOVER_HRS:
                pred["turnover_time_hrs"] = OS_TO_OS_TURNOVER_HRS
                n0 += 1
            continue
        # 规则2: 委外(或BP) 的前置若非同类, 周转 24
        if triggers_pred_turnover_24(r["op_name"]) and \
                not triggers_pred_turnover_24(pred["op_name"]):
            if pred["turnover_time_hrs"] != OS_PRED_TURNOVER_HRS:
                pred["turnover_time_hrs"] = OS_PRED_TURNOVER_HRS
                n24 += 1
    if n24 or n0:
        print(f"[周转] 调整 {n24} 道工序为 {OS_PRED_TURNOVER_HRS} 小时(后接委外工序), "
              f"{n0} 道为 {OS_TO_OS_TURNOVER_HRS} 小时(委外之间)")
    return ops_rows


def write_merge_report(merge_report, report_path):
    columns = ["类型", "task_id", "合并后op_id", "合并后op_name",
               "原op_id", "原op_name", "process_type", "WORK编号",
               "原processing_time_hrs", "原turnover_time_hrs",
               "原initial_assigned_machine_id",
               "合并后processing_time_hrs", "合并后turnover_time_hrs",
               "未合并原因"]
    if not merge_report:
        df = pd.DataFrame(columns=columns)
    else:
        df = pd.DataFrame(merge_report, columns=columns)
        df = df.sort_values(by=["类型", "task_id", "合并后op_id", "WORK编号"],
                            ascending=[False, True, True, True],
                            kind="mergesort").reset_index(drop=True)
    df.to_excel(report_path, index=False, sheet_name="合并清单")
    merged_n = int((df["类型"] == "已合并").sum()) if len(df) else 0
    blocked_n = int((df["类型"] == "未合并").sum()) if len(df) else 0
    print(f"[清单] 已输出合并工序清单: {report_path} "
          f"(已合并 {merged_n} 行, 未合并 {blocked_n} 行)")

def build_all(source_path, base_path, merge_report=None):
    if merge_report is None:
        merge_report = []
    src = pd.read_excel(source_path)
    src.columns = [norm_str(c) for c in src.columns]
    base = read_base_table(base_path)

    for col in (COL_ORDER_TIME, COL_CRAFT_TIME, COL_PLAN, COL_TASK, COL_WORK, COL_STATION,
                COL_LATEST_END, COL_PROC, COL_REMAIN, COL_MES_DUE, COL_QUALIFY_CHAIN, COL_PART):
        if col not in src.columns:
            sys.exit(f"[错误] 源表缺少列: {col}\n实际列: {list(src.columns)}")
    for col in (COL_TYPE_ID, COL_TYPE_NAME, COL_CRIT, COL_MACHINE_ID, COL_MACHINE_NAME):
        if col not in base.columns:
            sys.exit(f"[错误] 固定基表缺少列: {col}\n实际列: {list(base.columns)}")

    src = src[src[COL_PLAN].apply(plan_no_ok)]
    src = src[src[COL_TASK].apply(task_len_ok)].copy()
    src[COL_TASK] = src[COL_TASK].apply(norm_str)
    src[COL_PLAN] = src[COL_PLAN].apply(norm_str)
    src[COL_ORDER_TIME] = pd.to_datetime(src[COL_ORDER_TIME], errors="coerce")
    src[COL_CRAFT_TIME] = pd.to_datetime(src[COL_CRAFT_TIME], errors="coerce")
    src[COL_MES_DUE] = pd.to_datetime(src[COL_MES_DUE], errors="coerce")
    src[COL_LATEST_END] = pd.to_datetime(src[COL_LATEST_END], errors="coerce")
    src[COL_WORK] = src[COL_WORK].apply(norm_str)

    print(f"[筛选] 源表初始行数 {len(src)}")

    all_task_set = set(src[COL_TASK].unique())
    src = src[src[COL_TASK].apply(lambda t: len(t) == 11 or t[:11] in all_task_set)].copy()
    print(f"[筛选] 规则3后剩余 {len(src)} 行")

    src['order_id'] = src[COL_PLAN].apply(parse_order_id)
    src['part_prefix'] = src[COL_PART].apply(get_part_prefix)
    src['work_num'] = src[COL_WORK].apply(extract_work_number)

    src_valid = src[src['work_num'].notna()].copy()
    if len(src_valid) < len(src):
        print(f"[警告] 有 {len(src)-len(src_valid)} 行无法解析WORK编号，已丢弃")

    op_id_map = {}
    for _, row in src_valid.iterrows():
        op_id_map[(row[COL_TASK], row['work_num'])] = f"op_{row[COL_TASK]}-{row['work_num']}"

    task_work_list = {}
    for _, row in src_valid.iterrows():
        task_work_list.setdefault(row[COL_TASK], []).append(row['work_num'])
    for t in task_work_list:
        task_work_list[t].sort()

    pred_op_map = {}
    for t, wn_list in task_work_list.items():
        for i, wn in enumerate(wn_list):
            pred_op_map[(t, wn)] = "" if i == 0 else op_id_map.get((t, wn_list[i-1]), "")

    ops_rows, init_rows = [], []
    os_pool = build_os_pool()
    zz_pool = build_zz_pool()
    unmatched_procs = set()

    for _, row in src_valid.iterrows():
        t = row[COL_TASK]
        wn = row['work_num']
        op_id = op_id_map[(t, wn)]
        proc = row[COL_PROC]
        ops_rows.append({
            "op_id": op_id,
            "task_id": t,
            "op_name": proc,
            "process_type": match_process_type(proc),
            "processing_time_hrs": minutes_to_hours(row[COL_REMAIN]),
            "predecessor_ops": pred_op_map.get((t, wn), ""),
            "predecessor_tasks": "",
            "eligible_machine_ids": eligible_machines(proc, os_pool, zz_pool, unmatched_procs),
            "turnover_time_hrs": TURNOVER_TIME_HRS,
            "_work_num": int(wn),
        })
        init_rows.append({
            "op_id": op_id,
            "initial_start_time": craft_release_offset(row[COL_CRAFT_TIME]),
            "initial_end_time": clamp_non_negative(to_offset_hours(row[COL_MES_DUE])),
            "initial_status": "PENDING",
            "initial_remaining_processing_time": minutes_to_hours(row[COL_REMAIN]),
            "initial_assigned_machine_id": norm_str(row[COL_STATION]),
        })

    if unmatched_procs:
        print(f"[警告] 以下工序未匹配到可用机台，已兜底填充 OS_1~OS_{OS_MACHINE_COUNT}: {sorted(unmatched_procs)}")

    task_info = {}
    for _, row in src_valid.iterrows():
        t = row[COL_TASK]
        task_info[t] = {'plan': row[COL_PLAN], 'order': row['order_id'],
                        'part_pref': row['part_prefix'], 'len': len(t)}

    task_min_work = {t: min(wn_list) for t, wn_list in task_work_list.items()}
    pred_task_map = {}

    def add_pred_task(target_task, target_work, pred_task):
        pred_task_map.setdefault((target_task, target_work), set()).add(pred_task)

    _special_pattern = re.compile(r'(ED\d+$|TL\d+$|W\d*$|-W)')

    for t, info in task_info.items():
        if info['len'] > 11:
            main_t = t[:11]
            if main_t in task_info:
                add_pred_task(main_t, task_min_work[main_t], t)

    for t, info in task_info.items():
        plan = info['plan']
        if re.search(r'ED\d+$', plan):
            order_pref = parse_order_id(plan)
            part_pref = info['part_pref']
            candidates = [t2 for t2, i2 in task_info.items()
                          if t2 != t and i2['order'] == order_pref and i2['part_pref'] == part_pref
                          and i2['len'] == 11 and not _special_pattern.search(i2['plan'])]
            if candidates:
                edm_rows = src_valid[src_valid[COL_TASK].isin(candidates)]
                edm_rows = edm_rows[edm_rows[COL_PROC].str.contains('EDM', case=False, na=False)]
                if not edm_rows.empty:
                    best = edm_rows.sort_values('work_num').iloc[0]
                    target_task, target_work = best[COL_TASK], best['work_num']
                else:
                    target_task = min(candidates, key=lambda t2: src_valid[src_valid[COL_TASK]==t2]['work_num'].min())
                    target_work = src_valid[src_valid[COL_TASK]==target_task]['work_num'].min()
                add_pred_task(target_task, target_work, t)

    for t, info in task_info.items():
        plan = info['plan']
        if re.search(r'TL\d+$', plan):
            order_pref = parse_order_id(plan)
            part_pref = info['part_pref']
            candidates = [t2 for t2, i2 in task_info.items()
                          if t2 != t and i2['order'] == order_pref and i2['part_pref'] == part_pref
                          and i2['len'] == 11 and not _special_pattern.search(i2['plan'])]
            if candidates:
                target_task = min(candidates, key=lambda t2: src_valid[src_valid[COL_TASK]==t2]['work_num'].min())
                target_work = src_valid[src_valid[COL_TASK]==target_task]['work_num'].min()
                add_pred_task(target_task, target_work, t)

    for t, info in task_info.items():
        plan = info['plan']
        if re.search(r'(W\d*$|-W)', plan):
            order_pref = parse_order_id(plan)
            part_pref = info['part_pref']
            candidates = [t2 for t2, i2 in task_info.items()
                          if t2 != t and i2['order'] == order_pref and i2['part_pref'] == part_pref
                          and i2['len'] == 11 and not _special_pattern.search(i2['plan'])]
            for target in candidates:
                add_pred_task(target, task_min_work[target], t)

    for (target_task, target_work), pred_set in pred_task_map.items():
        pred_str = ';'.join(sorted(pred_set))
        target_op = op_id_map.get((target_task, target_work))
        if target_op:
            for row in ops_rows:
                if row['op_id'] == target_op:
                    row['predecessor_tasks'] = pred_str
                    break

    # ---------- 工序合并 (连续 CNC / LC / LF) ----------
    ops_rows, init_rows = merge_continuous_ops(ops_rows, init_rows, merge_report)
    # ---------- OS 工序前置的周转时间调整 (必须在合并之后) ----------
    ops_rows = apply_os_pred_turnover(ops_rows)
    for r in ops_rows:
        r.pop("_work_num", None)

    order_info = {}
    for _, row in src_valid.iterrows():
        oid = row['order_id']
        if oid not in order_info:
            order_info[oid] = {'craft_time': row[COL_CRAFT_TIME], 'due_time': row[COL_MES_DUE]}
        else:
            if pd.notna(row[COL_CRAFT_TIME]) and (pd.isna(order_info[oid]['craft_time']) or row[COL_CRAFT_TIME] < order_info[oid]['craft_time']):
                order_info[oid]['craft_time'] = row[COL_CRAFT_TIME]
            if pd.notna(row[COL_MES_DUE]) and (pd.isna(order_info[oid]['due_time']) or row[COL_MES_DUE] > order_info[oid]['due_time']):
                order_info[oid]['due_time'] = row[COL_MES_DUE]

    orders_rows = [{
        "order_id": oid, "order_name": oid,
        "release_time": craft_release_offset(t['craft_time']),
        "due_date": clamp_non_negative(to_offset_hours(t['due_time'])),
        "priority": 1,
    } for oid, t in order_info.items()]

    tasks_rows = []
    for t, info in task_info.items():
        plan = info['plan']
        is_main = (info['len'] == 11 and not _special_pattern.search(plan))
        rows = src_valid[src_valid[COL_TASK] == t]
        if rows.empty:
            continue
        rep = rows.sort_values(COL_ORDER_TIME).iloc[0]
        craft_min = rows[COL_CRAFT_TIME].min()
        tasks_rows.append({
            "task_id": t, "order_id": info['order'],
            "task_name": t + "-Assembly-main" if is_main else t,
            "is_main": "Y" if is_main else "N",
            "predecessor_task_ids": "",
            "release_time": craft_release_offset(craft_min),
            "due_date": clamp_non_negative(to_offset_hours(rep[COL_MES_DUE])),
        })

    mt_rows, seen_mt = [], set()
    for _, r in base.iterrows():
        tid = norm_str(r[COL_TYPE_ID])
        if not tid: continue
        row = {"type_id": tid, "type_name": norm_str(r[COL_TYPE_NAME]), "is_critical": norm_str(r[COL_CRIT])}
        key = (row["type_id"], row["type_name"], row["is_critical"])
        if key in seen_mt: continue
        seen_mt.add(key)
        mt_rows.append(row)

    shift_24h, shift_normal = build_shifts(PLAN_START)
    sundays = sum(1 for d in range(SHIFT_DAY_COUNT) if is_sunday(PLAN_START, d))
    print(f"[班次] plan_start={PLAN_START:%Y-%m-%d %H:%M:%S} ({'一二三四五六日'[PLAN_START.weekday()]}), "
          f"{SHIFT_DAY_COUNT} 天内含 {sundays} 个周日(仅 OS 类型置0, 其余类型周日照常上班)")

    mc_rows = []
    for _, r in base.iterrows():
        tid = norm_str(r[COL_TYPE_ID])
        mid = norm_str(r[COL_MACHINE_ID])
        if not mid: continue
        mc_rows.append({
            "type_id": tid, "machine_id": mid,
            "machine_name": norm_str(r[COL_MACHINE_NAME]),
            "shifts": shift_24h if tid == OS_TYPE_ID else shift_normal,
        })

    return {
        "orders": orders_rows, "tasks": tasks_rows,
        "operations": ops_rows, "initial_state": init_rows,
        "machine_types": mt_rows, "machines": mc_rows,
    }

def write_plan_start(wb):
    """将 plan_start_at 写入 planning_context sheet(只改该字段所在列, 其余保留)"""
    if PLANNING_CONTEXT_SHEET not in wb.sheetnames:
        print(f"[警告] 模板中不存在 {PLANNING_CONTEXT_SHEET} sheet, 跳过 plan_start_at 写入")
        return
    ws = wb[PLANNING_CONTEXT_SHEET]
    headers = [norm_str(c.value) for c in ws[1]]
    if PLAN_START_FIELD not in headers:
        print(f"[警告] {PLANNING_CONTEXT_SHEET} 中未找到 {PLAN_START_FIELD} 列, 跳过写入")
        return
    col_idx = headers.index(PLAN_START_FIELD) + 1
    value = PLAN_START.strftime("%Y-%m-%d %H:%M:%S")
    ws.cell(row=2, column=col_idx, value=value)
    print(f"[写入] {PLANNING_CONTEXT_SHEET}.{PLAN_START_FIELD} = {value}")

def write_to_template(template_path, output_path, data):
    wb = load_workbook(template_path)
    write_plan_start(wb)
    for sheet_name, rows in data.items():
        if sheet_name not in wb.sheetnames:
            if sheet_name == "initial_state":
                ws = wb.create_sheet(sheet_name)
                ws.append(["op_id","initial_start_time","initial_end_time",
                           "initial_status","initial_remaining_processing_time",
                           "initial_assigned_machine_id"])
            else:
                sys.exit(f"[错误] 模板中不存在 sheet: {sheet_name}")
        ws = wb[sheet_name]
        headers = [norm_str(c.value) for c in ws[1]]
        if not any(headers):
            sys.exit(f"[错误] sheet {sheet_name} 第一行没有表头")
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        missing = [k for k in rows[0].keys() if k not in headers and not k.startswith("_")] if rows else []
        if missing:
            print(f"[警告] sheet {sheet_name} 表头缺少字段 {missing}, 这些字段将无法写入")
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        print(f"[写入] {sheet_name}: {len(rows)} 行")
    wb.save(output_path)
    print(f"[完成] 已输出: {output_path}")

def main():
    global PLAN_START
    here = Path(__file__).resolve().parent
    ap = argparse.ArgumentParser(description="源表+固定基表 -> instance_template_v2 转换")
    ap.add_argument("--source", default=None,
                    help="源表路径(默认自动选取当前目录下日期最大的 YYYYMMDD.xlsx)")
    ap.add_argument("--base", default=str(here / "预排工位.xlsx"))
    ap.add_argument("--template", default=str(here / "instance_template_v2.xlsx"))
    ap.add_argument("--output", default=str(here / "instance_output.xlsx"))
    ap.add_argument("--report", default=str(here / "merge_report.xlsx"),
                    help="合并工序清单输出路径")
    args = ap.parse_args()

    # 确定源表与 plan_start
    if args.source:
        src_path = Path(args.source).resolve()
        m = re.match(SOURCE_FILE_PATTERN, src_path.name)
        if not m:
            sys.exit(f"[错误] 源表文件名需形如 20260729.xlsx, 实际: {src_path.name}")
        src_date = datetime.strptime(m.group(1), "%Y%m%d")
        print(f"[源表] 指定 {src_path.name}")
    else:
        src_path, src_date = find_latest_source(here)
    PLAN_START = src_date.replace(hour=PLAN_START_HOUR, minute=0, second=0)

    os_pool_len = len(build_os_pool())
    if os_pool_len > 32767:
        print(f"[警告] OS_1~OS_{OS_MACHINE_COUNT} 拼接后长度为 {os_pool_len} 字符, "
              f"超过 Excel 单元格上限 32767, 该单元格会被截断")

    merge_report = []
    data = build_all(str(src_path), args.base, merge_report)
    write_to_template(args.template, args.output, data)
    write_merge_report(merge_report, args.report)

if __name__ == "__main__":
    main()
