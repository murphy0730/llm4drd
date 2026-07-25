import io
import re
import unittest

import openpyxl

from llm4drd.api import server
from llm4drd.core.time_utils import format_display_datetime
from .shop_fixtures import make_graph_context_shop


DISPLAY_DATETIME = re.compile(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$")


def _sheet_rows(payload: bytes, sheet_name: str) -> list[list]:
    wb = openpyxl.load_workbook(io.BytesIO(payload))
    return [list(row) for row in wb[sheet_name].iter_rows(values_only=True)]


def _sheet_names(payload: bytes) -> list[str]:
    return openpyxl.load_workbook(io.BytesIO(payload)).sheetnames


class FormatDisplayDatetimeTest(unittest.TestCase):
    def test_iso_with_offset(self):
        self.assertEqual(format_display_datetime("2026-07-21T14:00:00+08:00"), "2026-07-21 14:00:00")

    def test_iso_with_zulu(self):
        self.assertEqual(format_display_datetime("2026-07-21T14:00:00Z"), "2026-07-21 14:00:00")

    def test_none_passthrough(self):
        self.assertIsNone(format_display_datetime(None))

    def test_non_datetime_passthrough(self):
        self.assertEqual(format_display_datetime("-"), "-")
        self.assertEqual(format_display_datetime(""), "")


class SimExportHeaderTest(unittest.TestCase):
    def _payload(self) -> bytes:
        gantt = [
            {
                "op_id": "OP-11", "op_name": "下料", "task_id": "T-11", "order_id": "O-1",
                "order_name": "Order 1", "machine_id": "M-C1", "machine_name": "Cutter 1",
                "process_type": "cut", "start": 0.0, "end": 4.0, "duration": 4.0,
                "start_at": "2026-07-21T14:00:00+08:00", "end_at": "2026-07-21T18:00:00+08:00",
                "due_at": "2026-07-25T08:00:00+08:00", "priority": 3,
                "is_tardy": True, "due_date": 3.0, "is_main": False,
            },
            # 前端回传的精简条目：无 op_name / machine_name，导出应回退到 ID
            {"op_id": "OP-21", "task_id": "T-21", "machine_id": "M-C2", "start": 1.0, "end": 6.0},
        ]
        return server._build_sim_export_excel({"gantt": gantt, "metrics": {}, "rule": "EDD"}, None)

    def test_schedule_headers(self):
        rows = _sheet_rows(self._payload(), "排程结果")
        self.assertEqual(rows[0], [
            "工序ID", "工序", "前工序ID", "前任务令ID", "任务令", "计划号", "订单名称",
            "计划工位", "机器名称", "工序类型",
            "开始(小时)", "计划开工时间", "结束(小时)", "计划完工时间", "时长(小时)",
            "状态", "优先级", "排产交期", "是否延误", "是否主订单",
        ])

    def test_names_and_datetime_format(self):
        rows = _sheet_rows(self._payload(), "排程结果")
        header, first = rows[0], rows[1]
        self.assertEqual(first[header.index("工序")], "下料")
        self.assertEqual(first[header.index("机器名称")], "Cutter 1")
        for column in ("计划开工时间", "计划完工时间", "排产交期"):
            self.assertRegex(first[header.index(column)], DISPLAY_DATETIME)

    def test_missing_names_fall_back_to_ids(self):
        rows = _sheet_rows(self._payload(), "排程结果")
        header = rows[0]
        simplified = next(row for row in rows[1:] if row[header.index("工序ID")] == "OP-21")
        self.assertEqual(simplified[header.index("工序")], "OP-21")
        self.assertEqual(simplified[header.index("机器名称")], "M-C2")

    def test_predecessors_come_from_instance(self):
        """排程条目不带前置关系，导出时按 op_id 回查实例的 operations 定义。"""
        shop = make_graph_context_shop()
        gantt = [
            {"op_id": "OP-12", "task_id": "T-11", "start": 4.0, "end": 6.0},  # predecessor_ops=["OP-11"]
            {"op_id": "OP-13", "task_id": "T-12", "start": 6.0, "end": 9.0},  # predecessor_tasks=["T-11"]
        ]
        rows = _sheet_rows(server._build_sim_export_excel({"gantt": gantt, "metrics": {}}, shop), "排程结果")
        header = rows[0]
        by_op = {row[header.index("工序ID")]: row for row in rows[1:]}
        self.assertEqual(by_op["OP-12"][header.index("前工序ID")], "OP-11")
        self.assertEqual(by_op["OP-13"][header.index("前任务令ID")], "T-11")
        # 无前置的列留空而不是报错
        self.assertIn(by_op["OP-12"][header.index("前任务令ID")], (None, ""))

    def test_predecessors_blank_without_shop(self):
        rows = _sheet_rows(self._payload(), "排程结果")
        header = rows[0]
        self.assertIn(rows[1][header.index("前工序ID")], (None, ""))

    def test_tardy_sheet_headers(self):
        rows = _sheet_rows(self._payload(), "延误明细")
        self.assertEqual(rows[0], ["计划号", "订单名称", "工序ID", "计划工位", "计划完工时间", "排产交期", "超出(小时)"])
        self.assertRegex(rows[1][4], DISPLAY_DATETIME)
        self.assertRegex(rows[1][5], DISPLAY_DATETIME)


class PredecessorCoverageLogTest(unittest.TestCase):
    """前置列为空时用这条日志区分：op_id 对不上（op_found=False）vs 实例本就没有前置（True）。"""

    def _log_line(self, gantt: list[dict]) -> str:
        shop = make_graph_context_shop()
        with self.assertLogs(level="INFO") as captured:
            server._build_sim_export_excel({"gantt": gantt, "metrics": {}}, shop)
        return next(line for line in captured.output if "sim export" in line)

    def test_unknown_op_id_reports_not_found(self):
        line = self._log_line([{"op_id": "OP-NOT-EXIST", "start": 0.0, "end": 1.0}])
        self.assertIn("predecessor_hits=0", line)
        self.assertIn("op_found=False", line)

    def test_known_op_without_predecessors_reports_found(self):
        line = self._log_line([{"op_id": "OP-11", "start": 0.0, "end": 1.0}])  # 实例里无前置
        self.assertIn("predecessor_hits=0", line)
        self.assertIn("op_found=True", line)

    def test_counts_hits(self):
        line = self._log_line([
            {"op_id": "OP-12", "start": 0.0, "end": 1.0},  # 有前置工序
            {"op_id": "OP-11", "start": 1.0, "end": 2.0},  # 无前置
        ])
        self.assertIn("predecessor_hits=1", line)


class SolutionExportHeaderTest(unittest.TestCase):
    def _payload(self) -> bytes:
        shop = make_graph_context_shop()
        solution = {
            "solution_id": "S-1",
            "schedule": [
                {
                    "op_id": "OP-11", "op_name": "Cut", "task_id": "T-11", "order_id": "O-1",
                    "order_name": "Order 1", "machine_id": "M-C1", "machine_name": "Cutter 1",
                    "tooling_ids": [], "personnel_ids": [], "start": 0.0, "end": 4.0,
                    "start_at": shop.time_label(0.0), "end_at": shop.time_label(4.0),
                    "duration": 4.0, "elapsed_duration": 4.0, "is_main": False,
                    "due_at": shop.time_label(24.0),
                },
                # 前置关系按 op_id 从实例回查：OP-12 依赖工序 OP-11，OP-13 依赖任务 T-11
                {"op_id": "OP-12", "task_id": "T-11", "start": 4.0, "end": 6.0},
                {"op_id": "OP-13", "task_id": "T-12", "start": 6.0, "end": 9.0},
            ],
            "objectives": {"makespan": 4.0},
            "summary": {},
        }
        return server._build_solution_export_bytes(shop, "task-1", {"objective_keys": ["makespan"]}, solution)

    def test_sheet_names(self):
        self.assertEqual(
            _sheet_names(self._payload()),
            ["方案摘要", "排程结果", "机器日历", "规则参数", "状态图例"],
        )

    def test_schedule_headers(self):
        rows = _sheet_rows(self._payload(), "排程结果")
        self.assertEqual(rows[0], [
            "计划号", "订单名称", "任务令", "工序ID", "工序", "前工序ID", "前任务令ID",
            "计划工位", "机器名称",
            "工装ID", "人员ID", "开始(小时)", "结束(小时)", "计划开工时间", "计划完工时间",
            "时长(小时)", "占用时长(小时)", "是否主订单", "排产交期", "状态", "状态说明",
        ])

    def test_datetime_format(self):
        rows = _sheet_rows(self._payload(), "排程结果")
        header, first = rows[0], rows[1]
        for column in ("计划开工时间", "计划完工时间", "排产交期"):
            self.assertRegex(first[header.index(column)], DISPLAY_DATETIME)

    def test_predecessors_come_from_instance(self):
        rows = _sheet_rows(self._payload(), "排程结果")
        header = rows[0]
        by_op = {row[header.index("工序ID")]: row for row in rows[1:]}
        self.assertEqual(by_op["OP-12"][header.index("前工序ID")], "OP-11")
        self.assertEqual(by_op["OP-13"][header.index("前任务令ID")], "T-11")

    def test_calendar_headers(self):
        rows = _sheet_rows(self._payload(), "机器日历")
        self.assertEqual(rows[0], ["计划工位", "机器名称", "窗口类型", "开始(小时)", "结束(小时)", "开始时间", "结束时间"])
        self.assertRegex(rows[1][5], DISPLAY_DATETIME)


class SharedColumnLabelTest(unittest.TestCase):
    def test_two_exports_share_column_labels(self):
        sim_header = _sheet_rows(SimExportHeaderTest()._payload(), "排程结果")[0]
        solution_header = _sheet_rows(SolutionExportHeaderTest()._payload(), "排程结果")[0]
        for key in ("order_id", "task_id", "op_id", "op_name", "machine_id", "machine_name", "start_at", "end_at", "due_at"):
            label = server.SCHEDULE_COLUMN_LABELS[key]
            self.assertIn(label, sim_header)
            self.assertIn(label, solution_header)


if __name__ == "__main__":
    unittest.main()
