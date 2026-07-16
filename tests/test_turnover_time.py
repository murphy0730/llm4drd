"""工序流转等待时间（turnover_time）语义测试。

对应设计：docs/superpowers/specs/2026-07-16-operation-turnover-time-design.md
"""
import io
import sqlite3
import tempfile
import unittest
from pathlib import Path

import openpyxl

from llm4drd.api.server import _validate_instance
from llm4drd.core.models import Machine, MachineType, Operation, OpStatus, Order, Shift, ShopFloor, Task
from llm4drd.core.rules import BUILTIN_RULES
from llm4drd.core.simulator import Simulator
from llm4drd.data.db import InstanceStore, _float_or_default, init_db
from llm4drd.data.template_builder import build_instance_template_bytes
from llm4drd.knowledge.canonical import CanonicalGraphBuilder
from llm4drd.scheduling.online import OnlineSchedulerV3
from llm4drd.tests.shop_fixtures import make_graph_context_shop
from llm4drd.tests.test_simulator_robustness import _build_shop, _full_calendar


class TestTurnoverField(unittest.TestCase):
    def test_turnover_time_defaults_to_zero(self):
        """既有构造点不传 turnover_time 时必须落 0，这是零回归的锚点。"""
        op = Operation(id="OP1", task_id="T1", name="OP1", process_type="turning", processing_time=5.0)
        self.assertEqual(op.turnover_time, 0.0)

    def test_turnover_time_is_settable(self):
        op = Operation(
            id="OP1", task_id="T1", name="OP1", process_type="turning",
            processing_time=5.0, turnover_time=3.5,
        )
        self.assertEqual(op.turnover_time, 3.5)


class TestTemplateColumn(unittest.TestCase):
    def _operations_headers(self) -> list[str]:
        wb = openpyxl.load_workbook(io.BytesIO(build_instance_template_bytes()))
        ws = wb["operations"]
        return [cell.value for cell in ws[1]]

    def test_turnover_column_follows_processing_time(self):
        headers = self._operations_headers()
        self.assertIn("turnover_time_hrs", headers)
        self.assertEqual(
            headers.index("turnover_time_hrs"),
            headers.index("processing_time_hrs") + 1,
            "turnover_time_hrs 必须紧跟在 processing_time_hrs 之后",
        )

    def test_template_demonstrates_nonzero_turnover(self):
        wb = openpyxl.load_workbook(io.BytesIO(build_instance_template_bytes()))
        ws = wb["operations"]
        headers = [cell.value for cell in ws[1]]
        column = headers.index("turnover_time_hrs")
        values = [row[column].value for row in ws.iter_rows(min_row=2)]
        self.assertTrue(any(v for v in values), "模板应至少有一行非零 turnover 以演示语义")


class TestTurnoverPersistence(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.db_path = str(Path(self._tmp.name) / "test.db")
        init_db(self.db_path)

    def tearDown(self):
        self._tmp.cleanup()

    def test_ddl_has_turnover_column(self):
        with sqlite3.connect(self.db_path) as conn:
            columns = {row[1] for row in conn.execute("PRAGMA table_info(inst_operations)")}
        self.assertIn("turnover_time", columns)

    def test_null_turnover_in_legacy_row_loads_as_zero(self):
        """模拟旧库：turnover_time 为 NULL 的行加载必须落 0 而非崩溃。"""
        with sqlite3.connect(self.db_path) as conn:
            conn.execute(
                "INSERT INTO inst_operations (op_id, task_id, op_name, process_type, processing_time, turnover_time) "
                "VALUES ('OP1','T1','OP1','turning',5.0,NULL)"
            )
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            row = dict(conn.execute("SELECT * FROM inst_operations WHERE op_id='OP1'").fetchone())
        self.assertEqual(_float_or_default(row.get("turnover_time"), 0.0), 0.0)

    def test_migration_is_idempotent_on_legacy_table(self):
        """旧库缺列时 init_db 必须补列且可重复执行。"""
        legacy = str(Path(self._tmp.name) / "legacy.db")
        with sqlite3.connect(legacy) as conn:
            conn.execute(
                "CREATE TABLE inst_operations (op_id TEXT PRIMARY KEY, task_id TEXT, "
                "op_name TEXT, process_type TEXT, processing_time REAL)"
            )
        init_db(legacy)
        init_db(legacy)  # 第二次必须不报错
        with sqlite3.connect(legacy) as conn:
            columns = {row[1] for row in conn.execute("PRAGMA table_info(inst_operations)")}
        self.assertIn("turnover_time", columns)

    def test_excel_import_without_turnover_column_defaults_to_zero(self):
        """旧 xlsx 缺列时导入必须落 0。"""
        legacy_row = {"op_id": "OP1", "task_id": "T1", "processing_time_hrs": 5.0}
        self.assertEqual(
            _float_or_default(legacy_row.get("turnover_time_hrs", legacy_row.get("turnover_time", 0)), 0.0),
            0.0,
        )

    def test_excel_import_reads_turnover_column(self):
        row = {"op_id": "OP1", "task_id": "T1", "processing_time_hrs": 5.0, "turnover_time_hrs": 2.5}
        self.assertEqual(
            _float_or_default(row.get("turnover_time_hrs", row.get("turnover_time", 0)), 0.0),
            2.5,
        )

    def test_save_and_load_roundtrip_preserves_turnover_time(self):
        """save_from_shopfloor -> build_shopfloor 必须保值非零 turnover_time。"""
        shop = make_graph_context_shop()
        shop.operations["OP-11"].turnover_time = 2.5
        store = InstanceStore(self.db_path)
        store.save_from_shopfloor(shop)
        loaded = store.build_shopfloor()
        self.assertEqual(loaded.operations["OP-11"].turnover_time, 2.5)


def _shop_with_two_ops(turnover: float) -> ShopFloor:
    """OP1 -> OP2 串行，OP1 的 turnover 可调。OP1 已完工于 t=10。"""
    op1 = Operation(id="OP1", task_id="T1", name="OP1", process_type="turning",
                    processing_time=5.0, turnover_time=turnover)
    op2 = Operation(id="OP2", task_id="T1", name="OP2", process_type="milling",
                    processing_time=3.0, predecessor_ops=["OP1"])
    task = Task(id="T1", order_id="O1", name="T1", due_date=100.0, operations=[op1, op2])
    calendar = [Shift(day=d, start_hour=0.0, hours=24.0) for d in range(30)]
    shop = ShopFloor(
        machine_types={"turning": MachineType(id="turning", name="turning"),
                       "milling": MachineType(id="milling", name="milling")},
        machines={"m1": Machine(id="m1", name="m1", type_id="turning", shifts=calendar),
                  "m2": Machine(id="m2", name="m2", type_id="milling", shifts=calendar)},
        orders={"O1": Order(id="O1", name="O1", due_date=100.0, task_ids=["T1"], main_task_id="T1")},
        tasks={"T1": task},
        operations={"OP1": op1, "OP2": op2},
    )
    shop.build_indexes()
    op1.end_time = 10.0
    return shop


class TestFlowReadyGate(unittest.TestCase):
    def test_gate_adds_predecessor_turnover_to_its_end_time(self):
        shop = _shop_with_two_ops(turnover=4.0)
        self.assertEqual(shop.get_operation_flow_ready_time(shop.operations["OP2"]), 14.0)

    def test_zero_turnover_gate_equals_predecessor_end(self):
        """零回归锚点：turnover=0 时闸门退化为前驱完工时刻。"""
        shop = _shop_with_two_ops(turnover=0.0)
        op2 = shop.operations["OP2"]
        self.assertEqual(
            shop.get_operation_flow_ready_time(op2),
            max(shop.get_operation_release_time(op2), 10.0),
        )

    def test_gate_honors_release_time_when_larger(self):
        shop = _shop_with_two_ops(turnover=1.0)
        shop.tasks["T1"].release_time = 50.0
        self.assertEqual(shop.get_operation_flow_ready_time(shop.operations["OP2"]), 50.0)

    def test_gate_accepts_precomputed_release_time(self):
        """调用方传入缓存的 release_time 时，结果必须与内部计算一致。"""
        shop = _shop_with_two_ops(turnover=4.0)
        op2 = shop.operations["OP2"]
        self.assertEqual(
            shop.get_operation_flow_ready_time(op2, release_time=0.0),
            shop.get_operation_flow_ready_time(op2),
        )

    def test_gate_tolerates_predecessor_without_end_time(self):
        """前驱尚未完工（end_time is None）时不得崩溃。"""
        shop = _shop_with_two_ops(turnover=4.0)
        shop.operations["OP1"].end_time = None
        self.assertEqual(shop.get_operation_flow_ready_time(shop.operations["OP2"]), 0.0)

    def test_gate_uses_max_over_task_predecessor_operations(self):
        """口径 2：任务级前驱取 max(各工序 end + 各自 turnover)。

        刻意让「完工最晚的工序」(OP4, end=12) 与「流转最晚的工序」
        (OP3, end=5 但 turnover=9) 不是同一道——若实现错写成「取任务
        completion_time 再加末工序 turnover」，此测试会得 13 而非 14。
        """
        shop = _shop_with_two_ops(turnover=1.0)
        early = Operation(id="OP3", task_id="T2", name="OP3", process_type="turning",
                          processing_time=2.0, turnover_time=9.0)
        late = Operation(id="OP4", task_id="T2", name="OP4", process_type="turning",
                         processing_time=2.0, turnover_time=1.0)
        early.end_time = 5.0   # 5 + 9 = 14
        late.end_time = 12.0   # 12 + 1 = 13
        shop.tasks["T2"] = Task(id="T2", order_id="O1", name="T2", due_date=100.0,
                                operations=[early, late])
        shop.operations["OP3"] = early
        shop.operations["OP4"] = late
        shop.operations["OP2"].predecessor_tasks = ["T2"]
        # max(OP1: 10+1=11, OP3: 14, OP4: 13) = 14
        self.assertEqual(shop.get_operation_flow_ready_time(shop.operations["OP2"]), 14.0)


def _run(shop):
    result = Simulator(shop, BUILTIN_RULES["FIFO"]).run()
    return {entry["op_id"]: entry for entry in result.schedule}


class TestSimulatorTurnover(unittest.TestCase):
    def test_successor_waits_for_predecessor_turnover(self):
        shop = _build_shop(
            [("OP1", "turning", 5.0, [], 3.0),
             ("OP2", "milling", 2.0, ["OP1"])],
            [("m1", "turning", _full_calendar()), ("m2", "milling", _full_calendar())],
        )
        entries = _run(shop)
        self.assertGreaterEqual(
            entries["OP2"]["start"], entries["OP1"]["end"] + 3.0 - 1e-9,
            "后继必须等满前驱的 turnover",
        )

    def test_zero_turnover_is_unchanged(self):
        """零回归：turnover=0 时后继紧接前驱开工。"""
        shop = _build_shop(
            [("OP1", "turning", 5.0, [], 0.0),
             ("OP2", "milling", 2.0, ["OP1"])],
            [("m1", "turning", _full_calendar()), ("m2", "milling", _full_calendar())],
        )
        entries = _run(shop)
        self.assertAlmostEqual(entries["OP2"]["start"], entries["OP1"]["end"], places=6)

    def test_machine_is_free_during_turnover(self):
        """口径 3：turnover 期间机床可接别的活——这是本字段的核心动机。"""
        shop = _build_shop(
            [("OP1", "turning", 5.0, [], 100.0),   # 超长 turnover
             ("OP2", "milling", 2.0, ["OP1"]),
             ("OP3", "turning", 4.0, [])],          # 无前驱，抢同一台车床
            [("m1", "turning", _full_calendar()), ("m2", "milling", _full_calendar())],
        )
        entries = _run(shop)
        self.assertLess(
            entries["OP3"]["start"], entries["OP1"]["end"] + 100.0,
            "OP3 不该等 OP1 的 turnover——turnover 不占用机床",
        )

    def test_turnover_elapses_on_wall_clock_not_shift_time(self):
        """口径 1：turnover 跨越非排班时段时不被拉长。

        机床每天只排 0-8h。OP1 需 6h，从 0 开工、8h 前完工。turnover=10h
        跨越了当天的非排班时段。若 turnover 错误地只在班次内计时，OP2 会被
        推迟一整天；按自然时间则闸门落在次日，OP2 应在闸门开启后的第一个
        排班窗口内开工。
        """
        short_shifts = [Shift(day=d, start_hour=0.0, hours=8.0) for d in range(30)]
        shop = _build_shop(
            [("OP1", "turning", 6.0, [], 10.0),
             ("OP2", "milling", 2.0, ["OP1"])],
            [("m1", "turning", short_shifts), ("m2", "milling", short_shifts)],
        )
        entries = _run(shop)
        gate = entries["OP1"]["end"] + 10.0
        self.assertGreaterEqual(entries["OP2"]["start"], gate - 1e-9)
        self.assertLess(entries["OP2"]["start"], gate + 24.0,
                        "OP2 应在闸门后的首个排班窗口开工，而非因 turnover 被班次化再推一天")

    def test_completed_predecessor_with_negative_end_time_gates_at_t0(self):
        """spec 用例 7：initial_state 中已完工前驱的 end_time 为负值。

        end_time=-2, turnover=1 -> 闸门 = -1 <= 0，后继在 t=0 即就绪。
        """
        op1 = Operation(
            id="OP1", task_id="T1", name="OP1", process_type="turning",
            processing_time=5.0, turnover_time=1.0,
        )
        op2 = Operation(
            id="OP2", task_id="T1", name="OP2", process_type="milling",
            processing_time=2.0, predecessor_ops=["OP1"],
        )
        from llm4drd.core.models import OpStatus
        op1.status = OpStatus.COMPLETED
        op1.end_time = -2.0
        task = Task(id="T1", order_id="O1", name="T1", due_date=100.0, operations=[op1, op2])
        shop = ShopFloor(
            machine_types={
                "turning": MachineType(id="turning", name="turning"),
                "milling": MachineType(id="milling", name="milling"),
            },
            machines={
                "m1": Machine(id="m1", name="m1", type_id="turning", shifts=_full_calendar()),
                "m2": Machine(id="m2", name="m2", type_id="milling", shifts=_full_calendar()),
            },
            orders={"O1": Order(id="O1", name="O1", due_date=100.0, task_ids=["T1"], main_task_id="T1")},
            tasks={"T1": task},
            operations={"OP1": op1, "OP2": op2},
        )
        shop.build_indexes()
        # predecessor 的贡献是 end_time + turnover_time = -2 + 1 = -1；
        # 与 release_time 默认值 0 取 max 后闸门为 0，即 <= 0，后继 t=0 即就绪。
        self.assertLessEqual(shop.get_operation_flow_ready_time(op2), 0.0)
        entries = _run(shop)
        self.assertLessEqual(entries["OP2"]["start"], 1e-9)


class TestDerivedTimesWithTurnover(unittest.TestCase):
    def test_earliest_start_of_successor_includes_predecessor_turnover(self):
        """任务内前推：后继的 earliest_start 必须含前驱 turnover。"""
        shop = _build_shop(
            [("OP1", "turning", 5.0, [], 4.0),
             ("OP2", "milling", 2.0, ["OP1"])],
            [("m1", "turning", _full_calendar()), ("m2", "milling", _full_calendar())],
        )
        # OP1: est=0, pt=5, turnover=4 -> OP2.est = 0+5+4 = 9
        self.assertAlmostEqual(shop.operations["OP2"].earliest_start_time, 9.0, places=6)

    def test_zero_turnover_leaves_earliest_start_unchanged(self):
        """零回归锚点。"""
        shop = _build_shop(
            [("OP1", "turning", 5.0, [], 0.0),
             ("OP2", "milling", 2.0, ["OP1"])],
            [("m1", "turning", _full_calendar()), ("m2", "milling", _full_calendar())],
        )
        self.assertAlmostEqual(shop.operations["OP2"].earliest_start_time, 5.0, places=6)

    def test_derived_start_time_backs_off_by_own_turnover(self):
        """反推：后继要在 t 开工，则本工序须在 t - turnover 前完工。"""
        shop = _build_shop(
            [("OP1", "turning", 5.0, [], 4.0),
             ("OP2", "milling", 2.0, ["OP1"])],
            [("m1", "turning", _full_calendar()), ("m2", "milling", _full_calendar())],
            due_date=100.0,
        )
        op1, op2 = shop.operations["OP1"], shop.operations["OP2"]
        # OP2.derived_start = 100 - 2 = 98；OP1 须在 98 - 4 = 94 前完工
        # 故 OP1.derived_due_date = 94, OP1.derived_start = 94 - 5 = 89
        self.assertAlmostEqual(op2.derived_start_time, 98.0, places=6)
        self.assertAlmostEqual(op1.derived_due_date, 94.0, places=6)
        self.assertAlmostEqual(op1.derived_start_time, 89.0, places=6)

    def _two_task_shop(self, t1_ops_spec, t2_due=100.0, t2_proc=2.0):
        """T2（含 due）通过 predecessor_task_ids 依赖 T1；T1 的工序由 spec 给定。

        t1_ops_spec: [(op_id, processing_time, turnover_time)]，T1 内无工序级前驱（并行）。
        """
        t1_ops = [
            Operation(id=op_id, task_id="T1", name=op_id, process_type="turning",
                      processing_time=proc, turnover_time=turnover)
            for op_id, proc, turnover in t1_ops_spec
        ]
        task1 = Task(id="T1", order_id="O1", name="T1", operations=t1_ops)
        op_b = Operation(id="OPB", task_id="T2", name="OPB", process_type="turning",
                         processing_time=t2_proc)
        task2 = Task(id="T2", order_id="O1", name="T2", due_date=t2_due,
                     predecessor_task_ids=["T1"], operations=[op_b])
        shop = ShopFloor(
            orders={"O1": Order(id="O1", name="O1", task_ids=["T1", "T2"])},
            tasks={"T1": task1, "T2": task2},
            operations={op.id: op for op in [*t1_ops, op_b]},
        )
        shop.build_indexes()
        return shop

    def test_cross_task_backward_pass_reserves_turnover(self):
        """跨任务反推：前驱任务的工序须为自身 turnover 预留时间。

        OPA (proc=5, turnover=4)；OPB (proc=2, due=100) 经 predecessor_tasks 依赖 T1。
        OPB.derived_start = 98 => OPA 须在 98 - 4 = 94 前完工，最迟 89 开工。
        修复前任务级反推直接传播 98，OPA.derived_start 高估为 93。
        """
        shop = self._two_task_shop([("OPA", 5.0, 4.0)])
        op_a, op_b = shop.operations["OPA"], shop.operations["OPB"]
        self.assertAlmostEqual(op_b.derived_start_time, 98.0, places=6)
        self.assertAlmostEqual(op_a.derived_due_date, 94.0, places=6)
        self.assertAlmostEqual(op_a.derived_start_time, 89.0, places=6)
        # 任务级标量同口径：T1 最迟 89 开工，自身工作最迟 94 完成
        self.assertAlmostEqual(shop.tasks["T1"].derived_start_time, 89.0, places=6)
        self.assertAlmostEqual(shop.tasks["T1"].derived_due_date, 94.0, places=6)

    def test_cross_task_backward_pass_is_per_operation(self):
        """异质 turnover：反推须保留工序级约束，不能给整个任务统一减一个值。

        T1 并行两道：A (proc=10, turnover=0)、B (proc=1, turnover=2)。
        OPB.derived_start = 99 => A 最迟 99 完工；B 最迟 99 - 2 = 97 完工。
        若只用任务级标量（cp_wt=10 与 cp=10 相同），B 会被高估为 99。
        """
        shop = self._two_task_shop([("OPA", 10.0, 0.0), ("OPB1", 1.0, 2.0)], t2_proc=1.0)
        self.assertAlmostEqual(shop.operations["OPB"].derived_start_time, 99.0, places=6)
        self.assertAlmostEqual(shop.operations["OPA"].derived_due_date, 99.0, places=6)
        self.assertAlmostEqual(shop.operations["OPB1"].derived_due_date, 97.0, places=6)

    def _cross_task_op_dep_shop(self):
        """T2.S 经 predecessor_ops 只依赖 T1.A；T1.B 是旁路工序，与 S 无关但 turnover 很大。

        A(proc=1, turnover=0) 与 B(proc=1, turnover=100) 在 T1 内并行，
        S(proc=1, due=100) 在 T2，predecessor_ops=["A"]。
        """
        op_a = Operation(id="A", task_id="T1", name="A", process_type="turning",
                         processing_time=1.0, turnover_time=0.0)
        op_b = Operation(id="B", task_id="T1", name="B", process_type="turning",
                         processing_time=1.0, turnover_time=100.0)
        task1 = Task(id="T1", order_id="O1", name="T1", operations=[op_a, op_b])
        op_s = Operation(id="S", task_id="T2", name="S", process_type="turning",
                         processing_time=1.0, predecessor_ops=["A"])
        task2 = Task(id="T2", order_id="O1", name="T2", due_date=100.0, operations=[op_s])
        shop = ShopFloor(
            orders={"O1": Order(id="O1", name="O1", task_ids=["T1", "T2"])},
            tasks={"T1": task1, "T2": task2},
            operations={"A": op_a, "B": op_b, "S": op_s},
        )
        shop.build_indexes()
        return shop

    def test_cross_task_op_dep_does_not_widen_to_whole_task(self):
        """跨任务 predecessor_ops 只应引入被引用工序的约束，不得取整个前驱任务的流转完工。

        只有 A 约束 S；B 的 turnover=100 与 S 无关。若把跨任务工序依赖折叠成
        任务边再取整个任务的 critical_path_with_turnover，S.earliest_start 会
        被误算为 101（= B 的 1+100）。
        """
        shop = self._cross_task_op_dep_shop()
        self.assertAlmostEqual(shop.operations["A"].earliest_start_time, 0.0, places=6)
        self.assertAlmostEqual(
            shop.operations["S"].earliest_start_time, 1.0, places=6,
            msg="S 只依赖 A（1h 完工、无 turnover），不应受旁路工序 B 的 turnover 影响",
        )

    def test_cross_task_op_dep_backward_spares_unreferenced_sibling(self):
        """反推同理：未被引用的旁路工序 B 不得继承 S 的流转反推界。"""
        shop = self._cross_task_op_dep_shop()
        # S.derived_start = 100 - 1 = 99；A 须在 99 - 0 = 99 前完工
        self.assertAlmostEqual(shop.operations["S"].derived_start_time, 99.0, places=6)
        self.assertAlmostEqual(shop.operations["A"].derived_due_date, 99.0, places=6)
        # B 无任何后继、任务也无外部交期 => 无界。修复前为 99 - 100 = -1。
        self.assertEqual(
            shop.operations["B"].derived_due_date, float("inf"),
            "B 未被 S 引用，不应被扣除其 turnover 而得到负的 derived_due_date",
        )

    def test_cross_task_backward_zero_turnover_unchanged(self):
        """零回归锚点：turnover=0 时跨任务反推与旧口径一致。"""
        shop = self._two_task_shop([("OPA", 5.0, 0.0)])
        op_a = shop.operations["OPA"]
        self.assertAlmostEqual(op_a.derived_due_date, 98.0, places=6)
        self.assertAlmostEqual(op_a.derived_start_time, 93.0, places=6)
        self.assertAlmostEqual(shop.tasks["T1"].derived_start_time, 93.0, places=6)


class TestApproxEvalTurnover(unittest.TestCase):
    def test_simulator_result_respects_turnover_as_reference(self):
        """参照锚点：仿真器的真实结果，供近似评价对齐。"""
        shop = _build_shop(
            [("OP1", "turning", 5.0, [], 4.0),
             ("OP2", "milling", 2.0, ["OP1"])],
            [("m1", "turning", _full_calendar()), ("m2", "milling", _full_calendar())],
        )
        entries = _run(shop)
        self.assertGreaterEqual(entries["OP2"]["start"], entries["OP1"]["end"] + 4.0 - 1e-9)

    def _evaluate(self, shop):
        from llm4drd.optimization.approx_eval import ApproximateScheduleEvaluator
        from llm4drd.optimization.solution_model import CandidateParameters, FEATURE_NAMES
        candidate = CandidateParameters(
            feature_weights={name: 0.0 for name in FEATURE_NAMES},
            destroy_weights={},
            repair_weights={},
        )
        evaluator = ApproximateScheduleEvaluator(
            shop, graph_features={}, time_scale=10.0, due_scale=100.0,
            priority_scale=1.0, keep_schedule_limit=100,
        )
        solution = evaluator.evaluate(candidate, source="test", generation=0)
        return {entry["op_id"]: entry for entry in solution.schedule}

    def test_approx_eval_respects_predecessor_op_turnover(self):
        """base_ready 必须计入工序级前驱的 turnover——直接验证 approx_eval 内部，不止仿真器。"""
        shop = _build_shop(
            [("OP1", "turning", 5.0, [], 4.0),
             ("OP2", "milling", 2.0, ["OP1"])],
            [("m1", "turning", _full_calendar()), ("m2", "milling", _full_calendar())],
        )
        entries = self._evaluate(shop)
        self.assertGreaterEqual(
            entries["OP2"]["start"], entries["OP1"]["end"] + 4.0 - 1e-9,
            "approx_eval 的 base_ready 必须计入前驱 turnover",
        )

    def test_approx_eval_task_level_predecessor_uses_max_over_operations(self):
        """口径 2 的 unroll 校验：不得退化为 task_completion（聚合完工时刻）+ 末工序 turnover。

        OP3 (end=5, turnover=9 -> 14) 与 OP4 (end=12, turnover=1 -> 13) 是同一
        预驱任务 T2 内并行的两道工序。若实现仍用聚合的 task_completion(=12，
        即 OP4 的完工时刻) 再加某一道工序的 turnover，只会得到 13；正确实现
        须遍历任务全部工序并取 max(14, 13) = 14。
        """
        early = Operation(id="OP3", task_id="T2", name="OP3", process_type="turning",
                           processing_time=5.0, turnover_time=9.0)
        late = Operation(id="OP4", task_id="T2", name="OP4", process_type="drilling",
                          processing_time=12.0, turnover_time=1.0)
        task2 = Task(id="T2", order_id="O1", name="T2", due_date=200.0, operations=[early, late])
        op2 = Operation(id="OP2", task_id="T1", name="OP2", process_type="milling",
                         processing_time=2.0, predecessor_tasks=["T2"])
        task1 = Task(id="T1", order_id="O1", name="T1", due_date=200.0, operations=[op2])
        shop = ShopFloor(
            machine_types={
                "turning": MachineType(id="turning", name="turning"),
                "drilling": MachineType(id="drilling", name="drilling"),
                "milling": MachineType(id="milling", name="milling"),
            },
            machines={
                "m1": Machine(id="m1", name="m1", type_id="turning", shifts=_full_calendar()),
                "m2": Machine(id="m2", name="m2", type_id="drilling", shifts=_full_calendar()),
                "m3": Machine(id="m3", name="m3", type_id="milling", shifts=_full_calendar()),
            },
            orders={"O1": Order(id="O1", name="O1", due_date=200.0, task_ids=["T1", "T2"], main_task_id="T1")},
            tasks={"T1": task1, "T2": task2},
            operations={"OP2": op2, "OP3": early, "OP4": late},
        )
        shop.build_indexes()
        entries = self._evaluate(shop)
        self.assertGreaterEqual(
            entries["OP2"]["start"], 14.0 - 1e-6,
            "任务级前驱须取 max(各工序 end+turnover)，不得退化为末工序/聚合完工时刻取值",
        )


class TestExactSolverTurnover(unittest.TestCase):
    def test_exact_respects_turnover(self):
        from llm4drd.optimization.exact import ExactSolver
        shop = _build_shop(
            [("OP1", "turning", 5.0, [], 4.0),
             ("OP2", "milling", 2.0, ["OP1"])],
            [("m1", "turning", _full_calendar()), ("m2", "milling", _full_calendar())],
        )
        result = ExactSolver(shop).solve()
        entries = {e["op_id"]: e for e in result.schedule}
        self.assertGreaterEqual(
            entries["OP2"]["start"], entries["OP1"]["end"] + 4.0 - 1e-6,
            "CP-SAT 必须与仿真器同口径地满足 turnover 约束",
        )

    def test_exact_horizon_covers_long_turnover(self):
        """horizon 估算须计入 turnover：100h 流转、未填交期的合法实例不得被误判 INFEASIBLE。

        仿真器对同一实例给出 OP2 于 101-102h 的可行排程。
        """
        from llm4drd.optimization.exact import ExactSolver
        shop = _build_shop(
            [("OP1", "turning", 1.0, [], 100.0),
             ("OP2", "milling", 1.0, ["OP1"])],
            [("m1", "turning", _full_calendar()), ("m2", "milling", _full_calendar())],
            due_date=float("inf"),
        )
        result = ExactSolver(shop).solve()
        entries = {e["op_id"]: e for e in result.schedule}
        self.assertIn("OP2", entries, f"合法实例被误判为 {result.status}")
        self.assertGreaterEqual(entries["OP2"]["start"], 101.0 - 1e-6)

    def test_exact_clamps_gate_not_end_time_for_negative_history(self):
        """历史完工 end_time=-2、turnover=1：闸门 = max(0, -2+1) = 0。

        仿真器对 end_time + turnover 的整体做非负截断；ExactSolver 若先把
        end_time 截为 0 再加 turnover，会得到 1，与仿真器不一致。
        """
        from llm4drd.optimization.exact import ExactSolver
        shop = _build_shop(
            [("OP1", "turning", 1.0, [], 1.0),
             ("OP2", "milling", 1.0, ["OP1"])],
            [("m1", "turning", _full_calendar()), ("m2", "milling", _full_calendar())],
        )
        op1 = shop.operations["OP1"]
        op1.status = OpStatus.COMPLETED
        op1.start_time = -3.0
        op1.end_time = -2.0
        self.assertLessEqual(shop.get_operation_flow_ready_time(shop.operations["OP2"]), 0.0)
        result = ExactSolver(shop).solve()
        entries = {e["op_id"]: e for e in result.schedule}
        self.assertLessEqual(
            entries["OP2"]["start"], 1e-6,
            "ExactSolver 必须与仿真器一致：闸门整体截断后为 0，OP2 应从 0 开工",
        )


class TestOnlineSchedulerTurnover(unittest.TestCase):
    def test_probe_and_ready_gate_respect_turnover(self):
        """滚动排产的 probe/就绪筛选/wait_time/release_time 四处取值须与仿真器同口径。"""
        shop = _build_shop(
            [("OP1", "turning", 5.0, [], 4.0),
             ("OP2", "milling", 2.0, ["OP1"])],
            [("m1", "turning", _full_calendar()), ("m2", "milling", _full_calendar())],
        )
        scheduler = OnlineSchedulerV3(shop, rule_name="FIFO")
        scheduler.advance(50.0)
        completed = {entry["op_id"]: entry for entry in scheduler.state.completed_ops}
        self.assertGreaterEqual(
            completed["OP2"]["start"], completed["OP1"]["end"] + 4.0 - 1e-9,
            "滚动排产必须与仿真器一致地等满 turnover",
        )

    def test_reschedule_window_preserves_trimmed_predecessor_turnover(self):
        """跨窗口裁剪：OP1 已完工被裁出窗口后，OP2 仍不得早于其 turnover 开工。

        OP1 turnover=20h，在 t=6 时已 COMPLETED（end=5）但 turnover 尚未走完。
        _build_remaining_shop 会把 OP1 从 remaining_shop.operations 中裁掉，
        OP2.predecessor_ops 里对 OP1 的引用也随之被过滤——若该 turnover 约束
        未被折算进窗口内的 release_time，OP2 会在新窗口里被允许提前开工。
        """
        shop = _build_shop(
            [("OP1", "turning", 5.0, [], 20.0),
             ("OP2", "milling", 2.0, ["OP1"])],
            [("m1", "turning", _full_calendar()), ("m2", "milling", _full_calendar())],
        )
        scheduler = OnlineSchedulerV3(shop, rule_name="FIFO")
        scheduler.advance(6.0)
        op1 = scheduler.sim_shop.operations["OP1"]
        self.assertEqual(op1.status, OpStatus.COMPLETED)
        self.assertAlmostEqual(op1.end_time, 5.0, places=6)

        remaining_shop = scheduler._build_remaining_shop()
        op2 = remaining_shop.operations["OP2"]
        # OP1 已被裁出窗口：predecessor_ops 中不再含 OP1。
        self.assertNotIn("OP1", op2.predecessor_ops)
        # 但 turnover 约束（5+20=25，折算到新窗口 25-6=19）必须仍然生效。
        self.assertGreaterEqual(
            remaining_shop.get_operation_flow_ready_time(op2), 19.0 - 1e-9,
            "被裁剪的前驱其 turnover 约束不得随窗口裁剪而丢失",
        )

        result = Simulator(remaining_shop, BUILTIN_RULES["FIFO"]).run()
        entries = {entry["op_id"]: entry for entry in result.schedule}
        self.assertGreaterEqual(
            entries["OP2"]["start"], 19.0 - 1e-9,
            "新窗口内重排时，OP2 不得早于原 turnover 折算后的时刻开工",
        )

    def test_trimmed_turnover_floor_does_not_delay_sibling_ops(self):
        """裁剪折算必须是工序级的：同任务内不依赖该前驱的旁路工序不得被连带推迟。

        OP2 依赖已完工的 OP1（窗口内剩余 turnover 19h）；OP3 与 OP2 同任务但
        无任何前驱，本应在新窗口起点即可排产。若把闸门提升为整个任务的
        release_time，OP3 会被错误推迟 19h。
        """
        shop = _build_shop(
            [("OP1", "turning", 5.0, [], 20.0),
             ("OP2", "milling", 2.0, ["OP1"]),
             ("OP3", "milling", 10.0, [])],
            [("m1", "turning", _full_calendar()), ("m2", "milling", _full_calendar())],
        )
        scheduler = OnlineSchedulerV3(shop, rule_name="FIFO")
        scheduler.advance(6.0)
        self.assertEqual(scheduler.sim_shop.operations["OP1"].status, OpStatus.COMPLETED)

        remaining_shop = scheduler._build_remaining_shop()
        if "OP3" not in remaining_shop.operations:
            self.skipTest("OP3 已在窗口内完工，场景未成立")
        self.assertGreaterEqual(
            remaining_shop.get_operation_flow_ready_time(remaining_shop.operations["OP2"]),
            19.0 - 1e-9,
        )
        self.assertLessEqual(
            remaining_shop.get_operation_flow_ready_time(remaining_shop.operations["OP3"]),
            1e-9,
            "旁路工序 OP3 不依赖被裁剪的前驱，不得继承其 turnover 闸门",
        )


class TestTurnoverValidation(unittest.TestCase):
    """turnover 允许 0、拒绝负值——与 processing_time 必须 > 0 的规则不同。"""

    def test_zero_turnover_is_valid(self):
        op = Operation(id="OP1", task_id="T1", name="OP1", process_type="turning",
                       processing_time=5.0, turnover_time=0.0)
        self.assertEqual(op.turnover_time, 0.0)

    def test_negative_turnover_is_rejected_by_import_validation(self):
        """turnover_time < 0 必须被 _validate_instance 报为 operations sheet 的数据完整性错误。"""
        shop = _build_shop(
            [("OP1", "turning", 5.0, [], -1.0)],
            [("m1", "turning", _full_calendar())],
        )
        result = _validate_instance(shop)
        matches = [
            e for e in result["errors"]
            if e["entity"] == "OP1" and e["sheet"] == "operations" and "流转等待时长非法" in e["message"]
        ]
        self.assertTrue(matches, f"未发现负值 turnover_time 的校验错误，实际 errors={result['errors']}")

    def test_zero_turnover_produces_no_turnover_validation_error(self):
        """零回归锚点：turnover=0 不得触发该校验。"""
        shop = _build_shop(
            [("OP1", "turning", 5.0, [], 0.0)],
            [("m1", "turning", _full_calendar())],
        )
        result = _validate_instance(shop)
        matches = [e for e in result["errors"] if "流转等待时长非法" in e["message"]]
        self.assertEqual(matches, [])

    def test_nonfinite_turnover_is_rejected_by_import_validation(self):
        """nan/inf 不能通过校验：inf 会让 ExactSolver 抛 OverflowError，nan 让闸门失效。"""
        for bad in (float("nan"), float("inf")):
            with self.subTest(turnover=bad):
                shop = _build_shop(
                    [("OP1", "turning", 5.0, [], bad)],
                    [("m1", "turning", _full_calendar())],
                )
                result = _validate_instance(shop)
                matches = [
                    e for e in result["errors"]
                    if e["entity"] == "OP1" and "流转等待时长非法" in e["message"]
                ]
                self.assertTrue(matches, f"turnover={bad} 未被校验拦截")

    def test_update_operation_rejects_invalid_turnover(self):
        """编辑写入边界：负值与非有限值都必须被拒，不能落库。"""
        import asyncio

        from fastapi import HTTPException

        from llm4drd.api.server import update_operation

        for bad in ("-1", "nan", "inf"):
            with self.subTest(turnover=bad):
                with self.assertRaises(HTTPException):
                    asyncio.run(update_operation("OP1", {"turnover_time": bad}))

    def test_update_operation_rejects_malformed_turnover_with_400(self):
        """非数字输入须报 400，而不是 float() 抛 ValueError/TypeError 冒泡成 500。"""
        import asyncio

        from fastapi import HTTPException

        from llm4drd.api.server import update_operation

        for bad in ("abc", {"x": 1}, [1, 2]):
            with self.subTest(turnover=bad):
                with self.assertRaises(HTTPException) as ctx:
                    asyncio.run(update_operation("OP1", {"turnover_time": bad}))
                self.assertEqual(ctx.exception.status_code, 400)


class TestGraphNodeTurnover(unittest.TestCase):
    """OP 节点属性须暴露 turnover_time，与既有 processing_time 并列。"""

    def _op_node_attrs(self, shop: ShopFloor, operation_id: str) -> dict:
        graph = CanonicalGraphBuilder().build(shop)
        node_id = f"OP:{operation_id}"
        for node in graph.nodes:
            if node.node_id == node_id:
                return dict(node.attrs)
        raise AssertionError(f"node {node_id} not found in built graph")

    def test_op_node_exposes_nonzero_turnover_time(self):
        shop = _build_shop(
            [("OP1", "turning", 5.0, [], 4.0)],
            [("m1", "turning", _full_calendar())],
        )
        shop.build_indexes()
        attrs = self._op_node_attrs(shop, "OP1")
        self.assertEqual(attrs["turnover_time"], 4.0)

    def test_op_node_exposes_zero_turnover_time_by_default(self):
        """零回归锚点：未显式设置 turnover_time 的工序节点落 0。"""
        shop = make_graph_context_shop()
        attrs = self._op_node_attrs(shop, "OP-11")
        self.assertEqual(attrs["turnover_time"], 0.0)


if __name__ == "__main__":
    unittest.main()
