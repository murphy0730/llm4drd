#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""按真实单号规则构造中等规模演示实例（命令行入口）。

产出一份符合导入模板格式的 Excel（planning_context / orders / tasks / operations /
machine_types / machines 六个 sheet），并可直接用 `InstanceStore.save_from_csv` 落库
—— 与界面 `POST /api/instance/import-excel` 走的是同一条代码路径。

本实例不含人员与工装：`toolings` / `tooling_types` / `personnel` sheet 全部省略，
工序的 required_tooling_types / required_personnel_skills 留空，选机只靠
`process_type` 与 machine_types.type_id 匹配（core/models.py get_eligible_machines）。

用法：
  # 生成 Excel 并落库（默认改写 llm4drd.db）
  python tools/make_demo_instance.py --xlsx ~/Desktop/演示实例.xlsx
  # 只看数据画像，不写库不写文件
  python tools/make_demo_instance.py --dry-run
"""
from __future__ import annotations

import argparse
import math
import pathlib
import random
import sys
from collections import defaultdict
from datetime import datetime

# 允许 `python tools/make_demo_instance.py` 直接运行：把含 llm4drd 包的目录加入 sys.path。
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))

import openpyxl  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402

from llm4drd.core.models import (  # noqa: E402
    Machine,
    MachineType,
    Operation,
    Order,
    Shift,
    ShopFloor,
    Task,
)
from llm4drd.core.rules import BUILTIN_RULES  # noqa: E402
from llm4drd.core.simulator import Simulator  # noqa: E402
from llm4drd.data.db import (  # noqa: E402
    BaselineSolutionStore,
    DowntimeStore,
    GraphStore,
    InstanceStore,
    init_db,
)
from llm4drd.data.graph_artifact_store import GraphArtifactStore  # noqa: E402


# === 规模常量 ===
PLAN_START_AT = "2026-07-30T08:00:00+08:00"
NUM_ORDERS = 50
NUM_TASKS = 500           # 含主任务
NUM_OPERATIONS = 2000     # 含主任务的装配/检测工序
SUB_TASKS_PER_ORDER = (6, 12)
OPS_PER_SUB_TASK = (2, 7)
PROCESSING_TIME_RANGE = (4.0, 48.0)   # 小时；按关键工种利用率约 50%、总工期约 50 天反推
CALENDAR_DAYS = 90                    # 班次逐天铺开的天数，需覆盖最晚交期
DAY_SHIFT = (8, 10)                   # 白班 08:00 起 10 小时
NIGHT_SHIFT = (20, 8)                 # 夜班 20:00 起 8 小时
RELEASE_HORIZON_H = 720.0             # 订单投放铺满首月（0~30 天）
RELEASE_SKEW = 2.2                    # >1 则前多后少，越往后越稀

# 交期：以参考派工规则的实际完工时刻为基准乘松弛系数，见 _assign_due_dates
DUE_REF_RULE = "ATC"
DUE_SLACK_RANGE = (0.85, 1.20)        # <1 的订单会成为需要抢排的紧单
DUE_SEED_SLACK = 2.5                  # 首轮参考排程用的宽松种子交期
DUE_REFINE_PASSES = 2

# === 工种表：type_id / type_name / is_critical / 台数 ===
MACHINE_TYPES = [
    ("CJ", "车加工", True, 12),
    ("XJ", "铣加工", True, 14),
    ("MJ", "磨加工", False, 8),
    ("ZK", "钻孔", False, 9),
    ("TJ", "镗加工", False, 6),
    ("XQ", "线切割", False, 5),
    ("DHJ", "电火花", False, 4),
    ("RCL", "热处理", True, 6),
    ("BMCL", "表面处理", False, 6),
    ("HJ", "焊接", False, 8),
    ("ZP", "装配", True, 12),
    ("JC", "检测", False, 10),
]
ASSEMBLY_TYPE = "ZP"
INSPECTION_TYPE = "JC"
CRITICAL_LOAD_TILT = 1.25   # 关键工种的接活权重上浮，制造温和瓶颈

# === ID 规则 ===
SUB_TASK_PREFIXES = ("QBY", "QWY", "QXY")
MAIN_TASK_PREFIX = "QZY"
TASK_SERIAL_START = 100001

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
HEADER_FONT = Font(bold=True)


# --------------------------------------------------------------------------- #
# ID 生成
# --------------------------------------------------------------------------- #
def _make_order_ids(rng: random.Random) -> list[str]:
    """一半是首位 5/6 的 6 位纯数字单号，一半是 PM + 6 位数字单号。"""
    half = NUM_ORDERS // 2
    numeric: set[str] = set()
    while len(numeric) < half:
        numeric.add(f"{rng.choice('56')}{rng.randint(0, 99999):05d}")
    pm: set[str] = set()
    while len(pm) < NUM_ORDERS - half:
        pm.add(f"PM{rng.randint(100000, 999999)}")
    order_ids = sorted(numeric) + sorted(pm)
    rng.shuffle(order_ids)
    return order_ids


def _split_total(total: int, buckets: int, low: int, high: int, rng: random.Random) -> list[int]:
    """把 total 拆成 buckets 份，每份尽量落在 [low, high]，总和精确等于 total。"""
    if not low * buckets <= total <= high * buckets:
        raise ValueError(f"无法把 {total} 拆成 {buckets} 份 [{low},{high}]")
    counts = [rng.randint(low, high) for _ in range(buckets)]
    while sum(counts) != total:
        delta = 1 if sum(counts) < total else -1
        index = rng.randrange(buckets)
        if low <= counts[index] + delta <= high:
            counts[index] += delta
    return counts


# --------------------------------------------------------------------------- #
# 车间与工件
# --------------------------------------------------------------------------- #
def _build_machines(shop: ShopFloor) -> None:
    shifts_template = []
    for day in range(CALENDAR_DAYS):
        shifts_template.append(Shift(day=day, start_hour=float(DAY_SHIFT[0]), hours=float(DAY_SHIFT[1])))
        shifts_template.append(Shift(day=day, start_hour=float(NIGHT_SHIFT[0]), hours=float(NIGHT_SHIFT[1])))

    for type_id, type_name, is_critical, count in MACHINE_TYPES:
        shop.machine_types[type_id] = MachineType(id=type_id, name=type_name, is_critical=is_critical)
        for index in range(1, count + 1):
            machine_id = f"{type_id}{index:02d}"
            shop.machines[machine_id] = Machine(
                id=machine_id,
                name=f"{type_name}-{index:02d}",
                type_id=type_id,
                shifts=list(shifts_template),
            )


def _process_type_weights() -> list[tuple[str, float]]:
    """子任务工序的接活权重：按台数加权，关键工种再上浮，形成温和瓶颈。"""
    weights = []
    for type_id, _, is_critical, count in MACHINE_TYPES:
        if type_id in (ASSEMBLY_TYPE, INSPECTION_TYPE):
            continue
        weights.append((type_id, count * (CRITICAL_LOAD_TILT if is_critical else 1.0)))
    return weights


def _build_workpieces(shop: ShopFloor, rng: random.Random) -> None:
    """造订单 / 任务（含主任务）/ 工序，交期先留 inf，由 _assign_due_dates 反推。"""
    order_ids = _make_order_ids(rng)
    sub_task_counts = _split_total(
        NUM_TASKS - NUM_ORDERS, NUM_ORDERS, *SUB_TASKS_PER_ORDER, rng=rng
    )
    sub_op_counts = _split_total(
        NUM_OPERATIONS - 2 * NUM_ORDERS, NUM_TASKS - NUM_ORDERS, *OPS_PER_SUB_TASK, rng=rng
    )

    type_pool = [type_id for type_id, _ in _process_type_weights()]
    type_weights = [weight for _, weight in _process_type_weights()]
    type_names = {type_id: name for type_id, name, _, _ in MACHINE_TYPES}

    task_serial = TASK_SERIAL_START
    op_serial = 1
    sub_task_cursor = 0

    for order_index, order_id in enumerate(order_ids):
        release_time = float(round(RELEASE_HORIZON_H * rng.random() ** RELEASE_SKEW))
        order = Order(
            id=order_id,
            name=f"生产订单-{order_index + 1:03d}",
            release_time=release_time,
            priority=rng.randint(1, 5),
        )

        sub_task_ids: list[str] = []

        for _ in range(sub_task_counts[order_index]):
            task_id = f"{SUB_TASK_PREFIXES[task_serial % len(SUB_TASK_PREFIXES)]}{task_serial}"
            task_serial += 1
            task = Task(
                id=task_id,
                order_id=order_id,
                name=f"零件-{len(sub_task_ids) + 1:02d}",
                release_time=release_time,
            )
            predecessor_op_id = None
            for _ in range(sub_op_counts[sub_task_cursor]):
                process_type = rng.choices(type_pool, weights=type_weights, k=1)[0]
                duration = round(rng.uniform(*PROCESSING_TIME_RANGE), 2)
                op = Operation(
                    id=f"op_{op_serial:06d}",
                    task_id=task_id,
                    name=f"{type_names[process_type]}工序",
                    process_type=process_type,
                    processing_time=duration,
                    predecessor_ops=[predecessor_op_id] if predecessor_op_id else [],
                )
                op_serial += 1
                predecessor_op_id = op.id
                task.operations.append(op)
                shop.operations[op.id] = op
            sub_task_cursor += 1
            shop.tasks[task_id] = task
            order.task_ids.append(task_id)
            sub_task_ids.append(task_id)

        # 主任务：装配（依赖全部子任务）→ 检测
        main_task_id = f"{MAIN_TASK_PREFIX}{task_serial}"
        task_serial += 1
        main_task = Task(
            id=main_task_id,
            order_id=order_id,
            name="总装与总检",
            is_main=True,
            predecessor_task_ids=list(sub_task_ids),
            release_time=release_time,
        )
        assembly_duration = round(rng.uniform(6.0, 30.0), 2)
        assembly_op = Operation(
            id=f"op_{op_serial:06d}",
            task_id=main_task_id,
            name="总装",
            process_type=ASSEMBLY_TYPE,
            processing_time=assembly_duration,
            predecessor_tasks=list(sub_task_ids),
        )
        op_serial += 1
        inspection_duration = round(rng.uniform(2.0, 10.0), 2)
        inspection_op = Operation(
            id=f"op_{op_serial:06d}",
            task_id=main_task_id,
            name="总检",
            process_type=INSPECTION_TYPE,
            processing_time=inspection_duration,
            predecessor_ops=[assembly_op.id],
        )
        op_serial += 1
        for op in (assembly_op, inspection_op):
            main_task.operations.append(op)
            shop.operations[op.id] = op

        shop.tasks[main_task_id] = main_task
        order.task_ids.append(main_task_id)
        order.main_task_id = main_task_id
        shop.orders[order_id] = order


# --------------------------------------------------------------------------- #
# 交期
# --------------------------------------------------------------------------- #
def _reference_completion(shop: ShopFloor) -> dict[str, float]:
    """跑一遍参考派工规则，取每个订单主任务的实际完工时刻。"""
    result = Simulator(shop, BUILTIN_RULES[DUE_REF_RULE]).run()
    end_by_task: dict[str, float] = {}
    for row in result.schedule:
        end_by_task[row["task_id"]] = max(end_by_task.get(row["task_id"], 0.0), row["end"])
    return {
        order.id: end_by_task.get(order.main_task_id, 0.0)
        for order in shop.orders.values()
    }


def _apply_due_dates(shop: ShopFloor, target: dict[str, float], factors: dict[str, float]) -> None:
    """due = 放行 + (目标完工 - 放行) × 松弛系数，并保证严格晚于关键路径下界。

    关键路径下界取自 derive_internal_targets（build_indexes 内部会调用），与
    api/server.py 的 _validate_instance 判定"必然延误"用的是同一个函数，
    因此下面的 max(..., cp + 1) 保证不会产生必然延误警告。
    """
    for order in shop.orders.values():
        critical_path_finish = float(shop.tasks[order.main_task_id].earliest_finish_time)
        due = order.release_time + (target[order.id] - order.release_time) * factors[order.id]
        due = float(math.ceil(max(due, critical_path_finish + 1.0)))
        order.due_date = due
        for task_id in order.task_ids:
            shop.tasks[task_id].due_date = due
    shop.build_indexes()


def _assign_due_dates(shop: ShopFloor, rng: random.Random) -> None:
    """以参考排程的实际完工时刻为基准反推交期。

    不走 data/generator.py 的产能反推：那套按工种产能 + 排队积压估出的交期对本实例
    过于宽松（EDD/ATC 之类的规则一律零延误，KPI 没有区分度）。改用"跑一遍参考规则、
    按实际完工时刻乘松弛系数"，松弛 <1 的订单自然成为需要抢排的紧单。

    交期变化会反过来影响 ATC 的排程，所以先用宽松种子交期起步，再迭代收敛几轮；
    每个订单的松弛系数在迭代中固定不变，保证收敛且结果可由 seed 复现。
    """
    factors = {order_id: rng.uniform(*DUE_SLACK_RANGE) for order_id in shop.orders}
    shop.build_indexes()
    seed_target = {
        order.id: order.release_time
        + (float(shop.tasks[order.main_task_id].earliest_finish_time) - order.release_time) * DUE_SEED_SLACK
        for order in shop.orders.values()
    }
    _apply_due_dates(shop, seed_target, factors)
    for _ in range(DUE_REFINE_PASSES):
        _apply_due_dates(shop, _reference_completion(shop), factors)


def build_shop(seed: int) -> ShopFloor:
    rng = random.Random(seed)
    shop = ShopFloor(plan_start_at=datetime.fromisoformat(PLAN_START_AT))
    _build_machines(shop)
    _build_workpieces(shop, rng)
    _assign_due_dates(shop, rng)
    return shop


# --------------------------------------------------------------------------- #
# Excel（列名与顺序严格照 data/template_builder.py）
# --------------------------------------------------------------------------- #
def _add_sheet(workbook, name: str, headers: list[str], rows: list[list]) -> None:
    ws = workbook.create_sheet(title=name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for index, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = min(max(len(header) + 2, 12), 42)
    ws.freeze_panes = "A2"


def _shifts_cell(machine: Machine) -> str:
    return ";".join(
        f"{shift.day}/{shift.start_hour:g}/{shift.hours:g}" for shift in machine.shifts
    )


def build_workbook(shop: ShopFloor):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _add_sheet(
        wb,
        "planning_context",
        ["template_version", "plan_start_at", "note"],
        [["2026.07.16.1", PLAN_START_AT, "时间字段均为相对 plan_start_at 的偏移小时数"]],
    )
    _add_sheet(
        wb,
        "orders",
        ["order_id", "order_name", "release_time", "due_date", "priority"],
        [
            [order.id, order.name, order.release_time, order.due_date, order.priority]
            for order in shop.orders.values()
        ],
    )
    _add_sheet(
        wb,
        "tasks",
        ["task_id", "order_id", "task_name", "is_main", "predecessor_task_ids", "release_time", "due_date"],
        [
            [
                task.id,
                task.order_id,
                task.name,
                "Y" if task.is_main else "N",
                ";".join(task.predecessor_task_ids),
                task.release_time,
                task.due_date,
            ]
            for task in shop.tasks.values()
        ],
    )
    _add_sheet(
        wb,
        "operations",
        [
            "op_id",
            "task_id",
            "op_name",
            "process_type",
            "processing_time_hrs",
            "turnover_time_hrs",
            "predecessor_ops",
            "predecessor_tasks",
            "eligible_machine_ids",
            "required_tooling_types",
            "required_personnel_skills",
        ],
        [
            [
                op.id,
                op.task_id,
                op.name,
                op.process_type,
                op.processing_time,
                op.turnover_time,
                ";".join(op.predecessor_ops),
                ";".join(op.predecessor_tasks),
                "",
                "",
                "",
            ]
            for op in shop.operations.values()
        ],
    )
    _add_sheet(
        wb,
        "machine_types",
        ["type_id", "type_name", "is_critical"],
        [
            [mt.id, mt.name, "Y" if mt.is_critical else "N"]
            for mt in shop.machine_types.values()
        ],
    )
    _add_sheet(
        wb,
        "machines",
        ["machine_id", "machine_name", "type_id", "shifts"],
        [
            [machine.id, machine.name, machine.type_id, _shifts_cell(machine)]
            for machine in shop.machines.values()
        ],
    )
    return wb


def _sheet_rows(wb, name: str) -> list[dict]:
    """按 api/server.py import_excel 的口径把 sheet 读成字符串字典列表。"""
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h or "").strip() for h in rows[0]]
    return [
        {headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)}
        for row in rows[1:]
        if any(v is not None for v in row)
    ]


def import_workbook(wb, db_path: str) -> ShopFloor:
    """复刻 POST /api/instance/import-excel 的落库序列，不依赖服务进程。"""
    init_db(db_path)
    inst_store = InstanceStore(db_path)
    inst_store.save_from_csv(
        _sheet_rows(wb, "orders"),
        _sheet_rows(wb, "tasks"),
        _sheet_rows(wb, "operations"),
        _sheet_rows(wb, "machine_types"),
        _sheet_rows(wb, "machines"),
        plan_start_at=datetime.fromisoformat(PLAN_START_AT),
    )
    DowntimeStore(db_path).clear_all()
    # 旧实例的图谱产物与基准方案库已失效；workflow_progress / rule_reference_cache
    # 按 inst_version 自动失效，不需要手动清。
    GraphStore(db_path).clear_all()
    GraphArtifactStore(db_path).mark_invalid("instance_regenerated")
    BaselineSolutionStore(db_path).clear_all()
    return inst_store.build_shopfloor()


# --------------------------------------------------------------------------- #
# 画像报告
# --------------------------------------------------------------------------- #
def report(shop: ShopFloor) -> None:
    print(f"plan_start_at = {shop.plan_start_at.isoformat()}")
    print(
        f"订单 {len(shop.orders)} / 任务 {len(shop.tasks)}"
        f"（主任务 {sum(1 for t in shop.tasks.values() if t.is_main)}）"
        f" / 工序 {len(shop.operations)} / 机器 {len(shop.machines)}"
        f" / 工种 {len(shop.machine_types)}"
    )

    def histogram(label: str, values: list[float]) -> None:
        weeks = defaultdict(int)
        for value in values:
            weeks[int(value // 168)] += 1
        print(f"  {label}（按周）:")
        for week in sorted(weeks):
            print(f"    第 {week + 1:>2} 周 [{week * 168:>4.0f}h~{(week + 1) * 168:>4.0f}h] "
                  f"{'█' * weeks[week]} {weeks[week]}")

    histogram("release_time", [o.release_time for o in shop.orders.values()])
    histogram("due_date", [o.due_date for o in shop.orders.values()])
    due_dates = [o.due_date for o in shop.orders.values()]
    print(f"  交期区间: {min(due_dates):.0f}h ~ {max(due_dates):.0f}h"
          f"（{min(due_dates) / 24:.1f} ~ {max(due_dates) / 24:.1f} 天）")

    work_by_type = defaultdict(float)
    for op in shop.operations.values():
        work_by_type[op.process_type] += op.processing_time
    total_work = sum(work_by_type.values())
    horizon_days = max(due_dates) / 24.0
    daily_hours = DAY_SHIFT[1] + NIGHT_SHIFT[1]
    print(f"  总工作量 {total_work:,.0f} 机时；理论利用率（按最晚交期 {horizon_days:.0f} 天为窗口）:")
    for type_id, type_name, is_critical, count in MACHINE_TYPES:
        capacity_hours = count * daily_hours * horizon_days
        util = work_by_type.get(type_id, 0.0) / capacity_hours if capacity_hours else 0.0
        print(f"    {type_id:<5}{type_name:<6}{count:>3}台  工作量 {work_by_type.get(type_id, 0.0):>8,.0f}h"
              f"  利用率 {util:>5.1%}{'  ← 关键' if is_critical else ''}")
    print(f"    合计 100 台平均利用率 {total_work / (100 * daily_hours * horizon_days):.1%}")


def main() -> None:
    parser = argparse.ArgumentParser(description="构造中等规模演示实例（Excel + 落库）")
    parser.add_argument("--seed", type=int, default=20260730)
    parser.add_argument("--db", default="llm4drd.db")
    parser.add_argument("--xlsx", default="", help="Excel 输出路径；留空则不写文件")
    parser.add_argument("--no-import", action="store_true", help="只产出 Excel，不写库")
    parser.add_argument("--dry-run", action="store_true", help="只打印画像，不写文件也不写库")
    args = parser.parse_args()

    shop = build_shop(args.seed)
    report(shop)
    if args.dry_run:
        print("\n[dry-run] 未写入任何文件或数据库")
        return

    wb = build_workbook(shop)
    if args.xlsx:
        path = pathlib.Path(args.xlsx).expanduser()
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        print(f"\nExcel 已写出: {path}")

    if not args.no_import:
        loaded = import_workbook(wb, args.db)
        print(f"已落库: {args.db}")
        print(f"库内实例: {loaded.summary()}")


if __name__ == "__main__":
    main()
